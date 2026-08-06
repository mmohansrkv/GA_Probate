"""
LNGA_ProbateCourts_Combined_Tool.py
====================================================================
Combined GA / LNGA Probate Courts Tool
====================================================================

This file merges the two previously separate tools into one launcher:

  1. CITATION SEARCH  (originally LNGA_ProbateCourts_CitationSearch_4.py)
     Drives a real Chrome browser (Selenium) against
     https://georgiaprobaterecords.com/Traffic/SearchCitations.aspx for
     every County/Start/End row in an Input Excel/txt file, scrapes the
     results grid, and writes an "Output" workbook of every case number
     found plus a "Result" (Match/New/Missing) workbook.

  2. VALIDATION  (originally the GA Probate Court Offline HTML vs Output
     Excel Validation Tool)
     Compares offline HTML case files against an Output Excel and flags
     field-by-field mismatches, plus Input-vs-Output and same-Case#/
     different-County cross-checks.

Running this file with no arguments opens a small launcher window with
two buttons — "Citation Search" and "Validation" — each of which opens
its original GUI in its own window so both tools can be used
independently (and even run at the same time). Both windows now show a
live RUNNING TIME indicator while a job is in progress, and report the
total elapsed time when the job finishes.

The Validation module's headless CLI mode is preserved unchanged:
    python LNGA_ProbateCourts_Combined_Tool.py --cli [--offline PATH] [--output PATH] [--input PATH] [--result-base PATH]

Requirements:
    pip install selenium openpyxl pandas beautifulsoup4
    (Selenium 4.6+ auto-downloads a matching Chrome driver — just need
    Google Chrome installed for the Citation Search module.)
====================================================================
"""

import os
import re
import sys
import glob
import time
import queue
import logging
import threading
import subprocess
import traceback
from datetime import datetime

import pandas as pd

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("\u274c openpyxl is required. Install it with:  pip install openpyxl")
    sys.exit(1)

try:
    from bs4 import BeautifulSoup
except ImportError:
    print("\u274c beautifulsoup4 is required. Install it with:  pip install beautifulsoup4")
    sys.exit(1)

import tkinter as tk
from tkinter import ttk, filedialog, messagebox

# ══════════════════════════════════════════════════════
#  GUI THEME  (shared by both modules)
# ══════════════════════════════════════════════════════
BG_MAIN      = "#F7F8FA"
BG_SIDEBAR   = "#1D3A6A"
BG_LOG       = "#1E2330"
BG_CARD      = "#FFFFFF"
ACCENT       = "#185FA5"
ACCENT_HOVER = "#1D3A6A"
C_GREEN      = "#276221";  C_GREEN_BG  = "#EAF3DE"
C_RED        = "#9C0006";  C_RED_BG    = "#FCEBEB"
C_ORANGE     = "#7F3F00";  C_ORANGE_BG = "#FEF3CD"
C_GRAY       = "#4B5563";  C_GRAY_BG   = "#F3F4F6"
TEXT_PRI     = "#1A1A1A";  TEXT_SEC    = "#6B7280"
BORDER_GUI   = "#E5E7EB"
LOG_OK       = "#D1FAE5";  LOG_WARN = "#FDE68A"
LOG_ERR      = "#FCA5A5";  LOG_INFO = "#93C5FD"

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

# ============================================================
# LOGGING  (used by the Validation module; the Citation Search module
# uses its own in-window log queue/callback and is unaffected by this)
# ============================================================
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-7s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("lnga_probate_tool")

# ================================================================
# MODULE: Validation Tool (Offline HTML vs Output Excel Validation)
# ================================================================
# WHAT THIS MODULE DOES
# ----------------------
# 1. Reads every offline .html/.htm case file in OFFLINE_HTML_DIR and extracts
#    field values using the ASP.NET control-id "tag" patterns you supplied
#    (e.g. an id ENDING IN "_lblCourtName" -> County, "_lblCitationNo_0" ->
#    first citation, "_lblCitationNo_1" -> second citation, etc, up to
#    index 50 - one citation per "row" within the case).
#
# 2. Reads every .xlsx/.xls file in OUTPUT_EXCEL_DIR. Each row in that Excel
#    represents ONE citation for a case (a case with 3 citations = 3 rows,
#    all sharing the same Case #/County/etc). The columns this script reads
#    from that file are the "(Output)" ones, e.g. "Case # (Output)",
#    "Fine (Output)", etc. Any pre-existing "(Tool_Input)", "Match", or
#    "Row Status" columns already in that file are ignored/overwritten -
#    this script regenerates them fresh from the HTML every run.
#
# 3. For each case, HTML citations are matched to Output rows first by
#    Citation Number (when both sides have one), then by position for
#    anything left over. Every field is compared and produces:
#        "<Field> (Tool_Input)"   -> value from HTML
#        "<Field> (Output)"       -> value from Output Excel
#        "<Field> Match"          -> "Match" / "Mismatch"
#    plus a final "Row Status" column ("Match" or "Mismatch (Field1, Field2)").
#
# 4. Reads every .xlsx/.xls file in INPUT_EXCEL_DIR (must contain a Case #
#    and County column) and checks that each Input row's Case #+County
#    was actually produced correctly in the Output Excel -> Match / Mismatch
#    / Not Found in Output.
#
# 5. Writes everything into a new folder:  Result_DDMMYYYY  (today's date)
#    inside RESULT_BASE_DIR, as a single Excel workbook with 3 sheets:
#        - "Validation_Result"        (the Tool_Input/Output/Match/Row Status sheet)
#        - "Case#_Diff_County_Check"  (Case number same but county different check)
#        - "Input_vs_Output"          (Case #/County check from step 4)
#
# If your Output/Input Excel files use different column header wording than
# assumed here, edit the OUTPUT_COLUMN_CANDIDATES / INPUT_COLUMN_CANDIDATES
# lists in the CONFIG section below.
# ================================================================

# ============================================================
# CONFIGURATION -- edit paths / column names here only
# ============================================================

OFFLINE_HTML_DIR = r"D:\Mohan\GA_Probate\New\Offline"
OUTPUT_EXCEL_DIR = r"D:\Mohan\GA_Probate\New\Output"
INPUT_EXCEL_DIR  = r"D:\Mohan\GA_Probate\New\Input Folder"
RESULT_BASE_DIR  = r"D:\Mohan\GA_Probate\New"

MAX_ROW_INDEX = 50   # HTML citation rows are indexed _0 through _50 inclusive

# Field order -- this drives the exact column order in the result sheet,
# matching your Output Excel's header pattern.
FIELD_ORDER = [
    "Case #", "County", "Violation Date", "Court Date", "Name", "Address",
    "Race", "Sex", "Citation Number", "Code", "Description", "Fine",
    "Disposition", "Disposition Date", "Sentence",
]

# --- HTML tag (control id SUFFIX) map for single (non-repeating, case-level) fields ---
SINGLE_FIELD_SUFFIXES = {
    "County":          "_lblCourtName",
    "Case #":          "_lblCaseNo",
    "Violation Date":  "lblViolationDate",
    "Court Date":      "_lblCourtDate",
    "Name":            "_lblOffender",
    "Address":         "_lblCityStateZip",
    "Race":            "_lblRace",
    "Sex":             "_lblSex",
}

# --- HTML tag (control id PREFIX, index appended) map for repeating (per-citation) fields ---
ROW_FIELD_PREFIXES = {
    "Citation Number": "_lblCitationNo_",
    "Code":            "_lblCitCode_",
    "Description":     "lblCitCodeDesc_",
    "Fine":            "lblFineAmt_",
    "Disposition":     "_lblDispDesc_",
    "Disposition Date":"_lblDispDate_",
    "Sentence":        "_lblSentencing_",
}

SINGLE_FIELDS = set(SINGLE_FIELD_SUFFIXES.keys())
ROW_FIELDS = set(ROW_FIELD_PREFIXES.keys())

# --- Column header candidates to look for in the OUTPUT excel, for each field. ---
# The script first tries "<field> (Output)" (matching your exact header pattern);
# if that's not found it falls back to these synonyms as a bare column name.
OUTPUT_COLUMN_CANDIDATES = {
    "Case #":            ["case #", "case no", "case number"],
    "County":            ["county", "court name"],
    "Violation Date":    ["violation date"],
    "Court Date":        ["court date"],
    "Name":              ["name", "offender", "offender name"],
    "Address":           ["address", "city state zip"],
    "Race":              ["race"],
    "Sex":               ["sex", "gender"],
    "Citation Number":   ["citation number", "citation no"],
    "Code":              ["code", "cit code"],
    "Description":       ["description", "cit code desc"],
    "Fine":              ["fine", "fine amt", "fine amount"],
    "Disposition":       ["disposition", "disp desc"],
    "Disposition Date":  ["disposition date", "disp date"],
    "Sentence":          ["sentence", "sentencing"],
}

# --- Column header candidates to look for in the INPUT excel ---
INPUT_COLUMN_CANDIDATES = {
    "Case #": ["case #", "case no", "case number"],
    "County": ["county", "court name"],
}

# NOTE: GUI theme constants, SCRIPT_DIR, and the shared `log` logger are
# defined once in the SHARED HEADER section near the top of this combined
# file, and reused by both modules.

# ============================================================
# HELPERS
# ============================================================

def normalize_key(value):
    """Normalize a value for use as a join key (case #, county names etc)."""
    if value is None:
        return ""
    s = str(value).strip()
    if s.lower() in ("nan", "none", "nat"):
        return ""
    s = re.sub(r"\s+", " ", s).upper()
    return s


def normalize_header(value):
    """Normalize a column header for fuzzy matching (lowercase, strip punctuation/spaces)."""
    s = str(value).strip().lower()
    s = re.sub(r"[^a-z0-9]+", "", s)
    return s


NO_RECORDS_MARKERS = (
    "no records to display",
    "no records found",
    "no record found",
    "no results found",
    "no data to display",
    "no matching records found",
)


def is_no_records_placeholder(value):
    """True if value is one of the site's 'zero results for this county'
    placeholder strings (e.g. 'No records to display.') rather than an
    actual Case #. These show up as a single placeholder row per empty
    county in the Input Excel and must NOT be treated as a real case
    that is 'New'/'Not Found' in the Output."""
    s = normalize_value_for_compare(value).lower().rstrip(".")
    return s in NO_RECORDS_MARKERS


def normalize_value_for_compare(value):
    """Normalize a field value before comparing HTML vs Excel (trim, collapse spaces, blank-safe)."""
    if value is None:
        return ""
    s = str(value).strip()
    if s.lower() in ("nan", "none", "nat"):
        return ""
    s = re.sub(r"\s+", " ", s)
    return s


def find_output_column(df_columns, field):
    """
    Find the Output Excel column for a given field. Tries, in order:
        1. exact "<field> (Output)"  (your standard header pattern)
        2. "<field> (Output)" using each synonym from OUTPUT_COLUMN_CANDIDATES
        3. bare synonym column names (no "(Output)" suffix), as a fallback
    Returns the actual column name, or None if nothing matches.
    """
    norm_map = {normalize_header(c): c for c in df_columns}

    exact_key = normalize_header(f"{field} (Output)")
    if exact_key in norm_map:
        return norm_map[exact_key]

    for cand in OUTPUT_COLUMN_CANDIDATES.get(field, []):
        key = normalize_header(f"{cand} (Output)")
        if key in norm_map:
            return norm_map[key]

    for cand in OUTPUT_COLUMN_CANDIDATES.get(field, [field]):
        key = normalize_header(cand)
        if key in norm_map:
            return norm_map[key]

    return None


def find_column(df_columns, candidates):
    """Generic fuzzy header match (used for the Input Excel, which has no (Output) suffix)."""
    norm_map = {normalize_header(c): c for c in df_columns}
    for cand in candidates:
        norm_cand = normalize_header(cand)
        if norm_cand in norm_map:
            return norm_map[norm_cand]
    return None


# ============================================================
# STEP 1: EXTRACT DATA FROM OFFLINE HTML FILES
# ============================================================

def get_text_by_id_suffix(soup, suffix):
    """Find the first element whose id ATTRIBUTE ENDS WITH suffix, return its text."""
    pattern = re.compile(re.escape(suffix) + r"$")
    el = soup.find(id=pattern)
    if el is None:
        return ""
    return el.get_text(separator=" ", strip=True)


COUNTY_SUFFIX_PATTERN = re.compile(r"\s*County\s+Probate\s+Court\s*$", re.IGNORECASE)


def clean_county_value(value):
    """Strip a trailing 'County Probate Court' from a County value, e.g.
    'Fulton County Probate Court' -> 'Fulton'."""
    if not value:
        return value
    return COUNTY_SUFFIX_PATTERN.sub("", value).strip()


# ------------------------------------------------------------------
# NAME / OFFENDER COLUMN CLEANUP  +  GLOBAL "?" / "€" STRIPPER
#
#   1. clean_name_value() -- for the Name/Offender field specifically:
#        - normalizes accented Unicode letters to plain ASCII
#          (e.g. "Á" -> "A", "é" -> "e", "Ñ" -> "N")
#        - strips "do not use" style flags that sometimes get appended
#          to a name in the source data (e.g. "DON'T USE*BURGESS",
#          "** DO NOT USE**BRUMLEY", "::DO NOT USE::")
#   2. strip_junk_chars() -- applied to EVERY column of extracted data:
#        removes any literal "?" and "€" characters outright.
# ------------------------------------------------------------------

import unicodedata

# Any of these appearing anywhere in the name (case-insensitive) marks it
# as a "do not use" flagged record. Written as regex fragments so minor
# spacing/punctuation variants still match.
_DO_NOT_USE_PATTERNS = [
    r"DON'?T\s*USE\s*\*",
    r"\*{1,2}\s*DO\s*NOT\s*USE\s*\*{1,2}",
    r"::\s*DO\s*NOT\s*USE\s*::",
]
_DO_NOT_USE_RE = re.compile("|".join(_DO_NOT_USE_PATTERNS), re.IGNORECASE)

# Characters to strip from every column, everywhere.
_JUNK_CHARS_RE = re.compile(r"[?€]")


def strip_junk_chars(value):
    """Cleans up any extracted column value:
      - normalizes accented Unicode letters (diacritics) to plain ASCII
        (e.g. "Á" -> "A", "é" -> "e", "Ñ" -> "N")
      - removes any literal '?' and '€' characters
    Safe to call on non-strings (returns them unchanged) or empty values."""
    if value is None:
        return value
    if not isinstance(value, str):
        return value
    s = unicodedata.normalize("NFKD", value)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = _JUNK_CHARS_RE.sub("", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def clean_name_value(name):
    """Clean up a Name/Offender field:
      - normalizes accented characters to plain ASCII, strips '?' / '€'
        (via strip_junk_chars, same as every other column)
      - strips 'do not use' style flags
    Safe to call on an empty string."""
    if not name:
        return name
    s = str(name).strip()
    if not s:
        return s

    # Strip "do not use" flags first (before accents are stripped, in
    # case a flag ever contains an accented character).
    s = _DO_NOT_USE_RE.sub("", s)

    # Diacritics + '?' / '€' -- same cleanup as every other column.
    s = strip_junk_chars(s)

    # Clean up stray punctuation/whitespace left behind by the strip.
    s = s.strip(" *:-")
    s = re.sub(r"\s+", " ", s).strip()

    return s


def parse_filename_county_case(filepath):
    """The offline HTML file is saved as '<COUNTY>_<CASE#>.html'
    (e.g. 'BUTTS_25T1999.html'). Some downloads have a numeric upload-id
    prefix in front of that (e.g. '1785921572890_BUTTS_25T1999.html') -
    that leading '<digits>_' is stripped before splitting.
    Returns (county, case_no) as upper-case strings, ('','') if the name
    doesn't follow the pattern."""
    stem = os.path.splitext(os.path.basename(filepath))[0]
    stem = re.sub(r"^\d+_", "", stem)  # strip leading numeric upload-id prefix, if any
    parts = stem.split("_", 1)
    if len(parts) == 2 and parts[0] and parts[1]:
        return parts[0].strip().upper(), parts[1].strip().upper()
    return "", ""


def extract_case_from_html(filepath):
    """Parse one offline HTML case file and return a dict of extracted data."""
    with open(filepath, "r", encoding="utf-8", errors="ignore") as f:
        html = f.read()
    soup = BeautifulSoup(html, "html.parser")

    case = {"_source_file": os.path.basename(filepath)}

    for field, suffix in SINGLE_FIELD_SUFFIXES.items():
        case[field] = strip_junk_chars(get_text_by_id_suffix(soup, suffix))

    # "Fulton County Probate Court" -> "Fulton"
    case["County"] = clean_county_value(case.get("County", ""))

    # Fix mangled accents / strip "DO NOT USE" flags on the offender name.
    case["Name"] = clean_name_value(case.get("Name", ""))

    # File name is expected to be '<COUNTY>_<CASE#>.html' - used as a
    # cross-check against what was actually extracted from inside the page,
    # and as a fallback for County/Case # if the page didn't yield one.
    fname_county, fname_case = parse_filename_county_case(filepath)
    case["_filename_county"] = fname_county
    case["_filename_case"] = fname_case
    if not case.get("County"):
        case["County"] = fname_county
    if not case.get("Case #"):
        case["Case #"] = fname_case

    citations = []
    for i in range(MAX_ROW_INDEX + 1):
        row = {}
        has_value = False
        for field, prefix in ROW_FIELD_PREFIXES.items():
            val = strip_junk_chars(get_text_by_id_suffix(soup, f"{prefix}{i}"))
            row[field] = val
            if val:
                has_value = True
        row["_index"] = i
        if has_value:
            citations.append(row)

    case["Citations"] = citations
    return case


def load_all_html_cases(html_dir):
    """Load and extract every single-case .html/.htm file in html_dir.
    Search-results LIST pages (e.g. 'BUTTS_6-9-2026_6-15-2026_List_3.html'
    or 'BUTTS_LIST_3.html') are skipped -- they hold a whole page of
    citation rows, not one case, and would otherwise get misparsed as a
    case file (the date/list-number portion getting read as a bogus
    Case #), producing false 'Mismatch (Case #, County)' rows."""
    files = sorted(
        glob.glob(os.path.join(html_dir, "*.html"))
        + glob.glob(os.path.join(html_dir, "*.htm"))
    )
    if not files:
        log.warning("No .html/.htm files found in %s", html_dir)
    cases = []
    skipped_list_files = 0
    for fp in files:
        if is_offline_list_filename(fp):
            skipped_list_files += 1
            continue
        try:
            cases.append(extract_case_from_html(fp))
        except Exception as e:
            log.error("Failed to parse %s: %s", fp, e)
    if skipped_list_files:
        log.info("Skipped %d search-results LIST page(s) (not single-case files) in %s",
                  skipped_list_files, html_dir)
    log.info("Extracted %d case(s) from %d HTML file(s) in %s", len(cases), len(files), html_dir)
    return cases


# ============================================================
# STEP 2: LOAD OUTPUT / INPUT EXCEL (all files, combined)
# ============================================================

def get_output_excel_name_tag(excel_dir):
    """Build a filename tag from the Output Excel file name(s) in excel_dir,
    e.g. 'Fulton_Output' -> 'Fulton_Output'. If multiple output excel files
    exist, their names are joined with '_'. Falls back to 'Output' if none found."""
    files = sorted(
        glob.glob(os.path.join(excel_dir, "*.xlsx"))
        + glob.glob(os.path.join(excel_dir, "*.xls"))
    )
    files = [f for f in files if not os.path.basename(f).startswith("~$")]
    if not files:
        return "Output"
    names = [os.path.splitext(os.path.basename(f))[0] for f in files]
    tag = "_".join(names)
    tag = re.sub(r"[^A-Za-z0-9_\-]+", "_", tag).strip("_")
    return tag or "Output"


def load_excel_folder(excel_dir):
    """Read every .xlsx/.xls file in excel_dir into one combined DataFrame (all sheets, all files).
    Skips Excel's hidden lock/temp files (e.g. '~$Book1.xlsx'), which are created
    while a workbook is open elsewhere and are not readable spreadsheets."""
    files = sorted(
        glob.glob(os.path.join(excel_dir, "*.xlsx"))
        + glob.glob(os.path.join(excel_dir, "*.xls"))
    )
    files = [f for f in files if not os.path.basename(f).startswith("~$")]
    if not files:
        log.warning("No .xlsx/.xls files found in %s", excel_dir)
        return pd.DataFrame()

    frames = []
    for fp in files:
        try:
            xls = pd.ExcelFile(fp)
            for sheet in xls.sheet_names:
                df = xls.parse(sheet, dtype=str)
                df["_source_file"] = os.path.basename(fp)
                df["_source_sheet"] = sheet
                frames.append(df)
        except Exception as e:
            log.error("Failed to read %s: %s", fp, e)

    if not frames:
        return pd.DataFrame()

    combined = pd.concat(frames, ignore_index=True, sort=False)
    log.info("Loaded %d row(s) from %d excel file(s) in %s", len(combined), len(files), excel_dir)
    return combined


# ============================================================
# STEP 3: FIELD-LEVEL VALIDATION -- HTML (Tool_Input) vs OUTPUT EXCEL
# ============================================================

def resolve_output_columns(output_df):
    """Resolve the actual Output Excel column name for every field, once."""
    resolved = {}
    for field in FIELD_ORDER:
        resolved[field] = find_output_column(output_df.columns, field)
    return resolved


def case_composite_key(case_value, county_value):
    """Build the composite grouping/lookup key used everywhere a case is
    matched between HTML and Output/Input Excel: Case # + County. Using
    Case # alone would wrongly merge two different cases that happen to
    share the same Case # in two different counties (this DOES happen -
    e.g. Case # 25T1999 exists in both Butts County and Harris County)."""
    return f"{normalize_key(case_value)}||{normalize_key(county_value)}"


def group_output_rows_by_case(output_df, case_col, county_col=None):
    """Group Output Excel rows by (Case #, County) composite key, preserving
    row order. Also returns a secondary case-# only index (case_key ->
    list of composite keys seen) so callers can detect/fall back when a
    Case # exists under more than one County."""
    groups = {}
    case_only_index = {}
    for _, row in output_df.iterrows():
        case_key = normalize_key(row.get(case_col, ""))
        if not case_key:
            continue
        county_val = row.get(county_col, "") if county_col else ""
        composite = case_composite_key(case_key, county_val)
        groups.setdefault(composite, []).append(row)
        case_only_index.setdefault(case_key, set()).add(composite)
    return groups, case_only_index


def pair_citations(html_citations, output_rows):
    """
    Pair each HTML citation with an Output Excel row (one row = one citation).
    1) Exact match by normalized Citation Number where both sides have one.
    2) Leftover citations/rows are paired by original order (position).
    3) Any unmatched leftovers on either side are paired with None.
    Returns a list of (html_citation_or_None, output_row_or_None) tuples.
    """
    html_remaining = list(html_citations)
    output_remaining = list(output_rows)

    # exact Citation Number match first
    pairs = []
    used_output_idx = set()
    still_unmatched_html = []
    for cit in html_remaining:
        cit_no = normalize_key(cit.get("Citation Number", ""))
        matched_idx = None
        if cit_no:
            for i, out_row in enumerate(output_remaining):
                if i in used_output_idx:
                    continue
                out_cit_no = normalize_key(out_row.get("_Citation Number_resolved", ""))
                if out_cit_no and out_cit_no == cit_no:
                    matched_idx = i
                    break
        if matched_idx is not None:
            pairs.append((cit, output_remaining[matched_idx]))
            used_output_idx.add(matched_idx)
        else:
            still_unmatched_html.append(cit)

    unused_output = [r for i, r in enumerate(output_remaining) if i not in used_output_idx]

    # positional pairing for the rest
    for i in range(max(len(still_unmatched_html), len(unused_output))):
        h = still_unmatched_html[i] if i < len(still_unmatched_html) else None
        o = unused_output[i] if i < len(unused_output) else None
        pairs.append((h, o))

    if not pairs:
        # no citations on either side -- still emit one blank citation slot
        pairs = [(None, None)]

    return pairs


def field_match_status(tool_val, output_val):
    t = normalize_value_for_compare(tool_val)
    o = normalize_value_for_compare(output_val)
    if t == "" and o == "":
        return "Match"  # nothing to compare -- treated as match
    return "Match" if t.upper() == o.upper() else "Mismatch"


def build_validation_rows(html_cases, output_df):
    """
    Build the main result rows: one row per (case, citation) pair, with
    "<Field> (Tool_Input)", "<Field> (Output)", "<Field> Match" for every
    field in FIELD_ORDER, plus "Row Status".
    """
    if output_df.empty:
        log.warning("Output Excel is empty/missing - skipping field validation.")
        return [], []

    output_cols = resolve_output_columns(output_df)
    case_col = output_cols.get("Case #")
    if case_col is None:
        raise ValueError(
            "Could not find a 'Case # (Output)' (or fallback) column in the Output Excel. "
            "Check OUTPUT_COLUMN_CANDIDATES['Case #']."
        )

    # pre-resolve the citation number column so pair_citations can use it
    cit_no_col = output_cols.get("Citation Number")
    county_col = output_cols.get("County")

    output_groups, case_only_index = group_output_rows_by_case(output_df, case_col, county_col)
    # attach a normalized helper key for citation-number matching
    for key, rows in output_groups.items():
        for row in rows:
            row["_Citation Number_resolved"] = row.get(cit_no_col, "") if cit_no_col else ""

    result_rows = []
    case_summary = []

    for html_case in html_cases:
        case_only_key = normalize_key(html_case.get("Case #", ""))
        county_key = normalize_key(html_case.get("County", ""))
        composite_key = case_composite_key(case_only_key, county_key)

        output_rows_for_case = output_groups.get(composite_key, [])
        county_note = ""

        if not output_rows_for_case:
            # No exact (Case #, County) match. If this Case # exists in the
            # Output under a DIFFERENT county, that is the "same case #,
            # different county" situation - do NOT silently borrow those
            # rows (that's the bug this fix addresses). Just flag it.
            other_composites = case_only_index.get(case_only_key, set())
            other_counties = sorted({
                c.split("||", 1)[1] for c in other_composites if c != composite_key
            })
            if other_counties:
                county_note = (
                    f"Case # {html_case.get('Case #','')} exists in Output under different "
                    f"County/Counties: {', '.join(other_counties)} (HTML County: "
                    f"{html_case.get('County','') or '(blank)'})"
                )
                log.warning(county_note)

        found_in_output = bool(output_rows_for_case)

        pairs = pair_citations(html_case.get("Citations", []), output_rows_for_case)

        row_mismatch_count = 0
        for html_cit, out_row in pairs:
            row_data = {}
            mismatched_fields = []

            for field in FIELD_ORDER:
                if field in SINGLE_FIELDS:
                    tool_val = html_case.get(field, "")
                else:
                    tool_val = html_cit.get(field, "") if html_cit else ""

                output_val = out_row.get(output_cols.get(field), "") if (out_row is not None and output_cols.get(field)) else ""

                status = field_match_status(tool_val, output_val)
                if status == "Mismatch":
                    mismatched_fields.append(field)

                row_data[f"{field} (Tool_Input)"] = normalize_value_for_compare(tool_val)
                row_data[f"{field} (Output)"] = normalize_value_for_compare(output_val)
                row_data[f"{field} Match"] = status

            if mismatched_fields:
                row_data["Row Status"] = "Mismatch (" + ", ".join(mismatched_fields) + ")"
                row_mismatch_count += 1
            else:
                row_data["Row Status"] = "Match"

            row_data["_source_html_file"] = html_case.get("_source_file", "")
            result_rows.append(row_data)

        case_summary.append({
            "Case #": html_case.get("Case #", ""),
            "County": html_case.get("County", ""),
            "Rows Compared": len(pairs),
            "Rows Matched": len(pairs) - row_mismatch_count if found_in_output else 0,
            "Rows Mismatched": row_mismatch_count if found_in_output else len(pairs),
            "Overall Status": "Match" if (found_in_output and row_mismatch_count == 0) else
                              ("Case Not Found in Output Excel" if not found_in_output else "Mismatch"),
            "Source HTML File": html_case.get("_source_file", ""),
            "Case# Same # Diff County Note": county_note,
        })

    return result_rows, case_summary


def build_case_county_check(html_cases, output_df):
    """
    Cross-check sheet: for every distinct Case # seen (in the offline HTML
    files and/or the Output Excel), list every County it shows up under on
    each side. Flags any Case # that appears under MORE THAN ONE County -
    this is the 'Case number same but county different' check. It is not
    automatically wrong (the same Case # legitimately recurs across
    different counties' probate courts), but it needs a human glance to
    confirm it's not a data-entry mix-up.
    """
    case_data = {}  # case_key -> {"html": set(), "output": set(), "html_files": set()}

    for c in html_cases:
        case_key = normalize_key(c.get("Case #", ""))
        if not case_key:
            continue
        entry = case_data.setdefault(case_key, {"html": set(), "output": set(), "html_files": set()})
        county = normalize_value_for_compare(c.get("County", "")).upper()
        if county:
            entry["html"].add(county)
        if c.get("_source_file"):
            entry["html_files"].add(c["_source_file"])

    if output_df is not None and not output_df.empty:
        output_cols = resolve_output_columns(output_df)
        case_col = output_cols.get("Case #")
        county_col = output_cols.get("County")
        if case_col:
            for _, row in output_df.iterrows():
                case_key = normalize_key(row.get(case_col, ""))
                if not case_key:
                    continue
                entry = case_data.setdefault(case_key, {"html": set(), "output": set(), "html_files": set()})
                county = normalize_value_for_compare(row.get(county_col, "")).upper() if county_col else ""
                if county:
                    entry["output"].add(county)

    rows = []
    for case_key in sorted(case_data.keys()):
        entry = case_data[case_key]
        html_counties = sorted(entry["html"])
        output_counties = sorted(entry["output"])
        flags = []
        if len(html_counties) > 1:
            flags.append("Same Case # under multiple Counties in OFFLINE HTML")
        if len(output_counties) > 1:
            flags.append("Same Case # under multiple Counties in OUTPUT EXCEL")
        if html_counties and output_counties and set(html_counties) != set(output_counties):
            flags.append("HTML County set differs from Output County set")
        status = "; ".join(flags) if flags else "OK"
        rows.append({
            "Case #": case_key,
            "Counties in Offline HTML": ", ".join(html_counties),
            "Counties in Output Excel": ", ".join(output_counties),
            "Source HTML File(s)": ", ".join(sorted(entry["html_files"])),
            "Status": status,
        })

    columns = ["Case #", "Counties in Offline HTML", "Counties in Output Excel",
               "Source HTML File(s)", "Status"]
    df = pd.DataFrame(rows)
    if df.empty:
        df = pd.DataFrame(columns=columns)
    else:
        df = df[columns]
    return df


# ============================================================
# STEP 4: INPUT EXCEL vs OUTPUT EXCEL -- Case # + County VALIDATION
# ============================================================

def run_input_vs_output_validation(input_df, output_df):
    """Returns (comparison_df, not_found_count).
    not_found_count = number of "No records to display." placeholder rows in
    the Input Excel (counties the source site returned zero cases for). Those
    rows are NOT real cases, so they are excluded from comparison_df / the
    New Case count, but they ARE counted here so the GUI can surface them
    under NOT FOUND instead of silently dropping them."""
    if input_df.empty:
        log.warning("Input Excel is empty/missing - skipping input-vs-output validation.")
        return pd.DataFrame(), 0
    if output_df.empty:
        log.warning("Output Excel is empty/missing - skipping input-vs-output validation.")
        return pd.DataFrame(), 0

    in_case_col = find_column(input_df.columns, INPUT_COLUMN_CANDIDATES["Case #"])
    in_county_col = find_column(input_df.columns, INPUT_COLUMN_CANDIDATES["County"])

    output_cols = resolve_output_columns(output_df)
    out_case_col = output_cols.get("Case #")
    out_county_col = output_cols.get("County")

    if in_case_col is None or out_case_col is None:
        raise ValueError("Could not locate a 'Case #' column in the Input and/or Output Excel.")

    # Keep ALL Output rows per Case # (not just the first) - a Case # can
    # legitimately appear under more than one County, and collapsing to a
    # single row per Case # was silently discarding the other county's data
    # and could report a false County "Mismatch".
    output_lookup = {}
    for _, row in output_df.iterrows():
        key = normalize_key(row.get(out_case_col, ""))
        if not key:
            continue
        output_lookup.setdefault(key, []).append(row)

    results = []
    not_found_count = 0
    for _, in_row in input_df.iterrows():
        in_case = in_row.get(in_case_col, "")
        key = normalize_key(in_case)

        if not key:
            # blank Case # in the Input Excel - nothing to compare, skip this row entirely
            continue

        if is_no_records_placeholder(in_case):
            # source site's "no cases found for this county" placeholder row -
            # not a real case, so it must not be counted as New Case. Still
            # written into the result sheet (and counted) as "Not Found" so
            # it's visible in the Excel, not just the console/log.
            not_found_count += 1
            county_for_msg = in_row.get(in_county_col, "") if in_county_col else ""
            msg = f"'No records to display.' placeholder row for county={county_for_msg} -> counted as Not Found"
            log.info(msg)
            results.append({
                "Case #": in_case,
                "Input County": county_for_msg,
                "Output County": "",
                "Case # Status": "Not Found",
                "County Status": "Not Found",
                "Source Input File": in_row.get("_source_file", ""),
                "Notes": msg,
            })
            continue

        in_county = in_row.get(in_county_col, "") if in_county_col else ""

        out_candidates = output_lookup.get(key, [])
        if not out_candidates:
            results.append({
                "Case #": in_case,
                "Input County": in_county,
                "Output County": "",
                "Case # Status": "New Case",
                "County Status": "New Case",
                "Source Input File": in_row.get("_source_file", ""),
                "Notes": "",
            })
            continue

        # Same Case # can exist under more than one County in the Output -
        # match to the row whose County agrees with this Input row's County
        # first; only fall back to the first row if none matches.
        out_row = None
        for cand in out_candidates:
            cand_county = cand.get(out_county_col, "") if out_county_col else ""
            if normalize_value_for_compare(cand_county).upper() == normalize_value_for_compare(in_county).upper():
                out_row = cand
                break
        if out_row is None:
            out_row = out_candidates[0]

        out_county = out_row.get(out_county_col, "") if out_county_col else ""
        county_status = (
            "Match"
            if normalize_value_for_compare(in_county) == normalize_value_for_compare(out_county)
            else "Mismatch"
        )

        notes = ""
        if len(out_candidates) > 1:
            other_counties = sorted({
                normalize_value_for_compare(c.get(out_county_col, "")) for c in out_candidates
            } - {normalize_value_for_compare(out_county)})
            if other_counties:
                notes = (f"Case # also exists in Output under other County/Counties: "
                         f"{', '.join(other_counties)}")

        results.append({
            "Case #": in_case,
            "Input County": in_county,
            "Output County": out_county,
            "Case # Status": "Match",
            "County Status": county_status,
            "Source Input File": in_row.get("_source_file", ""),
            "Notes": notes,
        })

    columns = ["Case #", "Input County", "Output County", "Case # Status",
               "County Status", "Source Input File", "Notes"]
    result_df = pd.DataFrame(results)
    if result_df.empty:
        result_df = pd.DataFrame(columns=columns)
    else:
        result_df = result_df[columns]
    return result_df, not_found_count


# ============================================================
# STEP 5: WRITE RESULTS
# ============================================================

GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
# Prefixed VAL_ to avoid colliding with the Citation Search module's own
# (differently styled) HEADER_FILL / HEADER_FONT constants of the same name.
VAL_HEADER_FILL = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
VAL_HEADER_FONT = Font(color="FFFFFF", bold=True)


def _write_df_to_sheet(ws, df, match_col_predicate=None):
    """
    match_col_predicate: optional function(col_name) -> bool, marking which
    columns should be color-coded Match(green)/Mismatch(red).
    """
    if df.empty:
        ws.append(["No records to display."])
        return

    ws.append(list(df.columns))
    for cell in ws[1]:
        cell.fill = VAL_HEADER_FILL
        cell.font = VAL_HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    if match_col_predicate:
        status_col_idx = [i for i, c in enumerate(df.columns) if match_col_predicate(c)]
    else:
        status_col_idx = []

    for _, row in df.iterrows():
        ws.append(list(row))
        r = ws.max_row
        for idx in status_col_idx:
            cell = ws.cell(row=r, column=idx + 1)
            val = str(cell.value or "")
            if val == "Match":
                cell.fill = GREEN_FILL
            elif val.startswith("Mismatch") or val == "Not Found in Output" or "Not Found" in val or val == "New Case":
                cell.fill = RED_FILL

    for i, col in enumerate(df.columns, start=1):
        max_len = max([len(str(col))] + [len(str(v)) for v in df[col].astype(str).tolist()])
        ws.column_dimensions[get_column_letter(i)].width = min(max(10, max_len + 2), 40)

    ws.freeze_panes = "A2"


def build_validation_result_df(result_rows):
    """Build the final DataFrame with columns in the exact required order.
    Rows where every single Tool_Input AND Output value is blank (nothing
    extracted from HTML and nothing in the Output Excel) are dropped -
    they carry no information."""
    columns = []
    for field in FIELD_ORDER:
        columns += [f"{field} (Tool_Input)", f"{field} (Output)", f"{field} Match"]
    columns += ["Row Status"]

    df = pd.DataFrame(result_rows)
    for col in columns:
        if col not in df.columns:
            df[col] = ""
    if df.empty:
        return pd.DataFrame(columns=columns)

    value_cols = [f"{field} (Tool_Input)" for field in FIELD_ORDER] + [f"{field} (Output)" for field in FIELD_ORDER]
    non_empty_mask = df[value_cols].apply(lambda col: col.astype(str).str.strip() != "").any(axis=1)
    df = df[non_empty_mask]

    return df[columns]


def write_results(result_rows, case_summary, input_output_df, case_county_check_df,
                   result_dir, output_name_tag="Output"):
    os.makedirs(result_dir, exist_ok=True)
    out_path = os.path.join(result_dir, f"{output_name_tag}_Validation_Results.xlsx")

    validation_df = build_validation_result_df(result_rows)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws1 = wb.create_sheet("Validation_Result")
    _write_df_to_sheet(ws1, validation_df, match_col_predicate=lambda c: c.endswith("Match") or c == "Row Status")

    ws2 = wb.create_sheet("Case#_Diff_County_Check")
    _write_df_to_sheet(ws2, case_county_check_df)
    # "Status" isn't a plain Match/Mismatch column like the others (it can
    # read "OK" or a flag sentence), so re-color it here instead of relying
    # on _write_df_to_sheet's Match/Mismatch coloring.
    if not case_county_check_df.empty and "Status" in case_county_check_df.columns:
        status_idx = list(case_county_check_df.columns).index("Status") + 1
        flag_fill = PatternFill(start_color="FEF3CD", end_color="FEF3CD", fill_type="solid")
        for r in range(2, ws2.max_row + 1):
            cell = ws2.cell(row=r, column=status_idx)
            if str(cell.value or "").strip() not in ("OK", ""):
                cell.fill = flag_fill
            else:
                cell.fill = GREEN_FILL

    ws3 = wb.create_sheet("Input_vs_Output")
    _write_df_to_sheet(ws3, input_output_df, match_col_predicate=lambda c: c in ("Case # Status", "County Status"))

    wb.save(out_path)
    log.info("Results written to %s", out_path)
    return out_path


# ============================================================
# CORE RUN (used by both the GUI and the CLI fallback)
# ============================================================

def run_validation(offline_dir, output_dir, input_dir, result_base_dir):
    """Run the full validation pipeline against the given folders and return
    a summary dict. Raises on unrecoverable errors (caller should catch)."""
    log.info("=== GA Probate Validation Tool starting ===")

    html_cases = load_all_html_cases(offline_dir)
    output_df = load_excel_folder(output_dir)
    input_df = load_excel_folder(input_dir)

    result_rows, case_summary = build_validation_rows(html_cases, output_df)
    input_output_df, not_found_cases = run_input_vs_output_validation(input_df, output_df)
    case_county_check_df = build_case_county_check(html_cases, output_df)

    same_case_diff_county_count = 0
    if not case_county_check_df.empty:
        same_case_diff_county_count = int((case_county_check_df["Status"] != "OK").sum())
        if same_case_diff_county_count:
            log.warning(
                "%d Case #(s) flagged in Case#_Diff_County_Check (same Case # under "
                "different County / HTML vs Output County mismatch) - please review.",
                same_case_diff_county_count,
            )

    today_str = datetime.now().strftime("%d%m%Y")
    result_dir = os.path.join(result_base_dir, f"Result_{today_str}")

    output_name_tag = get_output_excel_name_tag(output_dir)
    out_path = write_results(result_rows, case_summary, input_output_df, case_county_check_df,
                              result_dir, output_name_tag)

    total_cases = len(case_summary)
    matched_cases = sum(1 for c in case_summary if c["Overall Status"] == "Match")
    mismatched_cases = sum(1 for c in case_summary if c["Overall Status"] == "Mismatch")

    # "New Case" = read straight from the Input_vs_Output sheet's "Case # Status"
    # column (an Input Excel case whose Case # doesn't exist anywhere in Output yet).
    if not input_output_df.empty and "Case # Status" in input_output_df.columns:
        new_case_count = int((input_output_df["Case # Status"] == "New Case").sum())
    else:
        new_case_count = 0

    # "Not Found" = "No records to display." placeholder rows in the Input
    # Excel (counties the source site returned zero cases for) - counted by
    # run_input_vs_output_validation, not part of input_output_df itself.

    log.info(
        "Done. %d case(s) processed, %d fully matched, %d mismatched, %d new case(s), %d not found (Input_vs_Output).",
        total_cases, matched_cases, mismatched_cases, new_case_count, not_found_cases,
    )
    log.info("Open the results file: %s", out_path)

    return {
        "out_path": out_path,
        "result_dir": result_dir,
        "total_cases": total_cases,
        "matched_cases": matched_cases,
        "mismatched_cases": mismatched_cases,
        "new_case_count": new_case_count,
        "not_found_cases": not_found_cases,
        "same_case_diff_county_count": same_case_diff_county_count,
    }


# ============================================================
# GUI
# ============================================================

class QueueLogHandler(logging.Handler):
    """Logging handler that pushes formatted records onto a thread-safe queue
    so the Tk main loop can safely display them in the log console."""

    def __init__(self, log_queue):
        super().__init__()
        self.log_queue = log_queue

    def emit(self, record):
        try:
            msg = self.format(record)
            self.log_queue.put((record.levelname, msg))
        except Exception:
            pass


class LabeledPathRow(tk.Frame):
    """Sidebar row: label + entry (with default path) + Browse button."""

    def __init__(self, parent, label_text, default_path):
        super().__init__(parent, bg=BG_SIDEBAR)
        tk.Label(
            self, text=label_text, bg=BG_SIDEBAR, fg="#C9D6EA",
            font=("Segoe UI", 9, "bold"), anchor="w",
        ).pack(fill="x", padx=2, pady=(8, 2))

        row = tk.Frame(self, bg=BG_SIDEBAR)
        row.pack(fill="x")

        self.var = tk.StringVar(value=default_path)
        entry = tk.Entry(
            row, textvariable=self.var, font=("Segoe UI", 9),
            bg="#FFFFFF", fg=TEXT_PRI, relief="flat",
            highlightthickness=1, highlightbackground=BORDER_GUI,
            highlightcolor=ACCENT,
        )
        entry.pack(side="left", fill="x", expand=True, ipady=4, padx=(0, 4))

        browse_btn = tk.Button(
            row, text="...", command=self._browse, bg=ACCENT, fg="white",
            activebackground=ACCENT_HOVER, activeforeground="white",
            relief="flat", font=("Segoe UI", 9, "bold"), width=3, cursor="hand2",
        )
        browse_btn.pack(side="right")

    def _browse(self):
        chosen = filedialog.askdirectory(initialdir=self.var.get() or SCRIPT_DIR)
        if chosen:
            self.var.set(chosen)

    def get(self):
        return self.var.get().strip()


class StatCard(tk.Frame):
    """Small colored summary card used in the results header."""

    def __init__(self, parent, title, fg, bg):
        super().__init__(parent, bg=bg, highlightbackground=BORDER_GUI, highlightthickness=1)
        self.value_lbl = tk.Label(
            self, text="0", bg=bg, fg=fg, font=("Segoe UI", 20, "bold"),
        )
        self.value_lbl.pack(pady=(12, 0))
        tk.Label(
            self, text=title, bg=bg, fg=fg, font=("Segoe UI", 9, "bold"),
        ).pack(pady=(0, 12))

    def set_value(self, value):
        self.value_lbl.configure(text=str(value))


class ValidationApp(tk.Toplevel):
    LEVEL_COLORS = {
        "INFO": LOG_INFO,
        "WARNING": LOG_WARN,
        "ERROR": LOG_ERR,
        "CRITICAL": LOG_ERR,
        "DEBUG": C_GRAY_BG,
    }

    def __init__(self, master=None):
        super().__init__(master)
        self.title("GA Probate Court - Validation Tool")
        self.geometry("1180x720")
        self.minsize(980, 620)
        self.configure(bg=BG_MAIN)

        self.log_queue = queue.Queue()
        self.result_info = None
        self.worker_thread = None
        self.run_start_time = None
        self._elapsed_running = False
        self._elapsed_job = None

        self._build_sidebar()
        self._build_main_area()

        self._attach_log_handler()
        self.after(100, self._poll_log_queue)

    # ---------------- sidebar ----------------

    def _build_sidebar(self):
        sidebar = tk.Frame(self, bg=BG_SIDEBAR, width=320)
        sidebar.pack(side="left", fill="y")
        sidebar.pack_propagate(False)

        tk.Label(
            sidebar, text="GA Probate", bg=BG_SIDEBAR, fg="white",
            font=("Segoe UI", 16, "bold"), anchor="w",
        ).pack(fill="x", padx=16, pady=(20, 0))
        tk.Label(
            sidebar, text="Validation Tool", bg=BG_SIDEBAR, fg="#9FB3D6",
            font=("Segoe UI", 10), anchor="w",
        ).pack(fill="x", padx=16, pady=(0, 16))

        paths_frame = tk.Frame(sidebar, bg=BG_SIDEBAR)
        paths_frame.pack(fill="x", padx=16)

        select_all_btn = tk.Button(
            paths_frame, text="📁  Select All Paths", command=self._select_all_paths,
            bg="#2A4A7F", fg="white", activebackground=ACCENT, activeforeground="white",
            relief="flat", font=("Segoe UI", 9, "bold"), cursor="hand2", pady=6,
        )
        select_all_btn.pack(fill="x", pady=(0, 6))

        self.row_offline = LabeledPathRow(paths_frame, "Offline HTML Folder", OFFLINE_HTML_DIR)
        self.row_offline.pack(fill="x")
        self.row_output = LabeledPathRow(paths_frame, "Output Excel Folder", OUTPUT_EXCEL_DIR)
        self.row_output.pack(fill="x")
        self.row_input = LabeledPathRow(paths_frame, "Input Excel Folder", INPUT_EXCEL_DIR)
        self.row_input.pack(fill="x")
        self.row_result = LabeledPathRow(paths_frame, "Result Base Folder", RESULT_BASE_DIR)
        self.row_result.pack(fill="x")

        tk.Frame(sidebar, bg="#2A4A7F", height=1).pack(fill="x", padx=16, pady=18)

        self.run_btn = tk.Button(
            sidebar, text="▶  RUN VALIDATION", command=self._on_run_clicked,
            bg=ACCENT, fg="white", activebackground=ACCENT_HOVER, activeforeground="white",
            relief="flat", font=("Segoe UI", 11, "bold"), cursor="hand2", pady=10,
        )
        self.run_btn.pack(fill="x", padx=16)

        self.open_folder_btn = tk.Button(
            sidebar, text="Open Results Folder", command=self._open_results_folder,
            bg=BG_SIDEBAR, fg="#9FB3D6", activebackground="#2A4A7F", activeforeground="white",
            relief="flat", font=("Segoe UI", 9, "underline"), cursor="hand2", state="disabled",
        )
        self.open_folder_btn.pack(fill="x", padx=16, pady=(8, 0))

        self.status_var = tk.StringVar(value="Ready")
        tk.Label(
            sidebar, textvariable=self.status_var, bg=BG_SIDEBAR, fg="#C9D6EA",
            font=("Segoe UI", 9), anchor="w", wraplength=280, justify="left",
        ).pack(fill="x", padx=16, pady=(20, 12), side="bottom")

    # ---------------- main area ----------------

    def _build_main_area(self):
        main = tk.Frame(self, bg=BG_MAIN)
        main.pack(side="left", fill="both", expand=True)

        cards_frame = tk.Frame(main, bg=BG_MAIN)
        cards_frame.pack(fill="x", padx=20, pady=(20, 10))

        self.card_total = StatCard(cards_frame, "TOTAL CASES", TEXT_PRI, BG_CARD)
        self.card_matched = StatCard(cards_frame, "MATCHED", C_GREEN, C_GREEN_BG)
        self.card_mismatched = StatCard(cards_frame, "MISMATCHED", C_RED, C_RED_BG)
        self.card_notfound = StatCard(cards_frame, "NOT FOUND", C_ORANGE, C_ORANGE_BG)
        self.card_newcase = StatCard(cards_frame, "NEW CASE", C_GRAY, C_GRAY_BG)
        self.card_diffcounty = StatCard(cards_frame, "CASE# DIFF COUNTY", C_ORANGE, C_ORANGE_BG)
        self.card_elapsed = StatCard(cards_frame, "RUNNING TIME", ACCENT, BG_CARD)
        self.card_elapsed.set_value("00:00:00")

        for card in (self.card_total, self.card_matched, self.card_mismatched, self.card_notfound,
                     self.card_newcase, self.card_diffcounty, self.card_elapsed):
            card.pack(side="left", fill="both", expand=True, padx=6)

        log_card = tk.Frame(main, bg=BG_CARD, highlightbackground=BORDER_GUI, highlightthickness=1)
        log_card.pack(fill="both", expand=True, padx=20, pady=(10, 20))

        tk.Label(
            log_card, text="Activity Log", bg=BG_CARD, fg=TEXT_PRI,
            font=("Segoe UI", 10, "bold"), anchor="w",
        ).pack(fill="x", padx=12, pady=(10, 4))

        log_container = tk.Frame(log_card, bg=BG_LOG)
        log_container.pack(fill="both", expand=True, padx=12, pady=(0, 12))

        self.log_text = tk.Text(
            log_container, bg=BG_LOG, fg=LOG_INFO, insertbackground="white",
            font=("Consolas", 9), relief="flat", wrap="word", state="disabled",
            padx=10, pady=8,
        )
        scrollbar = ttk.Scrollbar(log_container, command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=scrollbar.set)
        self.log_text.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        for level, color in self.LEVEL_COLORS.items():
            self.log_text.tag_configure(level, foreground=color)

    # ---------------- logging plumbing ----------------

    def _attach_log_handler(self):
        handler = QueueLogHandler(self.log_queue)
        handler.setFormatter(logging.Formatter("%(asctime)s  %(levelname)-7s  %(message)s", "%H:%M:%S"))
        log.addHandler(handler)

    def _poll_log_queue(self):
        try:
            while True:
                level, msg = self.log_queue.get_nowait()
                self._append_log(level, msg)
        except queue.Empty:
            pass
        self.after(100, self._poll_log_queue)

    def _append_log(self, level, msg):
        self.log_text.configure(state="normal")
        self.log_text.insert("end", msg + "\n", level if level in self.LEVEL_COLORS else "INFO")
        self.log_text.see("end")
        self.log_text.configure(state="disabled")

    # ---------------- run handling ----------------

    def _on_run_clicked(self):
        if self.worker_thread and self.worker_thread.is_alive():
            return

        offline_dir = self.row_offline.get()
        output_dir = self.row_output.get()
        input_dir = self.row_input.get()
        result_base_dir = self.row_result.get()

        missing = [p for p in (offline_dir, output_dir, input_dir, result_base_dir) if not p]
        if missing:
            messagebox.showerror("Missing Folder", "Please fill in all four folder paths before running.")
            return

        self.run_btn.configure(state="disabled", text="Running...")
        self.open_folder_btn.configure(state="disabled")
        self.status_var.set("Validation running - see Activity Log below.")
        self._reset_cards()
        self._start_elapsed_timer()

        self.worker_thread = threading.Thread(
            target=self._run_worker,
            args=(offline_dir, output_dir, input_dir, result_base_dir),
            daemon=True,
        )
        self.worker_thread.start()

    def _run_worker(self, offline_dir, output_dir, input_dir, result_base_dir):
        try:
            info = run_validation(offline_dir, output_dir, input_dir, result_base_dir)
            self.after(0, self._on_run_success, info)
        except Exception as e:
            log.error("Validation failed: %s", e)
            self.after(0, self._on_run_failure, str(e))

    def _on_run_success(self, info):
        self.result_info = info
        self.card_total.set_value(info["total_cases"])
        self.card_matched.set_value(info["matched_cases"])
        self.card_mismatched.set_value(info["mismatched_cases"])
        self.card_notfound.set_value(info["not_found_cases"])
        self.card_newcase.set_value(info["new_case_count"])
        self.card_diffcounty.set_value(info.get("same_case_diff_county_count", 0))
        elapsed_str = self._stop_elapsed_timer()

        self.run_btn.configure(state="normal", text="▶  RUN VALIDATION")
        self.open_folder_btn.configure(state="normal")
        self.status_var.set(f"Done in {elapsed_str}. Results saved to:\n{info['out_path']}")
        log.info("Total running time: %s", elapsed_str)

    def _on_run_failure(self, error_msg):
        elapsed_str = self._stop_elapsed_timer()
        self.run_btn.configure(state="normal", text="▶  RUN VALIDATION")
        self.status_var.set(f"Failed after {elapsed_str}: {error_msg}")
        messagebox.showerror("Validation Failed", error_msg)

    def _reset_cards(self):
        for card in (self.card_total, self.card_matched, self.card_mismatched, self.card_notfound,
                     self.card_newcase, self.card_diffcounty):
            card.set_value(0)
        self.card_elapsed.set_value("00:00:00")

    # ---------------- running-time tracking ----------------

    @staticmethod
    def _format_elapsed(seconds):
        seconds = max(0, int(seconds))
        h, rem = divmod(seconds, 3600)
        m, s = divmod(rem, 60)
        return f"{h:02d}:{m:02d}:{s:02d}"

    def _tick_elapsed(self):
        if not self._elapsed_running or self.run_start_time is None:
            return
        elapsed = time.time() - self.run_start_time
        self.card_elapsed.set_value(self._format_elapsed(elapsed))
        self._elapsed_job = self.after(500, self._tick_elapsed)

    def _start_elapsed_timer(self):
        self.run_start_time = time.time()
        self._elapsed_running = True
        self._tick_elapsed()

    def _stop_elapsed_timer(self):
        self._elapsed_running = False
        if self._elapsed_job is not None:
            try:
                self.after_cancel(self._elapsed_job)
            except Exception:
                pass
            self._elapsed_job = None
        if self.run_start_time is not None:
            self.card_elapsed.set_value(self._format_elapsed(time.time() - self.run_start_time))
        return self._format_elapsed(time.time() - self.run_start_time) if self.run_start_time else "00:00:00"

    def _select_all_paths(self):
        """Let the user pick ONE root folder (e.g. the 'New' case folder) and
        auto-fill all four path fields from it in a single action:
            root\\Offline        -> Offline HTML Folder
            root\\Output         -> Output Excel Folder
            root\\Input Folder   -> Input Excel Folder
            root                -> Result Base Folder
        Falls back to a case-insensitive search for likely subfolder names,
        and leaves a field untouched if nothing suitable is found under root."""
        root = filedialog.askdirectory(initialdir=SCRIPT_DIR, title="Select the root case folder")
        if not root:
            return

        def find_subfolder(candidates):
            try:
                entries = {e.lower(): e for e in os.listdir(root) if os.path.isdir(os.path.join(root, e))}
            except OSError:
                return None
            for cand in candidates:
                if cand.lower() in entries:
                    return os.path.join(root, entries[cand.lower()])
            return None

        offline_match = find_subfolder(["Offline", "Offline HTML", "HTML"])
        output_match = find_subfolder(["Output", "Output Excel"])
        input_match = find_subfolder(["Input Folder", "Input", "Input Excel"])

        if offline_match:
            self.row_offline.var.set(offline_match)
        if output_match:
            self.row_output.var.set(output_match)
        if input_match:
            self.row_input.var.set(input_match)
        self.row_result.var.set(root)

        found = sum(bool(m) for m in (offline_match, output_match, input_match))
        self.status_var.set(f"Root folder set. Auto-detected {found}/3 subfolders under:\n{root}")

    def _open_results_folder(self):
        if not self.result_info:
            return
        folder = self.result_info["result_dir"]
        try:
            if sys.platform.startswith("win"):
                os.startfile(folder)  # noqa: S606 (Windows-only convenience)
            elif sys.platform == "darwin":
                subprocess.Popen(["open", folder])
            else:
                subprocess.Popen(["xdg-open", folder])
        except Exception as e:
            messagebox.showerror("Could Not Open Folder", str(e))


def open_validation_window(master=None):
    """Factory used by the combined-tool launcher to open this module's
    window as a Toplevel of the launcher's root window."""
    return ValidationApp(master=master)


# ============================================================
# MAIN (CLI mode -- unchanged behavior, invoked via `--cli`)
# ============================================================

def main():
    """CLI / headless mode - runs once from Command Prompt with NO GUI window,
    logging progress straight to the console, then exits.

    USAGE (from a Command Prompt / cmd.exe, in the script's folder):

        python ga_probate_validation_tool.py --cli
            Runs using the CONFIG paths hardcoded at the top of this file
            (OFFLINE_HTML_DIR, OUTPUT_EXCEL_DIR, INPUT_EXCEL_DIR, RESULT_BASE_DIR).

        python ga_probate_validation_tool.py --cli ^
            --offline "D:\\Mohan\\GA_Probate\\New\\Offline" ^
            --output  "D:\\Mohan\\GA_Probate\\New\\Output" ^
            --input   "D:\\Mohan\\GA_Probate\\New\\Input Folder" ^
            --result-base "D:\\Mohan\\GA_Probate\\New"
            Overrides one or more folders for this run only (handy for Task
            Scheduler / batch files that point at a different case folder
            each time). The caret ^ is cmd.exe's line-continuation character.

        python ga_probate_validation_tool.py --cli --help
            Shows this same option list from the command line.

    Exit code is 0 on success, 1 on failure, so a .bat file or Task Scheduler
    job can check %ERRORLEVEL% / the process exit code to detect problems.

    Without --cli, the script opens the GUI instead (see launch_gui()).
    """
    import argparse

    parser = argparse.ArgumentParser(
        prog="ga_probate_validation_tool.py --cli",
        description="Run the GA Probate validation headlessly from a Command Prompt (no GUI window).",
    )
    parser.add_argument("--cli", action="store_true", help=argparse.SUPPRESS)
    parser.add_argument("--offline", default=OFFLINE_HTML_DIR, metavar="PATH",
                         help=f"Offline HTML folder (default: {OFFLINE_HTML_DIR})")
    parser.add_argument("--output", default=OUTPUT_EXCEL_DIR, metavar="PATH",
                         help=f"Output Excel folder (default: {OUTPUT_EXCEL_DIR})")
    parser.add_argument("--input", default=INPUT_EXCEL_DIR, metavar="PATH",
                         help=f"Input Excel folder (default: {INPUT_EXCEL_DIR})")
    parser.add_argument("--result-base", default=RESULT_BASE_DIR, metavar="PATH",
                         help=f"Result base folder, where Result_DDMMYYYY is created (default: {RESULT_BASE_DIR})")
    args = parser.parse_args(sys.argv[1:])

    print("=" * 60)
    print(" GA Probate Validation Tool - running in CLI mode (no GUI)")
    print("=" * 60)
    print(f"  Offline HTML Folder : {args.offline}")
    print(f"  Output Excel Folder : {args.output}")
    print(f"  Input Excel Folder  : {args.input}")
    print(f"  Result Base Folder  : {args.result_base}")
    print("-" * 60)

    try:
        info = run_validation(args.offline, args.output, args.input, args.result_base)
    except Exception as e:
        print(f"\nFAILED: {e}")
        sys.exit(1)

    print("-" * 60)
    print(f"  Total Cases    : {info['total_cases']}")
    print(f"  Matched        : {info['matched_cases']}")
    print(f"  Mismatched     : {info['mismatched_cases']}")
    print(f"  New Case       : {info['new_case_count']}")
    print(f"  Not Found      : {info['not_found_cases']}")
    print(f"  Case# Diff County (flagged) : {info.get('same_case_diff_county_count', 0)}")
    print("-" * 60)
    print(f"  Results file   : {info['out_path']}")
    print("=" * 60)
    sys.exit(0)


# ================================================================
# MODULE: Citation Search Tool (Live Selenium scrape + validation)
# ================================================================
# GUI tool for https://georgiaprobaterecords.com/Traffic/SearchCitations.aspx
#
# This site needs NO LOGIN — there is no credential/CAPTCHA/session step.
# The GUI runs a single pipeline:
#
#   "Live Search + Validation" - drives a real Chrome browser (Selenium)
#                           against the live site for every County/Start/End
#                           row in an INPUT EXCEL file, scrapes the results
#                           grid, and writes:
#                             - a combined "Output" workbook of every case
#                               number found (Output folder)
#                             - a "Result" workbook comparing those case
#                               numbers against the pre-existing case-number
#                               list already sitting in the Tool Input /
#                               Compare folder, with Match / New / Missing
#                               status + color coding (Result folder)
#
# If a county's search fails (or returns 0 rows), the raw page HTML is still
# saved to the Offline folder as debug_<county>.html for troubleshooting.
#
# Counties are read from an INPUT EXCEL file (columns: County, Starting
# Date, Ending Date — dates as DD-MM-YYYY, e.g. 20-06-2026), or a legacy
# tab-separated .txt file.
# ================================================================

# Selenium is only needed for the Live Search run — imported lazily so the
# GUI itself still launches even on a machine without Chrome/selenium
# installed (it will just error when the user tries to run a search).
try:
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.common.keys import Keys
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.common.exceptions import (
        TimeoutException, NoSuchElementException, StaleElementReferenceException
    )
    SELENIUM_AVAILABLE = True
except ImportError:
    SELENIUM_AVAILABLE = False

# BeautifulSoup is used by both modules; already imported in the shared
# header, so it's always available here too.
BS4_AVAILABLE = True

from openpyxl import Workbook

# NOTE: GUI theme constants, SCRIPT_DIR, and the shared `log` logger are
# defined once in the SHARED HEADER section near the top of this combined
# file, and reused by both modules.

# ─────────────────────────────────────────────────────────────────────────
# CONFIG — defaults (all overridable from the GUI, all auto-created)
# ─────────────────────────────────────────────────────────────────────────
SEARCH_URL = "https://georgiaprobaterecords.com/Traffic/SearchCitations.aspx"

INPUT_FOLDER      = r"D:\Mohan\GA_Probate\Input"      # holds the County/Date Excel or .txt
OUTPUT_FOLDER      = r"D:\Mohan\GA_Probate\Input Folder"      # combined scrape-results workbook (saved as result_MMDDYYYY.xlsx directly here)
OFFLINE_FOLDER      = r"D:\Mohan\GA_Probate\Offline"    # debug_<county>.html dumps
TOOL_INPUT_FOLDER  = r"D:\Mohan\GA_Probate\Tool_Input"  # existing case list + Result workbook

MATCH_STATUS_FILE_PREFIX = "CaseNumber_MatchStatus_"

DEBUG_HTML_BASE_HREF = "https://georgiaprobaterecords.com/Traffic/"
PAGE_LOAD_TIMEOUT = 25
ELEMENT_TIMEOUT = 20


def ensure_folder(path):
    os.makedirs(path, exist_ok=True)
    return path


# ─────────────────────────────────────────────────────────────────────────
# DUPLICATE-MASKED MISSING CASE DETECTION
#
# Root cause of the "page 7 data missing" bug: on a live postback, RadGrid
# can occasionally render the SAME row twice (identical case number, same
# underlying RECID) in place of what should have been the next distinct
# row. The tool's existing de-dup step (keep first occurrence, drop the
# rest) hides the symptom but silently loses the row that should have
# taken that slot -- e.g. page 7 rendered 26T0314 twice back-to-back, and
# 26T0313 -- which belonged in that first slot -- never appeared anywhere
# in the result set at all. A duplicate found in the DATA is therefore not
# just "one extra row to remove", it's also a signal that some OTHER case
# number is probably missing.
#
# Case numbers on this site follow a PREFIX + zero-padded NUMBER pattern
# (e.g. "26T0313", "26T0314", ...). When duplicates are found, this helper
# looks at the run of numbers sharing each prefix and reports any gaps in
# that numeric sequence -- the most likely candidates for the case
# number(s) that got silently dropped.
# ─────────────────────────────────────────────────────────────────────────
_CASE_NO_SPLIT_RE = re.compile(r"^([A-Za-z0-9]*?)(\d+)$")


def _split_case_no(case_no):
    """Splits e.g. '26T0313' -> ('26T', '0313', 313). Returns None if the
    case number doesn't end in a run of digits."""
    m = _CASE_NO_SPLIT_RE.match((case_no or "").strip())
    if not m:
        return None
    prefix, digits = m.group(1), m.group(2)
    return prefix, digits, int(digits)


def find_duplicate_case_numbers(case_numbers):
    """Given a list of case-number strings (in the order encountered),
    returns (duplicates, first_seen_index) where `duplicates` is a sorted
    list of every case number that appeared more than once, and
    `first_seen_index` maps each case number to the row index of its
    FIRST occurrence (so callers can report exactly where in the run the
    duplicate showed up)."""
    seen_at = {}
    dupes = set()
    for i, cn in enumerate(case_numbers):
        cn = (cn or "").strip()
        if not cn:
            continue
        if cn in seen_at:
            dupes.add(cn)
        else:
            seen_at[cn] = i
    return sorted(dupes), seen_at


def infer_likely_missing_case_numbers(case_numbers):
    """Best-effort guess at which case number(s) were silently dropped when
    a duplicate is found. Groups all case numbers by their non-numeric
    prefix, looks at the min/max of the numeric suffix seen for that
    prefix, and reports any integer in that inclusive range which never
    appeared at all. This intentionally only reports gaps INSIDE the
    observed min..max span (not a guess about what might exist before/after
    it), since that's the span the tool actually has evidence for.

    Returns a dict: {prefix: [missing_case_no, ...], ...}. Empty dict if
    every prefix's numeric run is unbroken (or nothing parses as
    prefix+number).
    """
    by_prefix = {}
    for cn in case_numbers:
        parsed = _split_case_no(cn)
        if not parsed:
            continue
        prefix, digits, num = parsed
        by_prefix.setdefault(prefix, {"width": len(digits), "nums": set()})
        by_prefix[prefix]["nums"].add(num)

    missing = {}
    for prefix, info in by_prefix.items():
        nums = info["nums"]
        if len(nums) < 2:
            continue
        width = info["width"]
        lo, hi = min(nums), max(nums)
        gap = [n for n in range(lo, hi + 1) if n not in nums]
        if gap:
            missing[prefix] = [f"{prefix}{str(n).zfill(width)}" for n in gap]
    return missing


# ─────────────────────────────────────────────────────────────────────────
# INPUT EXCEL loading  (County | Starting Date | Ending Date)
# ─────────────────────────────────────────────────────────────────────────
def find_latest_excel(folder, label=""):
    if not os.path.isdir(folder):
        raise FileNotFoundError(f"{label or 'Folder'} not found: {folder}")
    candidates = [
        os.path.join(folder, f) for f in os.listdir(folder)
        if f.lower().endswith((".xlsx", ".xls")) and not f.startswith("~$")
    ]
    if not candidates:
        raise FileNotFoundError(f"No Excel file found in {label or folder}: {folder}")
    return max(candidates, key=os.path.getmtime)


# Filenames that are OUTPUTS this same tool writes (combined results,
# match-status/validation workbooks, the seeded baseline case-number
# list). These must never be auto-picked as the County/Date INPUT file,
# even if they land in the same folder and happen to be the newest file
# there (e.g. Output/Tool Input folder pointed at the Input folder).
_NON_INPUT_NAME_PREFIXES = (
    "result_",
    MATCH_STATUS_FILE_PREFIX.lower(),
    "lnga_casenumber_baseline",
)


def _looks_like_output_file(fname):
    lower = fname.lower()
    return any(lower.startswith(p) for p in _NON_INPUT_NAME_PREFIXES)


def _file_has_input_headers(path):
    """Peeks the header row only (cheap) and reports whether it contains
    County / Start / End -style columns, without loading the whole file."""
    try:
        ext = os.path.splitext(path)[1].lower()
        if ext == ".txt":
            with open(path, "r", newline="", encoding="utf-8-sig") as f:
                header = f.readline().lower()
            return "county" in header
        else:
            df = pd.read_excel(path, dtype=object, nrows=0)
            cols = [str(c).strip().lower().replace("_", " ") for c in df.columns]
            return any("county" in c for c in cols)
    except Exception:
        return False


def find_latest_input_file(folder, label=""):
    """Same as find_latest_excel but also considers .txt input files (the
    original weekly tab-separated format), since the Input File field now
    accepts either.

    Rejects known OUTPUT-style filenames outright (result_*, the
    CaseNumber_MatchStatus_* validation workbook, the seeded baseline
    list) and then, among what's left, only picks a file whose header row
    actually contains a County-style column — newest-first. This prevents
    silently auto-selecting a leftover output/result file that happens to
    share the Input folder and have a newer timestamp."""
    if not os.path.isdir(folder):
        raise FileNotFoundError(f"{label or 'Folder'} not found: {folder}")

    all_candidates = [
        f for f in os.listdir(folder)
        if f.lower().endswith((".xlsx", ".xls", ".txt")) and not f.startswith("~$")
    ]
    if not all_candidates:
        raise FileNotFoundError(f"No .txt or .xlsx input file found in {label or folder}: {folder}")

    candidates = [f for f in all_candidates if not _looks_like_output_file(f)]
    if not candidates:
        raise FileNotFoundError(
            f"Only output-style files (e.g. result_*, "
            f"{MATCH_STATUS_FILE_PREFIX}*) were found in {label or folder} — "
            f"no County/Starting Date/Ending Date input file to auto-detect. "
            f"Use Browse to pick the correct input file.")

    candidates.sort(key=lambda f: os.path.getmtime(os.path.join(folder, f)), reverse=True)

    for f in candidates:
        full_path = os.path.join(folder, f)
        if _file_has_input_headers(full_path):
            return full_path

    raise FileNotFoundError(
        f"Found {len(candidates)} candidate file(s) in {label or folder}, but none "
        f"has a 'County' column in its header row — nothing looks like a valid "
        f"input file. Use Browse to pick the correct one.")


def _normalize_date_cell(value):
    """County/date Excel cells may come back from Excel as real datetimes
    (if the user formatted the column as a date) or as plain DD-MM-YYYY
    text. Either way, return DD-MM-YYYY text so the rest of the script
    (unchanged from the original) keeps working exactly as before."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    if isinstance(value, (pd.Timestamp, datetime)):
        return value.strftime("%d-%m-%Y")
    text = str(value).strip()
    # Excel sometimes hands back "2026-06-20 00:00:00" for a date cell that
    # pandas didn't parse as a Timestamp for some reason.
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})", text)
    if m:
        return f"{m.group(3)}-{m.group(2)}-{m.group(1)}"
    return text


def load_input_excel(path):
    """Reads County / Starting Date / Ending Date from an Excel file.
    Header matching is case-insensitive and tolerant of minor variations
    (e.g. 'Start Date' vs 'Starting Date'). Returns the same list-of-dicts
    shape ({"county", "start_date", "end_date"}) the rest of the script
    (scraping, workbook writing) already expects."""
    df = pd.read_excel(path, dtype=object)
    df.columns = [str(c).strip() for c in df.columns]

    def _find_col(*keywords):
        for c in df.columns:
            norm = c.lower().replace("_", " ")
            if all(k in norm for k in keywords):
                return c
        return None

    county_col = _find_col("county")
    start_col = _find_col("start") or _find_col("starting")
    end_col = _find_col("end") or _find_col("ending")

    missing = [n for n, c in (("County", county_col), ("Starting Date", start_col),
                               ("Ending Date", end_col)) if c is None]
    if missing:
        raise RuntimeError(
            f"Input Excel '{path}' is missing column(s): {', '.join(missing)}.\n"
            f"Header row read as: {list(df.columns)}\n\n"
            f"This does not look like an LNGA Probate input file. The Input Excel "
            f"needs exactly these columns: County, Starting Date, Ending Date "
            f"(dates as DD-MM-YYYY). If you picked a file from a different tool's "
            f"folder by mistake, use Browse to select the correct one from the "
            f"GA_Probate Input folder instead.")

    rows = []
    for _, r in df.iterrows():
        county = str(r.get(county_col, "")).strip()
        if not county or county.lower() == "nan":
            continue
        rows.append({
            "county": county,
            "start_date": _normalize_date_cell(r.get(start_col)),
            "end_date": _normalize_date_cell(r.get(end_col)),
        })
    return rows


def load_input_txt(path):
    """Reads County / Starting Date / Ending Date from the original
    tab-separated .txt format (County\tStarting Date\tEnding Date), exactly
    as the weekly input file already comes. Returns the same list-of-dicts
    shape as load_input_excel()."""
    import csv
    rows = []
    with open(path, "r", newline="", encoding="utf-8-sig") as f:
        reader = csv.reader(f, delimiter="\t")
        next(reader, None)  # header
        for line_num, row in enumerate(reader, start=2):
            if not row or not any(c.strip() for c in row):
                continue
            if len(row) < 3:
                continue
            county, start_date, end_date = row[0].strip(), row[1].strip(), row[2].strip()
            if county:
                rows.append({"county": county, "start_date": start_date, "end_date": end_date})
    return rows


def load_input_file(path):
    """Dispatches to the .txt or .xlsx loader based on the file's
    extension, so the GUI's Input File field accepts either the original
    weekly tab-separated .txt file or an Excel workbook."""
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xls"):
        return load_input_excel(path)
    if ext == ".txt":
        return load_input_txt(path)
    raise RuntimeError(
        f"Unsupported input file type '{ext}'. Use a .txt (tab-separated: "
        f"County / Starting Date / Ending Date) or .xlsx (same columns) file.")


def to_date_input_format(ddmmyyyy_str):
    """Convert 'DD-MM-YYYY' to the site's own display format, M/D/YYYY with
    NO leading zeros (e.g. '6/20/2026')."""
    dt = datetime.strptime(ddmmyyyy_str.strip(), "%d-%m-%Y")
    return f"{dt.month}/{dt.day}/{dt.year}"


# ─────────────────────────────────────────────────────────────────────────
# Debug-HTML helper (patches saved debug_<county>.html so it can be opened
# later in a browser for troubleshooting, with CSS/JS/images resolving
# back against the live site)
# ─────────────────────────────────────────────────────────────────────────
def patch_debug_html_for_offline_viewing(html, base_href=DEBUG_HTML_BASE_HREF):
    """Inserts <base href="..."> so a saved debug_<county>.html opened later
    resolves its CSS/JS/images back against the live site."""
    base_tag = f'<base href="{base_href}">'
    if re.search(r"<head[^>]*>", html, re.IGNORECASE):
        return re.sub(r"(<head[^>]*>)", r"\1" + base_tag, html, count=1, flags=re.IGNORECASE)
    return base_tag + html


# ─────────────────────────────────────────────────────────────────────────
# Selenium driver + search-form interaction  (unchanged behavior — verified
# against the live search page; see the original script's history)
# ─────────────────────────────────────────────────────────────────────────
ID_COUNTY_DROPDOWN   = "ctl00_cpMain_ddlCounty"
ID_COUNTY_POPUP      = "ctl00_cpMain_ddlCounty_DropDown"
ID_OFFENSE_START     = "ctl00_cpMain_txtViolationStartDate_dateInput"
ID_OFFENSE_END       = "ctl00_cpMain_txtViolationEndDate_dateInput"
ID_SEARCH_BUTTON     = "ctl00_cpMain_rbSearch_input"
ID_RESULTS_GRID_CONTAINER = "ctl00_cpMain_rapGrid"

RESULT_COLUMNS = ["CASE #", "OFFENDER", "OFFENSE DATE", "COURT DATE", "DISPOSED", "COUNTY"]
_HEADER_KEY_MAP = {
    "CASE #":       "case_number",
    "OFFENDER":     "offender",
    "OFFENSE DATE": "offense_date",
    "COURT DATE":   "court_date",
    "DISPOSED":     "disposed",
    "COUNTY":       "county",
}


def build_driver(headless=True):
    options = webdriver.ChromeOptions()
    if headless:
        options.add_argument("--headless=new")
    options.add_argument("--window-size=1440,1000")
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")
    options.add_argument("--log-level=3")
    driver = webdriver.Chrome(options=options)
    driver.set_page_load_timeout(PAGE_LOAD_TIMEOUT)
    return driver


def select_county(driver, county_name, match_timeout=6):
    wait = WebDriverWait(driver, ELEMENT_TIMEOUT)
    dropdown_div = wait.until(EC.element_to_be_clickable((By.ID, ID_COUNTY_DROPDOWN)))
    dropdown_div.click()

    popup = wait.until(EC.visibility_of_element_located((By.ID, ID_COUNTY_POPUP)))
    target = county_name.strip().lower()

    match = None
    last_seen_texts = []
    deadline = time.time() + match_timeout
    while time.time() < deadline:
        items = popup.find_elements(By.XPATH, ".//li")
        last_seen_texts = [li.text.strip() for li in items]
        match = next((li for li, txt in zip(items, last_seen_texts) if txt.lower() == target), None)
        if match is not None:
            break
        time.sleep(0.2)

    if match is None:
        raise RuntimeError(
            f"County '{county_name}' not found in RadDropDownList after "
            f"{match_timeout}s of polling. Options seen: {last_seen_texts}")
    match.click()

    try:
        WebDriverWait(driver, 6).until(EC.invisibility_of_element_located((By.ID, ID_COUNTY_POPUP)))
    except TimeoutException:
        pass
    wait_for_ajax_idle(driver)


def _set_date_via_telerik_api(driver, picker_client_id, date_str):
    script = """
        var clientId = arguments[0];
        var dateStr = arguments[1];
        var sysDefined = (typeof Sys !== 'undefined');
        var picker = (window.$find) ? $find(clientId) : null;
        var pickerFound = !!(picker && picker.set_selectedDate);
        var setSucceeded = false;
        if (pickerFound) {
            var parts = dateStr.split('/');
            var d = new Date(parseInt(parts[2], 10), parseInt(parts[0], 10) - 1, parseInt(parts[1], 10));
            picker.set_selectedDate(d);
            setSucceeded = true;
        }
        return {sys_defined: sysDefined, picker_found: pickerFound, set_succeeded: setSucceeded};
    """
    try:
        return driver.execute_script(script, picker_client_id, date_str)
    except Exception as e:
        return {"sys_defined": None, "picker_found": False, "set_succeeded": False, "js_error": str(e)}


def _type_into_date_input(driver, input_el, date_str, picker_client_id=None, field_label=""):
    driver.execute_script("arguments[0].value = '';", input_el)
    input_el.click()
    input_el.send_keys(date_str)
    input_el.send_keys(Keys.TAB)

    actual = input_el.get_attribute("value")
    if actual.strip() == date_str.strip() and picker_client_id:
        try:
            has_date = driver.execute_script(
                "var p = window.$find ? $find(arguments[0]) : null;"
                "return !!(p && p.get_selectedDate && p.get_selectedDate());",
                picker_client_id)
        except Exception:
            has_date = True
        if has_date:
            return

    m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{4})$", date_str.strip())
    if m:
        padded = f"{int(m.group(1)):02d}/{int(m.group(2)):02d}/{m.group(3)}"
        if padded != date_str.strip():
            driver.execute_script("arguments[0].value = '';", input_el)
            input_el.click()
            input_el.send_keys(padded)
            input_el.send_keys(Keys.TAB)
            actual = input_el.get_attribute("value")
            if actual.strip() in (date_str.strip(), padded):
                return

    diagnostics = None
    if picker_client_id:
        diagnostics = _set_date_via_telerik_api(driver, picker_client_id, date_str)
        if diagnostics.get("set_succeeded"):
            final_val = input_el.get_attribute("value")
            if final_val.strip() == date_str.strip():
                return

    driver.execute_script(
        "arguments[0].value = arguments[1];"
        "arguments[0].dispatchEvent(new Event('change', {bubbles:true}));"
        "arguments[0].dispatchEvent(new Event('blur', {bubbles:true}));",
        input_el, date_str)
    final_val = input_el.get_attribute("value")
    if final_val.strip() == date_str.strip():
        return

    raise RuntimeError(
        f"Could not get {field_label or 'date field'} to hold '{date_str}' after typing, "
        f"the Telerik API call, and the raw-value fallback. Final input value seen: "
        f"'{final_val}'. Telerik API diagnostics: {diagnostics}.")


def wait_for_ajax_idle(driver, timeout=8):
    try:
        WebDriverWait(driver, timeout).until(
            lambda d: d.execute_script(
                "var m = window.Sys && Sys.WebForms && Sys.WebForms.PageRequestManager;"
                "return (document.readyState === 'complete') && "
                "(!m || !m.getInstance().get_isInAsyncPostBack());"
            )
        )
    except TimeoutException:
        pass


def _find_fresh(driver, element_id, timeout=ELEMENT_TIMEOUT):
    return WebDriverWait(driver, timeout).until(
        EC.presence_of_element_located((By.ID, element_id)))


def set_offense_date_range(driver, start_ddmmyyyy, end_ddmmyyyy):
    start_str = to_date_input_format(start_ddmmyyyy)
    end_str = to_date_input_format(end_ddmmyyyy)

    start_picker_id = ID_OFFENSE_START[: -len("_dateInput")]
    end_picker_id = ID_OFFENSE_END[: -len("_dateInput")]

    start_input = _find_fresh(driver, ID_OFFENSE_START)
    _type_into_date_input(driver, start_input, start_str, picker_client_id=start_picker_id,
                           field_label="Offense Date Start")
    wait_for_ajax_idle(driver)

    end_input = _find_fresh(driver, ID_OFFENSE_END)
    _type_into_date_input(driver, end_input, end_str, picker_client_id=end_picker_id,
                           field_label="Offense Date End")
    wait_for_ajax_idle(driver)


def click_search(driver):
    wait = WebDriverWait(driver, ELEMENT_TIMEOUT)
    try:
        btn = wait.until(EC.element_to_be_clickable((By.ID, ID_SEARCH_BUTTON)))
        btn.click()
        wait_for_ajax_idle(driver, timeout=ELEMENT_TIMEOUT)
        return
    except TimeoutException:
        pass
    strategies = [
        (By.XPATH, "//input[@type='submit' and contains(translate(@value,"
                    "'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'search')]"),
        (By.XPATH, "//button[contains(translate(text(),"
                    "'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'search')]"),
    ]
    for by, sel in strategies:
        try:
            btn = wait.until(EC.element_to_be_clickable((by, sel)))
            btn.click()
            wait_for_ajax_idle(driver, timeout=ELEMENT_TIMEOUT)
            return
        except TimeoutException:
            continue
    raise RuntimeError("Could not locate the Search button (tried several selectors).")


def set_page_size(driver, grid_container_id=ID_RESULTS_GRID_CONTAINER, target_size=30, timeout=8):
    """Forces the results grid's page-size combo box (Telerik RadComboBox,
    e.g. 'PageSizeComboBox') to `target_size` (default 30) right after a
    search, so each page of results holds up to 30 rows before the
    Next-page loop in scrape_result_rows() starts clicking through pages.

    This tries several selector strategies since the exact client ID isn't
    guaranteed across the site's grid instances. If no page-size control is
    found (e.g. too few rows for one to render, or the site changed), this
    logs nothing and simply returns False — scrape_result_rows() still
    pages through all results at whatever the default page size is, so no
    data is lost, just possibly slower to page through."""
    wait = WebDriverWait(driver, timeout)
    try:
        containers = driver.find_elements(By.ID, grid_container_id)
        search_root = containers[0] if containers else driver
    except Exception:
        search_root = driver

    target_text = str(target_size)

    # Strategy 1: plain <select> page-size dropdown (some RadGrid skins
    # render the pager's page-size picker as a native <select>).
    try:
        selects = search_root.find_elements(
            By.XPATH,
            ".//select[contains(translate(@id,'PAGESIZE','pagesize'),'pagesize') or "
            "contains(translate(@class,'PAGESIZE','pagesize'),'pagesize')]"
        )
        for sel_el in selects:
            options = sel_el.find_elements(By.TAG_NAME, "option")
            for opt in options:
                if opt.text.strip() == target_text or opt.get_attribute("value") == target_text:
                    opt.click()
                    wait_for_ajax_idle(driver, timeout=timeout)
                    time.sleep(0.4)
                    return True
    except Exception:
        pass

    # Strategy 2: Telerik RadComboBox rendered as a clickable div/input with
    # an id containing "PageSize" — open its dropdown popup and click the
    # "30" list item.
    try:
        combo_inputs = search_root.find_elements(
            By.XPATH,
            ".//*[contains(@id,'PageSizeComboBox') or contains(@id,'PageSize') "
            "or contains(@id,'pageSize')]"
        )
        for combo in combo_inputs:
            try:
                combo.click()
            except Exception:
                driver.execute_script("arguments[0].click();", combo)
            time.sleep(0.3)
            popups = driver.find_elements(
                By.XPATH,
                "//div[contains(@id,'PageSizeComboBox') and contains(@id,'DropDown')]//li | "
                "//ul[contains(@id,'PageSizeComboBox')]//li"
            )
            for li in popups:
                if li.text.strip() == target_text:
                    li.click()
                    wait_for_ajax_idle(driver, timeout=timeout)
                    time.sleep(0.4)
                    return True
    except Exception:
        pass

    # Strategy 3: raw Telerik client-side API — find any RadComboBox on the
    # page whose client ID mentions PageSize and set its value via JS.
    try:
        result = driver.execute_script("""
            var target = arguments[0];
            if (!window.Sys || !Sys.Application) { return false; }
            var comps = Sys.Application.getComponents ? Sys.Application.getComponents() : [];
            for (var i = 0; i < comps.length; i++) {
                var c = comps[i];
                var id = c.get_id ? c.get_id() : "";
                if (id && id.toLowerCase().indexOf("pagesize") !== -1 && c.set_text && c.findItemByText) {
                    var item = c.findItemByText(target);
                    if (item) { item.select(); return true; }
                }
            }
            return false;
        """, target_text)
        if result:
            wait_for_ajax_idle(driver, timeout=timeout)
            time.sleep(0.4)
            return True
    except Exception:
        pass

    return False


_PAGER_TEXT_RE = re.compile(r"page\s*size", re.IGNORECASE)

# Matches the RadGrid pager's own summary text, e.g. "96 items in 4 pages",
# rendered in a <div class="rgInfoPart"> inside the same pager row that
# _is_pager_row() excludes. Used purely as a self-check AFTER paging is
# complete — comparing the number of rows actually collected against the
# total the grid itself reports, so a silently-truncated scrape (e.g. the
# Next-click loop stopping early because of a slow AJAX postback) shows up
# as a logged warning instead of quietly shipping a partial result set.
_GRID_TOTAL_RE = re.compile(r"(\d+)\s+items?\s+in\s+(\d+)\s+pages?", re.IGNORECASE)


def _get_grid_expected_total(driver, grid_container_id=ID_RESULTS_GRID_CONTAINER):
    """Reads the RadGrid's 'N items in M pages' summary (rgInfoPart) and
    returns (expected_item_count, expected_page_count), or (None, None) if
    the summary isn't present (e.g. grid too small to show a pager)."""
    try:
        containers = driver.find_elements(By.ID, grid_container_id)
        search_root = containers[0] if containers else driver
        info_els = search_root.find_elements(By.XPATH, ".//div[contains(@class,'rgInfoPart')]")
        for el in info_els:
            m = _GRID_TOTAL_RE.search(el.text.strip())
            if m:
                return int(m.group(1)), int(m.group(2))
    except Exception:
        pass
    return None, None


def _is_pager_row(tr):
    """Identify the RadGrid pager row — the row that renders the page-number
    links ('1 2 3 4 ...') and the 'Page size:' dropdown — so it can be
    excluded from body_rows.

    This row lives inside the SAME <table> as the header and data rows, but
    it is neither. It typically renders as a single <td> with a colspan
    across the whole grid width, containing the pager UI as one text blob.
    Previously nothing excluded this row, so its entire text (e.g.
    '1 2 3 4 Page size: select') was being zipped into the first data
    column and — because it was non-empty — appended to the results as a
    bogus row on every single page.
    """
    try:
        cls = (tr.get_attribute("class") or "").lower()
        if "pager" in cls:
            return True
        tds = tr.find_elements(By.TAG_NAME, "td")
        if len(tds) == 1:
            td = tds[0]
            td_cls = (td.get_attribute("class") or "").lower()
            colspan = td.get_attribute("colspan")
            txt = td.text.strip()
            if "pager" in td_cls or _PAGER_TEXT_RE.search(txt):
                return True
            if colspan and colspan.isdigit() and int(colspan) > 1 and _PAGER_TEXT_RE.search(txt):
                return True
    except Exception:
        pass
    return False


def _click_next_numeric_page(search_root, driver, current_page_num):
    """Fallback pager strategy for grids whose pager renders as PURE page
    NUMBERS with no Next/Prev arrow controls at all (e.g. a pager showing
    just "1  2  3  4   Page size: [30 v]   96 items in 4 pages" with no
    separate Next/">" element). Looks for a clickable page-number element
    whose text is exactly str(current_page_num + 1) and clicks it. Returns
    True if a click happened, False if there's no such page (i.e. we're
    already on the last page)."""
    target_text = str(current_page_num + 1)
    try:
        candidates = search_root.find_elements(
            By.XPATH,
            ".//a[normalize-space(text())='%s'] | "
            ".//span[normalize-space(text())='%s' and "
            "(contains(@onclick,'age') or contains(@class,'age'))]"
            % (target_text, target_text)
        )
        for el in candidates:
            cls = (el.get_attribute("class") or "").lower()
            # Skip anything that's actually the page-SIZE dropdown/list
            # (e.g. an option literally reading "4") rather than a page
            # NUMBER link in the pager.
            if "pagesize" in cls or el.tag_name.lower() == "option":
                continue
            try:
                el.click()
            except Exception:
                try:
                    driver.execute_script("arguments[0].click();", el)
                except Exception:
                    continue
            return True
    except Exception:
        pass
    return False


def _wait_grid_row_count_stable(driver, grid_container_id=ID_RESULTS_GRID_CONTAINER,
                                 timeout=6, stable_polls=2, poll_interval=0.3):
    """FIX (last-page cropping): after a postback swaps in a new page,
    RadGrid can attach its <tr> row elements in more than one paint pass
    instead of all at once. The page-advance check in scrape_result_rows()
    only waits for the row SET to differ from the previous page at all --
    which is satisfied the instant the FIRST new row attaches, not once
    every row for that page has attached.

    This is invisible on a full (e.g. 30-row) page because the grid is
    already mid-flight when the first row lands and the rest usually
    follow within the same tick. It becomes visible on the LAST page of a
    result set, which is typically SHORTER than a full page (e.g. page 7
    of 7 holding only 20 of the normal 30 rows): the very first of those
    20 rows attaching is enough to satisfy "different from before", so the
    scraper can end up reading the grid after only a handful of that
    page's rows have actually painted -- silently cropping the last page
    short of its real row count.

    Polls the live row count on `grid_container_id` until it holds steady
    across `stable_polls` consecutive polls (grid has finished painting),
    or gives up after `timeout` seconds (grid was just genuinely slow;
    caller proceeds with whatever is there rather than hanging forever).
    """
    try:
        containers = driver.find_elements(By.ID, grid_container_id)
        root = containers[0] if containers else driver
    except Exception:
        root = driver

    deadline = time.time() + timeout
    last_count = -1
    stable_hits = 0
    while time.time() < deadline:
        try:
            count = len(root.find_elements(
                By.XPATH,
                ".//tr[contains(@class,'rgRow') or contains(@class,'rgAltRow')]"
            ))
        except Exception:
            count = last_count
        if count == last_count:
            stable_hits += 1
            if stable_hits >= stable_polls:
                return
        else:
            stable_hits = 0
        last_count = count
        time.sleep(poll_interval)


def scrape_result_rows(driver, timeout=ELEMENT_TIMEOUT, max_pages=200, log=None):
    if log is None:
        log = lambda *a, **k: None
    results = []
    try:
        WebDriverWait(driver, timeout).until(
            lambda d: d.find_elements(By.ID, ID_RESULTS_GRID_CONTAINER)
            and d.find_element(By.ID, ID_RESULTS_GRID_CONTAINER).find_elements(By.TAG_NAME, "table")
        )
    except TimeoutException:
        return results

    # Same "don't read mid-paint" protection as page-advances (see
    # _wait_grid_row_count_stable) applied to the very first page too,
    # since a short FIRST page (e.g. a search with under 30 total results,
    # so page 1 is also the last/only page) can race exactly the same way.
    _wait_grid_row_count_stable(driver, timeout=5)

    stale_retries = 0
    MAX_STALE_RETRIES = 5
    page_num = 1

    expected_total, expected_pages = _get_grid_expected_total(driver)

    for _ in range(max_pages):
        try:
            containers = driver.find_elements(By.ID, ID_RESULTS_GRID_CONTAINER)
            search_root = containers[0] if containers else driver
            tables = search_root.find_elements(By.TAG_NAME, "table")

            target_table, header_keys, header_row_el = None, None, None
            for t in tables:
                all_trs = t.find_elements(By.XPATH, ".//tr")
                if not all_trs:
                    continue
                # Identify the header row by CONTENT (looking for "CASE #"),
                # not by assuming it's always tr[1]. Telerik RadGrid
                # sometimes renders an extra <tr> (a filter row, a hidden
                # group row, etc.) ahead of the real header in document
                # order. If we blindly treat "tr[1]" as the header and skip
                # it, that extra row gets skipped instead of the real
                # header, and the true header row then gets counted as a
                # body row — which pushes every genuine data row down by
                # one and causes the actual first data row (the one on top
                # of the visible grid) to be silently dropped, no matter
                # which county is searched.
                for tr in all_trs:
                    cells = tr.find_elements(By.XPATH, ".//th | .//td")
                    texts = [c.text.strip().upper() for c in cells]
                    if "CASE #" in texts or any("CASE" in t2 for t2 in texts):
                        target_table = t
                        header_row_el = tr
                        header_keys = [_HEADER_KEY_MAP.get(t2, t2.lower().replace(" ", "_")) for t2 in texts]
                        break
                if target_table is not None:
                    break

            if target_table is None:
                break

            # Exclude ONLY the specific <tr> element we identified as the
            # header (matched by its Selenium element id), rather than by
            # position. This guarantees every real data row is kept even
            # if the header isn't literally the first <tr> in the DOM.
            all_trs = target_table.find_elements(By.XPATH, ".//tr")
            if header_row_el is not None:
                body_rows = [tr for tr in all_trs if tr.id != header_row_el.id and not _is_pager_row(tr)]
            else:
                body_rows = [tr for tr in all_trs[1:] if not _is_pager_row(tr)]

            if not body_rows and not results:
                # The grid's header rendered but no data rows are there yet —
                # this can happen right after an AJAX postback where the
                # header binds a beat before the row data does. Give it one
                # short extra look before concluding the grid is truly empty.
                time.sleep(1.0)
                all_trs = target_table.find_elements(By.XPATH, ".//tr")
                if header_row_el is not None:
                    body_rows = [tr for tr in all_trs if tr.id != header_row_el.id and not _is_pager_row(tr)]
                else:
                    body_rows = [tr for tr in all_trs[1:] if not _is_pager_row(tr)]

            # FIX (page 7 / BUTTS-county bug): RadGrid can occasionally
            # render a SAME-PAGE duplicate -- two rows on the very page
            # currently on screen sharing the identical case number/RECID
            # -- in place of what should have been two DIFFERENT rows
            # (confirmed live: page 7 rendered case 26T0314 twice back to
            # back, and the case number that belonged in that first slot,
            # 26T0313, never appeared anywhere in the result set at all).
            # This is a stale/partial re-render, not a legitimate repeat,
            # so retry re-reading THIS SAME page (no navigation) a few
            # times before accepting it, giving the grid a chance to
            # finish painting the real distinct row.
            PAGE_DUPLICATE_RETRIES = 3
            page_rows = []
            page_dupes = []
            for attempt in range(1, PAGE_DUPLICATE_RETRIES + 1):
                page_rows = []
                for r in body_rows:
                    cells = r.find_elements(By.TAG_NAME, "td")
                    if not cells:
                        continue
                    # Defensive second check: even if a pager row slipped
                    # past _is_pager_row() (e.g. an unexpected skin/markup
                    # variant), never let a cell whose text matches the
                    # pager pattern (or is just the page-number sequence)
                    # get recorded as a case number.
                    values = [c.text.strip() for c in cells]
                    if len(values) == 1 and _PAGER_TEXT_RE.search(values[0]):
                        continue
                    values = (values + [""] * len(header_keys))[:len(header_keys)]
                    values = [strip_junk_chars(v) for v in values]
                    row_dict = dict(zip(header_keys, values))
                    if "offender" in row_dict:
                        row_dict["offender"] = clean_name_value(row_dict["offender"])
                    if row_dict.get("case_number") and not _PAGER_TEXT_RE.search(row_dict["case_number"]):
                        page_rows.append(row_dict)

                page_dupes, _ = find_duplicate_case_numbers(
                    [row.get("case_number", "") for row in page_rows]
                )
                if not page_dupes:
                    break

                log(f"    [WARNING] Page {page_num}: duplicate case "
                    f"number(s) rendered within this SAME page "
                    f"({', '.join(page_dupes)}) -- likely a stale/partial "
                    f"re-render masking a different row. Re-reading this "
                    f"page (attempt {attempt}/{PAGE_DUPLICATE_RETRIES})...")
                time.sleep(1.0)
                _wait_grid_row_count_stable(driver, timeout=5)
                try:
                    all_trs = target_table.find_elements(By.XPATH, ".//tr")
                    if header_row_el is not None:
                        body_rows = [tr for tr in all_trs if tr.id != header_row_el.id and not _is_pager_row(tr)]
                    else:
                        body_rows = [tr for tr in all_trs[1:] if not _is_pager_row(tr)]
                except Exception:
                    break

            if page_dupes:
                log(f"    [WARNING] Page {page_num}: duplicate case "
                    f"number(s) persisted after {PAGE_DUPLICATE_RETRIES} "
                    f"re-read attempt(s): {', '.join(page_dupes)}. A "
                    f"different case number was likely dropped in its "
                    f"place -- please spot-check this page against the "
                    f"live site.")

            results.extend(page_rows)

            # NOTE: on this site (georgiaprobaterecords.com) the Next/Last
            # page pager controls render as <input type="button"
            # class="rgPageNext"> / rgPageLast, NOT as <a> tags. The old
            # XPath only ever looked for an <a class="rgPageNext">, so it
            # never found this site's Next button, next_candidates came
            # back empty every time, and the loop below broke out after
            # page 1 (only ~30 rows), silently discarding every remaining
            # page (e.g. pages 2-7 of a 7-page / page-size-30 result set).
            # Matching the <input> variant fixes that.
            # Snapshot of the WHOLE current page's row content (not just the
            # first row) — used below to confirm the postback really swapped
            # in a new page. Comparing every row (not only row[0]) avoids a
            # false "unchanged" read on the rare page boundary where two
            # consecutive pages happen to share the same first case number
            # text, and avoids a false "changed" read from a half-rendered
            # partial repaint (which would only touch some rows).
            prev_rows_snapshot = tuple(
                r.get_attribute("outerHTML") for r in body_rows
            ) if body_rows else tuple()

            # FIX (root cause of "N items across M pages, but only X rows
            # collected" + "removed Y duplicate case numbers" appearing
            # TOGETHER on the same county): the old code clicked Next/click-
            # numeric ONCE, waited once for the grid's first row to change,
            # and if that wait timed out it just `pass`-ed through and still
            # incremented page_num. Two different failure modes then hid
            # behind that single silent `pass`:
            #   1) The click never actually registered (stale/intercepted
            #      element, slow site, or fired mid-postback) -> the grid
            #      never advances, so the NEXT loop pass re-scrapes the
            #      exact same page it just scraped -> duplicate case
            #      numbers.
            #   2) page_num was incremented even though the grid didn't
            #      move -> if the arrow Next control is gone on a later pass
            #      and the code falls back to the pure-numeric pager, it
            #      targets str(page_num + 1), which is now off-by-one from
            #      the page actually on screen -> an entire real page of
            #      rows gets skipped, never scraped at all.
            # Together those two bugs produce exactly the symptom seen in
            # the log: fewer rows collected than the grid reports AND
            # duplicate rows removed, in the same county, at the same time.
            #
            # This is worse on a slow/loaded live site, where a single
            # ELEMENT_TIMEOUT-length wait can legitimately not be enough.
            #
            # Fix, two layers:
            #  (a) After clicking, wait on ASP.NET's own PageRequestManager
            #      flag (wait_for_ajax_idle) FIRST — this is the site
            #      telling us the postback actually finished, rather than
            #      us guessing from polling the DOM.
            #  (b) THEN confirm the grid's row content actually changed
            #      (whole-page snapshot, not just row 0). Only trust and
            #      advance page_num once both signals agree. Retry the
            #      click+wait cycle several times, with a growing timeout
            #      budget each attempt, before giving up.
            NAV_RETRIES = 5
            advanced = False
            for nav_attempt in range(1, NAV_RETRIES + 1):
                attempt_timeout = timeout + (nav_attempt - 1) * 5  # 20,25,30,35,40s

                next_candidates = search_root.find_elements(
                    By.XPATH,
                    ".//input[contains(@class,'rgPageNext') and not(contains(@class,'Disabled')) and not(@disabled)] | "
                    ".//a[contains(@class,'rgPageNext') and not(contains(@class,'Disabled'))] | "
                    ".//a[normalize-space(text())='Next'] | .//a[normalize-space(text())='>']"
                )

                # IMPORTANT (confirmed from real site HTML, "37 items in 2
                # pages" case): on the LAST page, this site's Next button
                # keeps class="rgPageNext" — no "Disabled" class, no
                # disabled attribute at all. The ONLY difference is its
                # onclick handler is prefixed with "return false;" (e.g.
                # onclick="return false;__doPostBack('...ctl10','')"),
                # which silently blocks the postback client-side. The
                # enabled Next button's onclick has no such prefix
                # (onclick="javascript:__doPostBack('...ctl10','')").
                # Without filtering these out, next_candidates never comes
                # up empty on the last page.
                next_candidates = [
                    el for el in next_candidates
                    if not (el.get_attribute("onclick") or "").strip().startswith("return false;")
                ]

                if next_candidates:
                    try:
                        next_candidates[0].click()
                    except Exception:
                        try:
                            driver.execute_script("arguments[0].click();", next_candidates[0])
                        except Exception:
                            pass
                    clicked = True
                else:
                    # No usable Next/Prev arrow control — try the pure-
                    # numeric pager fallback ("1  2  3  4   Page size: ...
                    # 96 items in 4 pages"). page_num here always reflects
                    # the last CONFIRMED page (see below), so the target
                    # text stays correct even after a retry.
                    clicked = _click_next_numeric_page(search_root, driver, page_num)

                if not clicked:
                    # No usable Next/Prev control at all — genuinely the
                    # last page.
                    break

                # (a) Wait for the ASP.NET AJAX postback itself to finish —
                # the site's own ground-truth signal, not a DOM guess.
                wait_for_ajax_idle(driver, timeout=attempt_timeout)

                # (b) Confirm the grid's row content actually changed
                # (whole-page snapshot comparison).
                try:
                    WebDriverWait(driver, attempt_timeout).until(
                        lambda d: (
                            d.find_elements(By.ID, ID_RESULTS_GRID_CONTAINER)
                            and d.find_element(By.ID, ID_RESULTS_GRID_CONTAINER)
                                .find_elements(By.XPATH, ".//tr[contains(@class,'rgRow') or contains(@class,'rgAltRow')]")
                            and (
                                not prev_rows_snapshot or
                                tuple(
                                    tr.get_attribute("outerHTML")
                                    for tr in d.find_element(By.ID, ID_RESULTS_GRID_CONTAINER)
                                        .find_elements(By.XPATH, ".//tr[contains(@class,'rgRow') or contains(@class,'rgAltRow')]")
                                ) != prev_rows_snapshot
                            )
                        )
                    )
                    advanced = True
                    break
                except TimeoutException:
                    # Click may not have registered, or fired before the
                    # previous postback fully settled, or the site is just
                    # slow. Let the AJAX queue drain and retry the click
                    # (with a longer budget next attempt) rather than
                    # silently assuming we moved on.
                    log(f"    [INFO] Page {page_num} -> {page_num + 1} not "
                        f"confirmed on attempt {nav_attempt}/{NAV_RETRIES} "
                        f"(timeout {attempt_timeout}s) — retrying.")
                    wait_for_ajax_idle(driver, timeout=5)
                    time.sleep(1.0)
                    continue

            if not advanced:
                log(f"    [WARNING] Could not confirm the grid advanced past "
                    f"page {page_num} after {NAV_RETRIES} attempt(s) — "
                    f"stopping pagination here.")
                break

            page_num += 1

            # FIX (last-page cropping): don't trust the grid is fully
            # painted just because SOME row differs from the previous page
            # (see _wait_grid_row_count_stable docstring). This matters
            # most on a short final page, where the old single fixed
            # time.sleep(0.5)/(0.6) could fire before all of that page's
            # rows had actually attached, so the NEXT loop pass would read
            # (and scrape) an incomplete row set for the last page.
            _wait_grid_row_count_stable(driver, timeout=5)
            time.sleep(0.3)

        except StaleElementReferenceException:
            stale_retries += 1
            if stale_retries > MAX_STALE_RETRIES:
                break
            wait_for_ajax_idle(driver, timeout=3)
            time.sleep(0.3)
            continue
        except Exception:
            break

    if expected_total is not None and len(results) != expected_total:
        log(f"    [WARNING] Grid reported {expected_total} item(s) across "
            f"{expected_pages} page(s), but only {len(results)} row(s) were "
            f"collected — pagination may have stopped early.")

    # De-duplicate by case number. The RadGrid can occasionally render the
    # same row twice across a paging cycle (e.g. a partial AJAX re-render
    # that overlaps with the tail of the previous page, or a slow postback
    # that gets read before AND after it settles), which otherwise inflates
    # the Output workbook with repeated case numbers. Keep the first
    # occurrence of each case number and drop the rest.
    seen_case_numbers = set()
    deduped_results = []
    duplicate_count = 0
    duplicated_case_nos = []
    for row_dict in results:
        cn = row_dict.get("case_number", "").strip()
        if cn in seen_case_numbers:
            duplicate_count += 1
            duplicated_case_nos.append(cn)
            continue
        seen_case_numbers.add(cn)
        deduped_results.append(row_dict)

    if duplicate_count:
        log(f"    [INFO] Removed {duplicate_count} duplicate case number "
            f"row(s) (same case rendered more than once during paging): "
            f"{', '.join(sorted(set(duplicated_case_nos)))}")
        # FIX (page 7 / BUTTS-county bug): a duplicated row usually means a
        # DIFFERENT, adjacent case number was silently dropped in its
        # place -- simply removing the duplicate and moving on (the old
        # behavior) hides that loss completely. Surface the best-effort
        # guess at what's actually missing so it doesn't go unnoticed.
        likely_missing = infer_likely_missing_case_numbers(
            [r.get("case_number", "") for r in results]
        )
        if likely_missing:
            for prefix, nums in likely_missing.items():
                log(f"    [WARNING] Possible MISSING case number(s) (gap "
                    f"in the {prefix}#### sequence, likely dropped when a "
                    f"duplicate rendered in its place): {', '.join(nums)}")

    return deduped_results


# ─────────────────────────────────────────────────────────────────────────
# OFFLINE PAGINATION VERIFIER — no live browser needed
#
# Replays the SAME header-detection / pager-row-exclusion logic as
# scrape_result_rows() above, but against saved static HTML (e.g. the
# debug_<county>.html dumps this tool already writes to the Offline
# folder on a failed/0-row run, or a manual Save-As of each result page).
# Lets you confirm "did page 2 get cropped?" for a past run WITHOUT
# needing Selenium/Chrome or re-hitting the live site.
# ─────────────────────────────────────────────────────────────────────────
def _html_is_pager_row(tr):
    cls = " ".join(tr.get("class", []) or []).lower()
    if "pager" in cls:
        return True
    tds = tr.find_all("td")
    if len(tds) == 1:
        td = tds[0]
        td_cls = " ".join(td.get("class", []) or []).lower()
        colspan = td.get("colspan")
        txt = td.get_text(" ", strip=True)
        if "pager" in td_cls or _PAGER_TEXT_RE.search(txt):
            return True
        if colspan and colspan.isdigit() and int(colspan) > 1 and _PAGER_TEXT_RE.search(txt):
            return True
    return False


def _parse_offline_page(path, grid_container_id=ID_RESULTS_GRID_CONTAINER):
    """Parses one saved HTML page and returns (case_numbers, expected_total,
    expected_pages). expected_total/pages come from that page's own
    'N items in M pages' rgInfoPart summary, or (None, None) if absent."""
    with open(path, "r", encoding="utf-8", errors="replace") as f:
        html = f.read()
    soup = BeautifulSoup(html, "lxml")
    container = soup.find(id=grid_container_id)
    search_root = container if container else soup

    expected_total, expected_pages = None, None
    for el in search_root.find_all("div", class_="rgInfoPart"):
        m = _GRID_TOTAL_RE.search(el.get_text(" ", strip=True))
        if m:
            expected_total, expected_pages = int(m.group(1)), int(m.group(2))
            break

    tables = search_root.find_all("table")
    target_table, header_row_el = None, None
    for t in tables:
        all_trs = t.find_all("tr")
        if not all_trs:
            continue
        for tr in all_trs:
            cells = tr.find_all(["th", "td"])
            texts = [c.get_text(" ", strip=True).upper() for c in cells]
            if "CASE #" in texts or any("CASE" in t2 for t2 in texts):
                target_table, header_row_el = t, tr
                break
        if target_table is not None:
            break

    if target_table is None:
        return [], expected_total, expected_pages

    all_trs = target_table.find_all("tr")
    body_rows = [tr for tr in all_trs if tr is not header_row_el and not _html_is_pager_row(tr)]

    case_numbers = []
    for r in body_rows:
        cells = r.find_all("td")
        if not cells:
            continue
        values = [c.get_text(" ", strip=True) for c in cells]
        if len(values) == 1 and _PAGER_TEXT_RE.search(values[0]):
            continue
        case_no = values[0] if values else ""
        if case_no and not _PAGER_TEXT_RE.search(case_no):
            case_numbers.append(case_no)
    return case_numbers, expected_total, expected_pages


# ------------------------------------------------------------------
# OFFLINE "LIST" FILE RENAMER
#
# Manually-saved search-result pages (e.g. via a browser "Save page as")
# can show up in TWO forms:
#   1. With the date range tacked on (needs renaming):
#        COUNTY_1-6-2026_1-12-2026_LIST_1.html  ...  COUNTY_1-6-2026_1-12-2026_LIST_50.html
#   2. Already short, no date data tacked on (already correct, left alone):
#        COUNTY_LIST_1.html  ...  COUNTY_LIST_50.html
# This renames every file in form 1 down to form 2 (dropping the date
# range, normalizing "List" -> "LIST"), and leaves files already in
# form 2 untouched instead of flagging them as unrecognized. Handles
# list numbers 1 through 100.
# ------------------------------------------------------------------
_OFFLINE_LIST_NAME_RE = re.compile(
    r"^(?P<county>[A-Za-z]+)_"
    r"\d{1,2}-\d{1,2}-\d{4}_\d{1,2}-\d{1,2}-\d{4}_"
    r"List_(?P<num>\d{1,3})$",
    re.IGNORECASE,
)

# Already in the target "no date data" form, e.g. "COUNTY_LIST_7".
_OFFLINE_LIST_ALREADY_DONE_RE = re.compile(
    r"^(?P<county>[A-Za-z]+)_LIST_(?P<num>\d{1,3})$",
    re.IGNORECASE,
)


def build_offline_list_name(filepath):
    """Given an offline list filepath (or bare filename), return the new
    '<COUNTY>_LIST_<N>.html' name, or None if it doesn't match the
    '<COUNTY>_<start>_<end>_List_<N>' pattern, or N is outside 1-100."""
    base = os.path.basename(filepath)
    stem, ext = os.path.splitext(base)
    m = _OFFLINE_LIST_NAME_RE.match(stem)
    if not m:
        return None
    num = int(m.group("num"))
    if not (1 <= num <= 100):
        return None
    county = m.group("county").upper()
    return f"{county}_LIST_{num}{ext.lower()}"


def is_already_offline_list_name(filepath):
    """True if filepath is already in the short '<COUNTY>_LIST_<N>' form
    (no date data tacked on) -- nothing to rename."""
    stem = os.path.splitext(os.path.basename(filepath))[0]
    m = _OFFLINE_LIST_ALREADY_DONE_RE.match(stem)
    if not m:
        return False
    return 1 <= int(m.group("num")) <= 100


def is_offline_list_filename(filepath):
    """True if filepath is a search-results LIST page (e.g.
    'BUTTS_6-9-2026_6-15-2026_List_3.html' or the already-renamed
    'BUTTS_LIST_3.html') rather than a single-case offline file
    ('<COUNTY>_<CASE#>.html'). These LIST pages hold a whole page of
    citation rows, not one case's fields, so they must NOT be fed into
    the single-case extractor -- doing so misreads the date/number
    portion of the filename as a Case #, which is what was causing the
    spurious 'Mismatch (Case #, County)' Row Status entries."""
    stem = os.path.splitext(os.path.basename(filepath))[0]
    return bool(_OFFLINE_LIST_NAME_RE.match(stem) or _OFFLINE_LIST_ALREADY_DONE_RE.match(stem))


def rename_offline_list_files(folder, dry_run=False):
    """Renames every offline list HTML file in `folder` from
    '<COUNTY>_<start>-<end>_List_<N>.html' to '<COUNTY>_LIST_<N>.html'.
    Files already in the '<COUNTY>_LIST_<N>.html' form are left alone
    and are NOT reported as skipped/unmatched.
    Returns a list of (old_name, new_name) pairs that were (or would be)
    renamed, and a list of (name, reason) for files that genuinely don't
    match either recognized pattern."""
    renamed, skipped = [], []
    files = sorted(
        glob.glob(os.path.join(folder, "*.html")) + glob.glob(os.path.join(folder, "*.htm"))
    )
    for fp in files:
        old_name = os.path.basename(fp)

        if is_already_offline_list_name(fp):
            continue  # already correctly named -- nothing to do

        new_name = build_offline_list_name(fp)
        if new_name is None:
            skipped.append((old_name, "does not match '<COUNTY>_<start>_<end>_List_<N>' pattern"))
            continue
        if new_name == old_name:
            continue
        new_path = os.path.join(folder, new_name)
        if os.path.exists(new_path):
            skipped.append((old_name, f"target '{new_name}' already exists"))
            continue
        if not dry_run:
            os.rename(fp, new_path)
        renamed.append((old_name, new_name))
    return renamed, skipped


def verify_offline_html_pages(paths):
    """Runs the offline pagination check across one or more saved HTML
    pages belonging to the SAME search result set. Returns (ok, report_text).

    ok=True  -> total rows extracted across all supplied pages matches the
                grid's own reported total, and no case number repeats
                across pages (i.e. nothing was cropped or double-read).
    ok=False -> mismatch or duplicates found; report_text explains why.
    """
    if not BS4_AVAILABLE:
        return False, ("BeautifulSoup is required for this check.\n"
                        "Install it with:  pip install beautifulsoup4 lxml")
    if not paths:
        return False, "No HTML file(s) selected."

    lines = []
    all_case_numbers = []
    grid_total, grid_pages = None, None

    for path in paths:
        try:
            rows, exp_total, exp_pages = _parse_offline_page(path)
        except Exception as e:
            lines.append(f"{os.path.basename(path)}: ERROR reading/parsing file ({e})")
            continue
        summary = f"  (grid reports: {exp_total} items in {exp_pages} pages)" if exp_total else ""
        lines.append(f"{os.path.basename(path)}: {len(rows)} row(s) extracted{summary}")
        if exp_total is not None:
            grid_total, grid_pages = exp_total, exp_pages
        all_case_numbers.extend(rows)

    total_extracted = len(all_case_numbers)
    unique_case_numbers = set(all_case_numbers)
    dupe_count = total_extracted - len(unique_case_numbers)
    dupe_case_nos, _ = find_duplicate_case_numbers(all_case_numbers)

    lines.append("-" * 50)
    lines.append(f"TOTAL extracted across {len(paths)} page(s): {total_extracted}")
    if dupe_count:
        lines.append(f"Unique case numbers: {len(unique_case_numbers)}  "
                      f"({dupe_count} DUPLICATE(S) found across pages)")
        lines.append(f"  Duplicated case number(s): {', '.join(dupe_case_nos)}")
        # A duplicate render is a strong signal that some OTHER case
        # number was silently dropped in its place (see
        # infer_likely_missing_case_numbers docstring) -- surface the
        # best-effort guess right next to the duplicate that caused it.
        likely_missing = infer_likely_missing_case_numbers(all_case_numbers)
        if likely_missing:
            lines.append("  [LIKELY MISSING] A duplicate usually means a different case "
                         "number was dropped in its place. Gap(s) found in the numbered "
                         "sequence:")
            for prefix, nums in likely_missing.items():
                lines.append(f"    {prefix}...: {', '.join(nums)}")
    else:
        lines.append(f"Unique case numbers: {len(unique_case_numbers)}")

    ok = True
    if grid_total is not None:
        lines.append(f"Grid's own reported total: {grid_total} items in {grid_pages} pages")
        if total_extracted != grid_total:
            lines.append(f"[MISMATCH] Extracted {total_extracted} but grid says {grid_total} "
                          f"-- pagination may have stopped early / a page was cropped.")
            ok = False
        else:
            lines.append("[OK] Extracted count matches the grid's reported total.")
    else:
        lines.append("(No 'N items in M pages' summary found on any supplied page -- "
                      "can't cross-check total. Include the page that shows the pager row.)")

    if dupe_count:
        lines.append(f"[WARNING] {dupe_count} duplicate case number(s) across the supplied pages "
                      f"-- could mean the same page was read twice, overlapping page boundaries, "
                      f"OR (as seen on page 7 / BUTTS county) a stale grid re-render that duplicated "
                      f"one row and silently dropped a different, adjacent case number.")
        ok = False

    lines.append("-" * 50)
    lines.append("RESULT: PASS" if ok else "RESULT: FAIL")
    return ok, "\n".join(lines)


# ─────────────────────────────────────────────────────────────────────────
# Excel styling
# ─────────────────────────────────────────────────────────────────────────
HEADER_FONT  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
CELL_FONT    = Font(name="Arial", size=10)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN         = Side(style="thin", color="D9D9D9")
CELL_BORDER  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
NORMAL_FILL  = PatternFill("solid", fgColor="FFFFFF")
ALT_FILL     = PatternFill("solid", fgColor="F2F6FA")

STATUS_FONT = {
    "Match":                   Font(name="Arial", size=10, color="1E7B34"),
    "New (scraped only)":      Font(name="Arial", size=10, bold=True, color="9C5700"),
    "Missing (existing only)": Font(name="Arial", size=10, bold=True, color="C00000"),
}
STATUS_FILL = {
    "Match":                   PatternFill("solid", fgColor="E2F0D9"),
    "New (scraped only)":      PatternFill("solid", fgColor="FFF2CC"),
    "Missing (existing only)": PatternFill("solid", fgColor="FCE4E4"),
}


def build_output_path(output_folder, tag=""):
    """Saves a flat file directly in output_folder (e.g.
    D:\\Mohan\\GA_Probate\\Input Folder\\result_07212026_live.xlsx) —
    no dated subfolder is created."""
    stamp = datetime.now().strftime("%m%d%Y")
    ensure_folder(output_folder)
    suffix = f"_{tag}" if tag else ""
    return os.path.join(output_folder, f"result_{stamp}{suffix}.xlsx")


def build_match_status_path(tool_input_folder, tag=""):
    ensure_folder(tool_input_folder)
    stamp = datetime.now().strftime("%m%d%Y")
    suffix = f"_{tag}" if tag else ""
    return os.path.join(tool_input_folder, f"{MATCH_STATUS_FILE_PREFIX}{stamp}{suffix}.xlsx")


def write_combined_workbook(county_results, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Case Results"

    headers = ["Case #", "Offender", "Offense Date", "Court Date", "Disposed", "County"]
    row_keys = ["case_number", "offender", "offense_date", "court_date", "disposed", "county"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font, c.fill, c.alignment, c.border = HEADER_FONT, HEADER_FILL, CENTER_ALIGN, CELL_BORDER
    ws.freeze_panes = "A2"

    row_idx = 2
    toggle = False
    for r in county_results:
        if r["status"] != "ok" or not r["rows"]:
            continue
        toggle = not toggle
        fill = ALT_FILL if toggle else NORMAL_FILL
        for row_dict in r["rows"]:
            values = [row_dict.get(k, "") for k in row_keys]
            if not values[-1]:
                values[-1] = r["county"]
            for ci, v in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=ci, value=v)
                cell.font, cell.fill, cell.border, cell.alignment = CELL_FONT, fill, CELL_BORDER, CENTER_ALIGN
            row_idx += 1

    for ci, w in enumerate([14, 26, 14, 14, 14, 14], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    if row_idx > 2:
        ws.auto_filter.ref = f"A1:F{row_idx - 1}"

    ws2 = wb.create_sheet("Run Summary")
    summary_headers = ["County", "Starting Date", "Ending Date", "Status", "Rows Found"]
    for ci, h in enumerate(summary_headers, 1):
        c = ws2.cell(row=1, column=ci, value=h)
        c.font, c.fill, c.alignment, c.border = HEADER_FONT, HEADER_FILL, CENTER_ALIGN, CELL_BORDER
    for ri, r in enumerate(county_results, 2):
        values = [r["county"], r["start_date"], r["end_date"], r["status"], len(r["rows"])]
        for ci, v in enumerate(values, 1):
            cell = ws2.cell(row=ri, column=ci, value=v)
            cell.font, cell.border, cell.alignment = CELL_FONT, CELL_BORDER, CENTER_ALIGN
    for ci, w in enumerate([18, 14, 14, 12, 12], 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w

    wb.save(output_path)
    return output_path


def find_compare_file(tool_input_folder, explicit_path=None, log=print):
    if explicit_path:
        if os.path.isfile(explicit_path):
            return explicit_path
        log(f"⚠ Compare file was given but not found: {explicit_path}")
        return None

    if not os.path.isdir(tool_input_folder):
        log(f"⚠ Tool Input folder does not exist, skipping validation: {tool_input_folder}")
        return None

    candidates = [
        os.path.join(tool_input_folder, f)
        for f in os.listdir(tool_input_folder)
        if f.lower().endswith(".xlsx") and not f.startswith(MATCH_STATUS_FILE_PREFIX)
    ]
    if not candidates:
        log(f"⚠ No existing case-number workbook (.xlsx) found in {tool_input_folder} — "
            f"skipping match/mismatch validation.")
        return None
    if len(candidates) > 1:
        candidates.sort(key=os.path.getmtime, reverse=True)
        log(f"⚠ Multiple candidate case-number workbooks found — using most recent: "
            f"{os.path.basename(candidates[0])}")
    return candidates[0]


def read_existing_case_numbers(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
    case_col_idx = None
    for idx, h in enumerate(header_row):
        if h is None:
            continue
        norm = re.sub(r"[^A-Z]", "", str(h).upper())
        if norm in ("CASE", "CASENO", "CASENUMBER", "CASENUM"):
            case_col_idx = idx
            break
    if case_col_idx is None:
        raise RuntimeError(
            f"Could not find a 'Case #' column in {path}. Header row read as: {header_row}")

    numbers = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if case_col_idx < len(row) and row[case_col_idx] not in (None, ""):
            numbers.add(str(row[case_col_idx]).strip())
    return numbers


def write_match_status_workbook(existing_numbers, scraped_numbers, compare_file_path, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Match Status"

    headers = ["Case #", "Status"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font, c.fill, c.alignment, c.border = HEADER_FONT, HEADER_FILL, CENTER_ALIGN, CELL_BORDER
    ws.freeze_panes = "A2"

    all_numbers = sorted(existing_numbers | scraped_numbers)
    row_idx = 2
    match_count = new_count = missing_count = 0
    for case_no in all_numbers:
        in_existing = case_no in existing_numbers
        in_scraped = case_no in scraped_numbers
        if in_existing and in_scraped:
            status = "Match"
            match_count += 1
        elif in_scraped and not in_existing:
            status = "New (scraped only)"
            new_count += 1
        else:
            status = "Missing (existing only)"
            missing_count += 1

        c1 = ws.cell(row=row_idx, column=1, value=case_no)
        c2 = ws.cell(row=row_idx, column=2, value=status)
        for c in (c1, c2):
            c.font, c.border, c.alignment = STATUS_FONT[status], CELL_BORDER, CENTER_ALIGN
            c.fill = STATUS_FILL[status]
        row_idx += 1

    for ci, w in enumerate([26, 24], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    if row_idx > 2:
        ws.auto_filter.ref = f"A1:B{row_idx - 1}"

    ws2 = wb.create_sheet("Validation Summary")
    ws2["A1"], ws2["B1"] = "Compared against", compare_file_path
    ws2["A2"], ws2["B2"] = "Existing case numbers", len(existing_numbers)
    ws2["A3"], ws2["B3"] = "Scraped case numbers", len(scraped_numbers)
    ws2["A4"], ws2["B4"] = "Match", match_count
    ws2["A5"], ws2["B5"] = "New (scraped only)", new_count
    ws2["A6"], ws2["B6"] = "Missing (existing only)", missing_count
    for r in range(1, 7):
        ws2.cell(row=r, column=1).font = Font(bold=True, name="Arial", size=10)
        ws2.cell(row=r, column=2).font = CELL_FONT
    ws2.column_dimensions["A"].width = 26
    ws2.column_dimensions["B"].width = 60

    wb.save(output_path)

    report_rows = [{"Case Number": n,
                     "Status": (
                         "Match" if (n in existing_numbers and n in scraped_numbers) else
                         "New (scraped only)" if n in scraped_numbers else
                         "Missing (existing only)")}
                    for n in all_numbers]
    report_df = pd.DataFrame(report_rows) if report_rows else pd.DataFrame(
        columns=["Case Number", "Status"])

    return output_path, match_count, new_count, missing_count, report_df


# ─────────────────────────────────────────────────────────────────────────
# LIVE SEARCH pipeline (Selenium, drives the real site)
# ─────────────────────────────────────────────────────────────────────────
def run_live_pipeline(cfg, log, on_done):
    """cfg keys: input_file, output_folder, offline_folder, tool_input_folder,
    delay_seconds_minimum, delay_seconds_maximum, headless, limit"""
    try:
        if not SELENIUM_AVAILABLE:
            raise RuntimeError(
                "selenium is not installed. Install it with:  pip install selenium")

        log("=" * 62)
        log("  LNGA Probate Courts — Live Search")
        log("=" * 62)
        log(f"  Input Excel     : {cfg['input_file']}")
        log(f"  Output folder   : {cfg['output_folder']}")
        log(f"  Offline folder  : {cfg['offline_folder']}")
        log(f"  Tool Input      : {cfg['tool_input_folder']}")

        county_rows = load_input_file(cfg["input_file"])
        limit = cfg.get("limit") or 0
        if limit:
            county_rows = county_rows[:limit]
        if not county_rows:
            raise RuntimeError(f"No county rows parsed from: {cfg['input_file']}")

        log(f"  Counties to process : {len(county_rows)}")
        log(f"  Browser mode        : {'headless' if cfg.get('headless', True) else 'visible'}\n")

        ensure_folder(cfg["offline_folder"])
        driver = build_driver(headless=cfg.get("headless", True))
        county_results = []
        MAX_ATTEMPTS = 3
        delay_min = float(cfg.get("delay_seconds_minimum", 0) or 0)
        delay_max = float(cfg.get("delay_seconds_maximum", 0) or 0)

        try:
            for i, row in enumerate(county_rows, 1):
                county, start_date, end_date = row["county"], row["start_date"], row["end_date"]
                log("-" * 60)
                log(f"[{i}/{len(county_rows)}] {county}  ({start_date} -> {end_date})")

                last_error = None
                result_rows, succeeded = [], False
                for attempt in range(1, MAX_ATTEMPTS + 1):
                    try:
                        driver.get(SEARCH_URL)
                        select_county(driver, county)
                        set_offense_date_range(driver, start_date, end_date)
                        click_search(driver)
                        set_page_size(driver, target_size=30)
                        wait_for_ajax_idle(driver, timeout=8)
                        result_rows = scrape_result_rows(driver, log=log)

                        if not result_rows and attempt < MAX_ATTEMPTS:
                            # A clean 0-row scrape is ambiguous: it could be a
                            # genuinely empty result set, or the grid's AJAX
                            # postback simply hadn't finished rendering data
                            # rows yet when we read the table. Give it one
                            # more real search attempt (with an extra settle
                            # pause) before accepting 0 as the final answer,
                            # instead of trusting the first read.
                            log(f"  [WARNING] Attempt {attempt}/{MAX_ATTEMPTS}: 0 rows scraped — "
                                f"retrying in case the grid was still loading...")
                            time.sleep(1.5)
                            continue

                        succeeded = True
                        break
                    except StaleElementReferenceException as e:
                        last_error = e
                        log(f"  [WARNING] Attempt {attempt}/{MAX_ATTEMPTS}: stale element — retrying...")
                        time.sleep(1.0)
                        continue
                    except Exception as e:
                        last_error = e
                        break

                if succeeded:
                    log(f"  {len(result_rows)} row(s) found")
                    if len(result_rows) == 0:
                        dbg_path = os.path.join(cfg["offline_folder"], f"debug_{county}_zero_rows.html")
                        try:
                            patched = patch_debug_html_for_offline_viewing(driver.page_source)
                            with open(dbg_path, "w", encoding="utf-8") as f:
                                f.write(patched)
                            log(f"    (0 rows — page HTML saved to {os.path.basename(dbg_path)} for troubleshooting)")
                        except Exception:
                            pass
                    county_results.append({"county": county, "start_date": start_date, "end_date": end_date,
                                            "rows": result_rows, "status": "ok"})
                else:
                    log(f"  [ERROR] {county} failed ({type(last_error).__name__}): {last_error}")
                    dbg_path = os.path.join(cfg["offline_folder"], f"debug_{county}.html")
                    try:
                        patched = patch_debug_html_for_offline_viewing(driver.page_source)
                        with open(dbg_path, "w", encoding="utf-8") as f:
                            f.write(patched)
                        log(f"    (page HTML saved to {os.path.basename(dbg_path)} for troubleshooting)")
                    except Exception:
                        pass
                    county_results.append({"county": county, "start_date": start_date, "end_date": end_date,
                                            "rows": [], "status": f"failed: {type(last_error).__name__}: {last_error}"})

                if delay_max > 0 and i < len(county_rows):
                    import random
                    time.sleep(random.uniform(delay_min, delay_max))
        finally:
            driver.quit()

        _finish_pipeline(county_results, cfg, log, on_done, tag="live")

    except Exception:
        log(f"\n[FATAL ERROR]\n{traceback.format_exc()}")
        on_done(False, None, None, {})


def _finish_pipeline(county_results, cfg, log, on_done, tag):
    """Shared tail end for both pipelines: write the combined Output
    workbook, then the Match/New/Missing Result workbook against whatever
    existing case-number file is sitting in the Tool Input / Compare folder."""
    ok = sum(1 for r in county_results if r["status"] == "ok")
    failed = len(county_results) - ok
    total_cases = sum(len(r["rows"]) for r in county_results)
    log("\n" + "=" * 62)
    log(f"SUMMARY: {ok} OK   {failed} failed   {total_cases} row(s) total")
    log("=" * 62)

    output_path = build_output_path(cfg["output_folder"], tag=tag)
    write_combined_workbook(county_results, output_path)
    log(f"\nCombined workbook saved -> {output_path}")

    scraped_numbers = {
        row_dict["case_number"].strip()
        for r in county_results if r["status"] == "ok"
        for row_dict in r["rows"]
        if row_dict.get("case_number")
    }

    compare_file = find_compare_file(cfg["tool_input_folder"],
                                      explicit_path=cfg.get("compare_file"), log=log)
    stats = {"total": total_cases, "match": 0, "new": 0, "missing": 0,
              "report_df": pd.DataFrame(columns=["Case Number", "Status"])}

    match_path = None
    if compare_file:
        try:
            existing_numbers = read_existing_case_numbers(compare_file)
            match_path = build_match_status_path(cfg["tool_input_folder"], tag=tag)
            _, matched, new, missing, report_df = write_match_status_workbook(
                existing_numbers, scraped_numbers, compare_file, match_path)
            log(f"Result (match status) workbook saved -> {match_path}")
            log(f"  Match: {matched}   New (scraped only): {new}   Missing (existing only): {missing}")
            stats.update({"match": matched, "new": new, "missing": missing, "report_df": report_df})
        except Exception as e:
            log(f"[ERROR] Validation step failed ({type(e).__name__}): {e}")
    else:
        log("[WARNING] Skipped match/mismatch validation — no existing case-number "
            "workbook (.xlsx with a 'Case #' column) found in the Tool Input / "
            "Compare folder to check today's scrape against.")
        if scraped_numbers:
            try:
                baseline_path = os.path.join(cfg["tool_input_folder"], "LNGA_CaseNumber_Baseline.xlsx")
                wb = Workbook()
                ws = wb.active
                ws.title = "Case Numbers"
                ws.cell(row=1, column=1, value="Case #").font = Font(bold=True, name="Arial", size=10)
                for ri, num in enumerate(sorted(scraped_numbers), 2):
                    ws.cell(row=ri, column=1, value=num).font = CELL_FONT
                ws.column_dimensions["A"].width = 26
                ensure_folder(cfg["tool_input_folder"])
                wb.save(baseline_path)
                log(f"  Seeded baseline case-number workbook -> {baseline_path}")
                log(f"  ({len(scraped_numbers)} case number(s) — future runs will validate "
                    f"new scrapes against this file. Replace it any time with a more "
                    f"authoritative existing list if you have one.)")
            except Exception as e:
                log(f"  [WARNING] Could not write baseline workbook: {e}")

    log("\n" + "=" * 62)
    log("  COMPLETE")
    log("=" * 62)

    on_done(True, output_path, match_path, stats)


# ══════════════════════════════════════════════════════
#  GUI
# ══════════════════════════════════════════════════════
class App(tk.Toplevel):

    def __init__(self, master=None):
        super().__init__(master)
        self.title("LNGA Probate Courts — Citation Search  (GUI)")
        self.geometry("1280x800")
        self.minsize(1020, 660)
        self.configure(bg=BG_MAIN)

        # Shared folders
        self.var_input   = tk.StringVar(value="")
        self.var_output  = tk.StringVar(value=OUTPUT_FOLDER)
        self.var_offline = tk.StringVar(value=OFFLINE_FOLDER)
        self.var_tool_input = tk.StringVar(value=TOOL_INPUT_FOLDER)

        # Live-search-only options
        self.var_delay_min = tk.StringVar(value="1")
        self.var_delay_max = tk.StringVar(value="3")
        self.var_limit      = tk.StringVar(value="0")
        self.var_visible    = tk.BooleanVar(value=False)

        self.status_var = tk.StringVar(value="Idle")
        self.cnt_total   = tk.StringVar(value="—")
        self.cnt_match   = tk.StringVar(value="—")
        self.cnt_new     = tk.StringVar(value="—")
        self.cnt_missing = tk.StringVar(value="—")

        self._result_df          = None
        self._filter              = "all"
        self._last_output_path    = None
        self._last_match_path     = None
        self._log_queue           = queue.Queue()

        # running-time tracking
        self.run_start_time       = None
        self._elapsed_running     = False
        self._elapsed_job         = None

        self._build_ui()
        self._auto_detect_input()
        self._poll_log_queue()

    def _auto_detect_input(self):
        """On launch, if the Input Excel field is empty, try to auto-fill it
        with the most recently modified .xlsx in the GA_Probate Input folder
        — reduces the chance of the Browse dialog opening somewhere
        unrelated (e.g. a different tool's folder) and the wrong file
        getting picked."""
        if self.var_input.get().strip():
            return
        try:
            candidate = find_latest_input_file(INPUT_FOLDER, label="Input folder")
            self.var_input.set(candidate)
        except Exception:
            pass

    # ── BUILD UI ─────────────────────────────────────────

    def _build_ui(self):
        s = ttk.Style(self)
        s.theme_use("clam")
        s.configure("Treeview", font=("Segoe UI", 9), rowheight=24,
                    background=BG_CARD, fieldbackground=BG_CARD, foreground=TEXT_PRI)
        s.configure("Treeview.Heading", font=("Segoe UI", 9, "bold"),
                    background="#F0F4FA", foreground=TEXT_SEC, relief="flat")
        s.map("Treeview", background=[("selected", "#DBEAFE")])
        s.configure("bar.Horizontal.TProgressbar",
                    troughcolor=BORDER_GUI, background=ACCENT, thickness=4)

        outer = tk.Frame(self, bg=BG_MAIN)
        outer.pack(fill="both", expand=True)

        sidebar = tk.Frame(outer, bg=BG_SIDEBAR, width=320)
        sidebar.pack(side="left", fill="y")
        sidebar.pack_propagate(False)
        self._build_sidebar(sidebar)

        main = tk.Frame(outer, bg=BG_MAIN)
        main.pack(side="left", fill="both", expand=True)
        self._build_main(main)

    # ── SIDEBAR ──────────────────────────────────────────

    def _build_sidebar(self, parent):
        hdr = tk.Frame(parent, bg=BG_SIDEBAR, pady=18, padx=16)
        hdr.pack(fill="x")
        tk.Label(hdr, text="🔎", bg=BG_SIDEBAR, fg="#93C5FD",
                 font=("Segoe UI", 24)).pack(anchor="w")
        tk.Label(hdr, text="Probate Citation\nSearch", bg=BG_SIDEBAR,
                 fg="#FFFFFF", font=("Segoe UI", 13, "bold"),
                 justify="left").pack(anchor="w", pady=(4, 0))
        tk.Label(hdr, text="Georgia Probate Records  ·  no login required",
                 bg=BG_SIDEBAR, fg="#93C5FD",
                 font=("Segoe UI", 8)).pack(anchor="w")

        self._divider(parent)

        bf = tk.Frame(parent, bg=BG_SIDEBAR, padx=14, pady=5)
        bf.pack(fill="x")
        tk.Label(bf, text="Sets Input File + Output + Tool Input all at once",
                 bg=BG_SIDEBAR, fg="#64A0D4", font=("Segoe UI", 7),
                 wraplength=270, justify="left").pack(anchor="w")
        tk.Button(bf, text="📁  Select Base Folder (fills all 3 below)",
                  bg="#276221", fg="#FFFFFF", font=("Segoe UI", 8, "bold"),
                  relief="flat", pady=6, cursor="hand2",
                  command=self._browse_base_folder).pack(fill="x", pady=(4, 0))

        self._path_field(parent, "Input File (.txt or .xlsx)", self.var_input,
                          self._browse_input, "County / Starting Date / Ending Date")
        self._path_field(parent, "Output Folder", self.var_output,
                          self._browse_output, "result_MMDDYYYY_*.xlsx saved here")
        # "Offline Folder" field is hidden from this tool's UI -- debug HTML
        # dumps and the Verify/Rename Offline utilities below still use
        # self.var_offline internally (defaulted to OFFLINE_FOLDER / the
        # base-folder "Offline" subfolder if found), it's just no longer a
        # visible/editable row here.
        self._path_field(parent, "Tool Input / Compare Folder", self.var_tool_input,
                          self._browse_tool_input, "Existing case list + Result workbook")

        self._divider(parent)

        tk.Label(parent, text="LIVE SEARCH OPTIONS", bg=BG_SIDEBAR, fg="#64A0D4",
                 font=("Segoe UI", 8, "bold"), padx=16).pack(anchor="w")

        df = tk.Frame(parent, bg=BG_SIDEBAR, padx=14, pady=5)
        df.pack(fill="x")
        tk.Label(df, text="Delay Between Counties (sec)", bg=BG_SIDEBAR, fg="#E2E8F0",
                 font=("Segoe UI", 8, "bold")).pack(anchor="w")
        row = tk.Frame(df, bg=BG_SIDEBAR)
        row.pack(fill="x", pady=(3, 0))
        tk.Entry(row, textvariable=self.var_delay_min, width=6, bg="#253F70", fg="#E2E8F0",
                 insertbackground="#E2E8F0", font=("Consolas", 9), relief="flat", bd=4
                 ).pack(side="left")
        tk.Label(row, text=" to ", bg=BG_SIDEBAR, fg="#94A3B8",
                 font=("Segoe UI", 8)).pack(side="left")
        tk.Entry(row, textvariable=self.var_delay_max, width=6, bg="#253F70", fg="#E2E8F0",
                 insertbackground="#E2E8F0", font=("Consolas", 9), relief="flat", bd=4
                 ).pack(side="left")

        lf = tk.Frame(parent, bg=BG_SIDEBAR, padx=14, pady=5)
        lf.pack(fill="x")
        tk.Label(lf, text="Limit Counties (0 = all)", bg=BG_SIDEBAR, fg="#E2E8F0",
                 font=("Segoe UI", 8, "bold")).pack(anchor="w")
        tk.Entry(lf, textvariable=self.var_limit, width=6, bg="#253F70", fg="#E2E8F0",
                 insertbackground="#E2E8F0", font=("Consolas", 9), relief="flat", bd=4
                 ).pack(anchor="w", pady=(3, 0))

        cf = tk.Frame(parent, bg=BG_SIDEBAR, padx=14, pady=5)
        cf.pack(fill="x")
        tk.Checkbutton(cf, text="Show browser window (visible, not headless)",
                        variable=self.var_visible, bg=BG_SIDEBAR, fg="#E2E8F0",
                        selectcolor="#253F70", activebackground=BG_SIDEBAR,
                        activeforeground="#E2E8F0", font=("Segoe UI", 8),
                        wraplength=270, justify="left").pack(anchor="w")

        self._divider(parent)

        sf = tk.Frame(parent, bg=BG_SIDEBAR, padx=16)
        sf.pack(fill="x", pady=(0, 4))
        tk.Label(sf, text="STATUS", bg=BG_SIDEBAR, fg="#64A0D4",
                 font=("Segoe UI", 8, "bold")).pack(anchor="w")
        self._status_lbl = tk.Label(sf, textvariable=self.status_var,
                                    bg=BG_SIDEBAR, fg="#93C5FD",
                                    font=("Segoe UI", 10, "bold"))
        self._status_lbl.pack(anchor="w")

        self.elapsed_var = tk.StringVar(value="00:00:00")
        tk.Label(sf, text="RUNNING TIME", bg=BG_SIDEBAR, fg="#64A0D4",
                 font=("Segoe UI", 8, "bold")).pack(anchor="w", pady=(8, 0))
        tk.Label(sf, textvariable=self.elapsed_var, bg=BG_SIDEBAR, fg="#93C5FD",
                 font=("Consolas", 12, "bold")).pack(anchor="w")

        self._divider(parent)

        tk.Label(parent, text="RESULT (LATEST RUN)", bg=BG_SIDEBAR, fg="#64A0D4",
                 font=("Segoe UI", 8, "bold"), padx=16).pack(anchor="w")

        counters = [
            ("Case #s Scraped", self.cnt_total,   "#93C5FD"),
            ("Match",           self.cnt_match,   "#6EE7B7"),
            ("New (scraped)",   self.cnt_new,     "#FDE68A"),
            ("Missing",         self.cnt_missing, "#FCA5A5"),
        ]
        for lbl, var, color in counters:
            mf = tk.Frame(parent, bg="#253F70", padx=12, pady=3)
            mf.pack(fill="x", padx=14, pady=2)
            r = tk.Frame(mf, bg="#253F70")
            r.pack(fill="x")
            tk.Label(r, text=lbl, bg="#253F70", fg="#94A3B8",
                     font=("Segoe UI", 8)).pack(side="left")
            tk.Label(r, textvariable=var, bg="#253F70", fg=color,
                     font=("Segoe UI", 13, "bold")).pack(side="right")

        self._divider(parent)

        self._open_output_btn = tk.Button(
            parent, text="📂  Open Output Workbook",
            command=lambda: self._open_path(self._last_output_path),
            bg="#253F70", fg="#93C5FD",
            font=("Segoe UI", 8, "bold"), relief="flat", pady=6,
            cursor="hand2", state="disabled",
        )
        self._open_output_btn.pack(fill="x", padx=14, pady=2)

        self._open_match_btn = tk.Button(
            parent, text="🗂  Open Result Workbook",
            command=lambda: self._open_path(self._last_match_path),
            bg="#253F70", fg="#94A3B8",
            font=("Segoe UI", 8), relief="flat", pady=4,
            cursor="hand2", state="disabled",
        )
        self._open_match_btn.pack(fill="x", padx=14, pady=(0, 6))

        self._divider(parent)

        tk.Button(
            parent, text="🔍  Verify Offline HTML (Debug Pages)",
            command=self._verify_offline_html,
            bg="#253F70", fg="#93C5FD",
            font=("Segoe UI", 8, "bold"), relief="flat", pady=6,
            cursor="hand2",
        ).pack(fill="x", padx=14, pady=(0, 2))
        tk.Label(
            parent,
            text="Select saved page(s) (e.g. debug_<county>.html) to\n"
                 "check pagination for cropped rows — no live search needed.",
            bg=BG_SIDEBAR, fg="#64A0D4", font=("Segoe UI", 7),
            justify="left",
        ).pack(fill="x", padx=14, pady=(0, 8))

        tk.Button(
            parent, text="🏷  Rename Offline List Files",
            command=self._rename_offline_list_files,
            bg="#253F70", fg="#93C5FD",
            font=("Segoe UI", 8, "bold"), relief="flat", pady=6,
            cursor="hand2",
        ).pack(fill="x", padx=14, pady=(0, 2))
        tk.Label(
            parent,
            text="Renames files like 'BUTTS_1-6-2026_1-12-2026_List_1.html'\n"
                 "to 'BUTTS_LIST_1.html' (drops the date range, LIST_1-100)\n"
                 "in the Offline Folder above.",
            bg=BG_SIDEBAR, fg="#64A0D4", font=("Segoe UI", 7),
            justify="left",
        ).pack(fill="x", padx=14, pady=(0, 8))

    def _rename_offline_list_files(self):
        folder = self.var_offline.get().strip()
        if not folder or not os.path.isdir(folder):
            messagebox.showerror("Rename Offline List Files",
                                  "Set a valid Offline Folder first.")
            return

        renamed, skipped = rename_offline_list_files(folder)

        self._reset_log()
        self._log_ts("=" * 20 + " RENAME OFFLINE LIST FILES " + "=" * 20)
        for old_name, new_name in renamed:
            self._log_ts(f"  {old_name}  ->  {new_name}")
        for name, reason in skipped:
            self._log_ts(f"  SKIPPED {name}: {reason}")

        summary = f"Renamed {len(renamed)} file(s)."
        if skipped:
            summary += f"\nSkipped {len(skipped)} file(s) (see log)."
        messagebox.showinfo("Rename Offline List Files", summary)

    def _verify_offline_html(self):
        if not BS4_AVAILABLE:
            messagebox.showerror(
                "Missing Dependency",
                "This check needs BeautifulSoup.\n\nInstall it with:\n"
                "  pip install beautifulsoup4 lxml")
            return

        initial_dir = self.var_offline.get().strip() or SCRIPT_DIR
        paths = filedialog.askopenfilenames(
            title="Select saved HTML page(s) from the SAME search (e.g. debug_<county>.html)",
            initialdir=initial_dir if os.path.isdir(initial_dir) else SCRIPT_DIR,
            filetypes=[("HTML files", "*.html;*.htm"), ("All files", "*.*")],
        )
        if not paths:
            return

        ok, report = verify_offline_html_pages(list(paths))

        self._reset_log()
        self._log_ts("=" * 20 + " OFFLINE PAGINATION CHECK " + "=" * 20)
        for line in report.splitlines():
            self._log_ts(line)

        if ok:
            messagebox.showinfo("Verify Offline HTML — PASS ✓",
                                 "No cropping detected.\n\n" + report)
        else:
            messagebox.showwarning("Verify Offline HTML — FAIL ✗",
                                    "Possible cropping / pagination issue found.\n\n" + report)

    def _path_field(self, parent, label, var, browse_cmd, hint):
        fr = tk.Frame(parent, bg=BG_SIDEBAR, padx=14, pady=5)
        fr.pack(fill="x")
        tk.Label(fr, text=label, bg=BG_SIDEBAR, fg="#E2E8F0",
                 font=("Segoe UI", 8, "bold")).pack(anchor="w")
        tk.Label(fr, text=hint, bg=BG_SIDEBAR, fg="#64A0D4",
                 font=("Segoe UI", 7)).pack(anchor="w")
        row = tk.Frame(fr, bg=BG_SIDEBAR)
        row.pack(fill="x", pady=(3, 0))
        tk.Entry(row, textvariable=var, bg="#253F70", fg="#E2E8F0",
                 insertbackground="#E2E8F0", font=("Consolas", 7),
                 relief="flat", bd=4).pack(side="left", fill="x", expand=True)
        tk.Button(row, text="Browse…", bg="#2D5499", fg="#FFFFFF",
                  font=("Segoe UI", 8), relief="flat", padx=8,
                  command=browse_cmd).pack(side="left", padx=(4, 0))

    def _divider(self, parent):
        tk.Frame(parent, bg="#2D5499", height=1).pack(fill="x", padx=16, pady=6)

    # ── MAIN AREA ────────────────────────────────────────

    def _build_main(self, parent):
        rb = tk.Frame(parent, bg=BG_MAIN, padx=20, pady=14)
        rb.pack(fill="x")

        btn_row = tk.Frame(rb, bg=BG_MAIN)
        btn_row.pack(fill="x")

        self._run_live_btn = tk.Button(
            btn_row, text="▶  Run Live Search + Validation",
            command=self._start_live, bg=ACCENT, fg="#FFFFFF",
            font=("Segoe UI", 11, "bold"), relief="flat",
            pady=10, cursor="hand2",
            activebackground=ACCENT_HOVER, activeforeground="#FFFFFF",
        )
        self._run_live_btn.pack(side="left", fill="x", expand=True)

        self._progress = ttk.Progressbar(rb, style="bar.Horizontal.TProgressbar",
                                         mode="indeterminate")
        self._progress.pack(fill="x", pady=(8, 0))

        nb_frame = tk.Frame(parent, bg=BG_MAIN, padx=20, pady=4)
        nb_frame.pack(fill="both", expand=True)
        nb = ttk.Notebook(nb_frame)
        nb.pack(fill="both", expand=True)

        log_tab = tk.Frame(nb, bg=BG_LOG)
        nb.add(log_tab, text="  Run Log  ")
        self._log_widget = tk.Text(log_tab, bg=BG_LOG, fg="#D1FAE5",
                                   font=("Consolas", 9), relief="flat",
                                   wrap="none", state="disabled",
                                   insertbackground="#D1FAE5", bd=8)
        sc = tk.Scrollbar(log_tab, command=self._log_widget.yview)
        self._log_widget.configure(yscrollcommand=sc.set)
        sc.pack(side="right", fill="y")
        self._log_widget.pack(fill="both", expand=True)
        for tag, color in [("ok", LOG_OK), ("warn", LOG_WARN),
                            ("err", LOG_ERR), ("info", LOG_INFO)]:
            self._log_widget.tag_config(tag, foreground=color)

        res_tab = tk.Frame(nb, bg=BG_CARD)
        nb.add(res_tab, text="  Results  ")
        self._build_results(res_tab)

    def _build_results(self, parent):
        fbar = tk.Frame(parent, bg=BG_CARD, pady=8, padx=8)
        fbar.pack(fill="x")
        self._filter_btns = {}
        filters = [
            ("All",     "all",                      ACCENT,    "#FFFFFF"),
            ("Match",   "Match",                     "#276221", "#FFFFFF"),
            ("New",     "New (scraped only)",        "#9C5700", "#FFFFFF"),
            ("Missing", "Missing (existing only)",   "#9C0006", "#FFFFFF"),
        ]
        for label, key, bg_on, fg_on in filters:
            btn = tk.Button(fbar, text=label,
                            font=("Segoe UI", 9),
                            bg=bg_on if key == "all" else BORDER_GUI,
                            fg=fg_on if key == "all" else TEXT_SEC,
                            relief="flat", padx=10, pady=4,
                            command=lambda k=key, b=bg_on, f=fg_on:
                                self._set_filter(k, b, f))
            btn.pack(side="left", padx=2)
            self._filter_btns[key] = (btn, bg_on, fg_on)

        self._row_count_var = tk.StringVar(value="")
        tk.Label(fbar, textvariable=self._row_count_var,
                 bg=BG_CARD, fg=TEXT_SEC,
                 font=("Segoe UI", 8)).pack(side="right", padx=8)

        cols = ("Case Number", "Status")
        self._tree = ttk.Treeview(parent, columns=cols, show="headings", selectmode="browse")
        widths = {"Case Number": 220, "Status": 220}
        for col in cols:
            self._tree.heading(col, text=col, command=lambda c=col: self._sort_col(c))
            self._tree.column(col, width=widths.get(col, 150), anchor="center")

        self._tree.tag_configure("Match", background=C_GREEN_BG, foreground=C_GREEN)
        self._tree.tag_configure("New (scraped only)", background=C_ORANGE_BG, foreground=C_ORANGE)
        self._tree.tag_configure("Missing (existing only)", background=C_RED_BG, foreground=C_RED)

        vsb = ttk.Scrollbar(parent, orient="vertical",   command=self._tree.yview)
        hsb = ttk.Scrollbar(parent, orient="horizontal", command=self._tree.xview)
        self._tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        hsb.pack(side="bottom", fill="x")
        vsb.pack(side="right",  fill="y")
        self._tree.pack(fill="both", expand=True)

        self._sort_state = {}

    # ── BROWSE HANDLERS ──────────────────────────────────

    def _browse_base_folder(self):
        """One-click setup: pick a single base folder (e.g.
        D:\\Mohan\\GA_Probate\\New) and derive all four paths from it,
        matching the fixed layout:
            <base>                  -> Input File (auto-detected inside it)
            <base>\\Input Folder     -> Output Folder
            <base>\\Offline          -> Offline Folder
            <base>\\Tool_Input       -> Tool Input / Compare Folder
        Missing subfolders are created automatically."""
        start_dir = self.var_input.get().strip()
        start_dir = os.path.dirname(start_dir) if os.path.isfile(start_dir) else (start_dir or os.getcwd())
        base = filedialog.askdirectory(
            title="Select Base Folder (e.g. D:\\Mohan\\GA_Probate\\New)",
            initialdir=start_dir if os.path.isdir(start_dir) else os.getcwd(),
        )
        if not base:
            return
        base = os.path.normpath(base)

        output_folder     = os.path.join(base, "Input Folder")
        offline_folder     = os.path.join(base, "Offline")
        tool_input_folder = os.path.join(base, "Tool_Input")

        try:
            ensure_folder(output_folder)
            ensure_folder(offline_folder)
            ensure_folder(tool_input_folder)
        except Exception as e:
            messagebox.showerror("Folder Error", f"Could not create subfolders under:\n{base}\n\n{e}")
            return

        self.var_output.set(output_folder)
        self.var_offline.set(offline_folder)
        self.var_tool_input.set(tool_input_folder)

        try:
            candidate = find_latest_input_file(base, label="Base folder")
            self.var_input.set(candidate)
        except Exception as e:
            self.var_input.set("")
            messagebox.showwarning(
                "Input File Not Auto-Detected",
                f"Set Output/Offline/Tool Input folders under:\n{base}\n\n"
                f"But couldn't auto-detect a County/Starting Date/Ending Date "
                f"input file directly inside that folder:\n\n{e}\n\n"
                f"Use the Input File \u201cBrowse\u2026\u201d button to pick it manually.")

    def _browse_input(self):
        # Default to the GA_Probate Input folder (not wherever the last
        # dialog happened to be left, e.g. a different tool's folder) so
        # it's much harder to accidentally pick a file from another tool
        # that just happens to also be an .xlsx.
        start_dir = self.var_input.get().strip()
        start_dir = os.path.dirname(start_dir) if start_dir else ""
        if not start_dir or not os.path.isdir(start_dir):
            start_dir = INPUT_FOLDER if os.path.isdir(INPUT_FOLDER) else SCRIPT_DIR
        path = filedialog.askopenfilename(
            title="Select Input File (columns: County, Starting Date, Ending Date)",
            initialdir=start_dir,
            filetypes=[("Input files", "*.txt *.xlsx *.xls"),
                       ("Text files", "*.txt"),
                       ("Excel files", "*.xlsx *.xls"),
                       ("All files", "*.*")]
        )
        if path:
            self.var_input.set(os.path.normpath(path))

    def _browse_output(self):
        p = filedialog.askdirectory(title="Select Output Folder")
        if p:
            self.var_output.set(os.path.normpath(p))

    def _browse_offline(self):
        p = filedialog.askdirectory(title="Select Offline Folder")
        if p:
            self.var_offline.set(os.path.normpath(p))

    def _browse_tool_input(self):
        p = filedialog.askdirectory(title="Select Tool Input / Compare Folder")
        if p:
            self.var_tool_input.set(os.path.normpath(p))

    def _open_path(self, path):
        if path and os.path.isfile(path):
            try:
                os.startfile(path)
            except Exception:
                messagebox.showinfo("File Location", path)

    # ── FILTER / SORT / RENDER ───────────────────────────

    def _set_filter(self, key, bg_on, fg_on):
        self._filter = key
        for k, (btn, bg, fg) in self._filter_btns.items():
            active = (k == key)
            btn.configure(bg=bg if active else BORDER_GUI,
                          fg=fg if active else TEXT_SEC)
        self._render_table()

    def _sort_col(self, col):
        if self._result_df is None or self._result_df.empty:
            return
        ascending = not self._sort_state.get(col, False)
        self._sort_state[col] = ascending
        self._result_df = self._result_df.sort_values(col, ascending=ascending, ignore_index=True)
        self._render_table()

    def _render_table(self):
        for row in self._tree.get_children():
            self._tree.delete(row)
        if self._result_df is None or self._result_df.empty:
            self._row_count_var.set("")
            return
        df = self._result_df
        if self._filter != "all":
            df = df[df["Status"] == self._filter]
        self._row_count_var.set(f"{len(df)} row{'s' if len(df) != 1 else ''}")
        for _, row in df.iterrows():
            status = row.get("Status", "")
            self._tree.insert("", "end",
                               values=(row.get("Case Number", ""), status),
                               tags=(status,))

    # ── LOG QUEUE ────────────────────────────────────────

    def _log_ts(self, msg):
        self._log_queue.put(msg)

    def _poll_log_queue(self):
        while True:
            try:
                msg = self._log_queue.get_nowait()
            except queue.Empty:
                break
            tag = ""
            if "[FATAL" in msg or "[ERROR]" in msg:
                tag = "err"
            elif "COMPLETE" in msg:
                tag = "ok"
            elif msg.startswith("="):
                tag = "info"
            elif "[WARNING]" in msg:
                tag = "warn"
            ts = datetime.now().strftime("%H:%M:%S")
            self._log_widget.configure(state="normal")
            self._log_widget.insert("end", f"[{ts}]  {msg}\n", tag)
            self._log_widget.see("end")
            self._log_widget.configure(state="disabled")
        self.after(150, self._poll_log_queue)

    # ── VALIDATION ───────────────────────────────────────

    def _common_paths_ok(self):
        if not self.var_input.get().strip():
            messagebox.showwarning("Missing File", "Please select the Input File (.txt or .xlsx).")
            return False
        if not os.path.isfile(self.var_input.get().strip()):
            messagebox.showerror("File Not Found", f"Input Excel not found:\n{self.var_input.get().strip()}")
            return False
        if not self.var_output.get().strip() or not self.var_offline.get().strip() \
                or not self.var_tool_input.get().strip():
            messagebox.showwarning("Missing Folder", "Please set Output, Offline, and Tool Input folders.")
            return False
        return True

    def _cfg_from_ui(self):
        return {
            "input_file":            self.var_input.get().strip(),
            "output_folder":         self.var_output.get().strip(),
            "offline_folder":        self.var_offline.get().strip(),
            "tool_input_folder":     self.var_tool_input.get().strip(),
        }

    # ── RUN: LIVE ────────────────────────────────────────

    def _start_live(self):
        if not self._common_paths_ok():
            return
        try:
            delay_min = float(self.var_delay_min.get())
            delay_max = float(self.var_delay_max.get())
        except ValueError:
            messagebox.showwarning("Invalid Delay", "Delay values must be numbers.")
            return
        try:
            limit = int(self.var_limit.get() or 0)
        except ValueError:
            messagebox.showwarning("Invalid Limit", "Limit must be a whole number.")
            return

        cfg = self._cfg_from_ui()
        cfg.update({
            "delay_seconds_minimum": delay_min,
            "delay_seconds_maximum": delay_max,
            "limit": limit,
            "headless": not self.var_visible.get(),
        })

        self._reset_log()
        self._start_elapsed_timer()
        self._set_running(True)

        threading.Thread(
            target=run_live_pipeline,
            args=(cfg, self._log_ts, self._on_done),
            daemon=True,
        ).start()

    # ── SHARED RUN STATE ─────────────────────────────────

    def _reset_log(self):
        self._log_widget.configure(state="normal")
        self._log_widget.delete("1.0", "end")
        self._log_widget.configure(state="disabled")

    def _set_running(self, running):
        state = "disabled" if running else "normal"
        self._run_live_btn.configure(state=state)
        self._open_output_btn.configure(state="disabled")
        self._open_match_btn.configure(state="disabled")
        if running:
            self.status_var.set("Running (Live)…")
            self._status_lbl.configure(fg="#FDE68A")
            self._progress.start(12)
            for v in (self.cnt_total, self.cnt_match, self.cnt_new, self.cnt_missing):
                v.set("—")
        else:
            self._progress.stop()

    def _on_done(self, success, output_path, match_path, stats):
        self.after(0, lambda: self._finish(success, output_path, match_path, stats))

    # ---------------- running-time tracking ----------------

    @staticmethod
    def _format_elapsed(seconds):
        seconds = max(0, int(seconds))
        h, rem = divmod(seconds, 3600)
        m, s = divmod(rem, 60)
        return f"{h:02d}:{m:02d}:{s:02d}"

    def _tick_elapsed(self):
        if not self._elapsed_running or self.run_start_time is None:
            return
        elapsed = time.time() - self.run_start_time
        self.elapsed_var.set(self._format_elapsed(elapsed))
        self._elapsed_job = self.after(500, self._tick_elapsed)

    def _start_elapsed_timer(self):
        self.run_start_time = time.time()
        self._elapsed_running = True
        self.elapsed_var.set("00:00:00")
        self._tick_elapsed()

    def _stop_elapsed_timer(self):
        self._elapsed_running = False
        if self._elapsed_job is not None:
            try:
                self.after_cancel(self._elapsed_job)
            except Exception:
                pass
            self._elapsed_job = None
        if self.run_start_time is not None:
            final = self._format_elapsed(time.time() - self.run_start_time)
            self.elapsed_var.set(final)
            return final
        return "00:00:00"

    def _finish(self, success, output_path, match_path, stats):
        self._progress.stop()
        self._run_live_btn.configure(state="normal")
        elapsed_str = self._stop_elapsed_timer()

        if success:
            self._last_output_path = output_path
            self._last_match_path  = match_path
            self.status_var.set(f"Complete ✓  (Running time: {elapsed_str})")
            self._status_lbl.configure(fg=LOG_OK)

            total   = stats.get("total", 0)
            match   = stats.get("match", 0)
            new     = stats.get("new", 0)
            missing = stats.get("missing", 0)

            self.cnt_total.set(str(total))
            self.cnt_match.set(str(match))
            self.cnt_new.set(str(new))
            self.cnt_missing.set(str(missing))

            self._result_df = stats.get("report_df")
            self._render_table()
            if output_path:
                self._open_output_btn.configure(state="normal")
            if match_path:
                self._open_match_btn.configure(state="normal")

            messagebox.showinfo("Done ✓",
                f"Run complete!\n\n"
                f"Case #s scraped : {total}\n"
                f"Match           : {match}\n"
                f"New (scraped)   : {new}\n"
                f"Missing         : {missing}\n"
                f"Running time    : {elapsed_str}\n\n"
                f"Output workbook saved:\n  {os.path.basename(output_path) if output_path else '(none)'}")
        else:
            self.status_var.set(f"Error ✗  (Running time: {elapsed_str})")
            self._status_lbl.configure(fg=LOG_ERR)
            messagebox.showerror("Error", f"Run failed after {elapsed_str}.\nSee the Run Log tab for details.")


def open_citation_search_window(master=None):
    """Factory used by the combined-tool launcher to open this module's
    window as a Toplevel of the launcher's root window."""
    return App(master=master)


# ════════════════════════════════════════════════════════════════
#  COMBINED LAUNCHER
#  A single root window that opens either module in its own
#  Toplevel window. Both modules can be open (and running) together.
# ════════════════════════════════════════════════════════════════

class CombinedLauncher(tk.Tk):
    """Home screen for the merged LNGA Probate Courts tool."""

    def __init__(self):
        super().__init__()
        self.title("LNGA Probate Courts — Combined Tool")
        self.geometry("680x460")
        self.minsize(600, 420)
        self.configure(bg=BG_MAIN)

        self._citation_win = None
        self._validation_win = None

        tk.Label(self, text="LNGA Probate Courts", bg=BG_MAIN, fg=TEXT_PRI,
                 font=("Segoe UI", 20, "bold")).pack(pady=(36, 4))
        tk.Label(self, text="Combined Tool — choose a module to launch",
                 bg=BG_MAIN, fg=TEXT_SEC, font=("Segoe UI", 10)).pack(pady=(0, 28))

        card_frame = tk.Frame(self, bg=BG_MAIN)
        card_frame.pack(expand=True, fill="both", padx=40)

        self._make_launch_card(
            card_frame, "\U0001F50E  Citation Search",
            "Live-scrapes georgiaprobaterecords.com and builds the "
            "Output / Result workbooks. Shows running time while the "
            "search is in progress.",
            self._open_citation_search,
        ).pack(side="left", expand=True, fill="both", padx=(0, 12))

        self._make_launch_card(
            card_frame, "\u2705  Validation",
            "Compares offline HTML case files against the Output Excel "
            "and flags mismatches. Shows running time while validation "
            "is in progress.",
            self._open_validation,
        ).pack(side="left", expand=True, fill="both", padx=(12, 0))

        tk.Label(self, text="Each module opens in its own window and can be run independently — "
                             "you can even run both at the same time.",
                 bg=BG_MAIN, fg=TEXT_SEC, font=("Segoe UI", 8),
                 wraplength=600, justify="center").pack(pady=(20, 16))

    def _make_launch_card(self, parent, title, desc, command):
        card = tk.Frame(parent, bg=BG_CARD, highlightbackground=BORDER_GUI,
                         highlightthickness=1)
        tk.Label(card, text=title, bg=BG_CARD, fg=TEXT_PRI,
                 font=("Segoe UI", 13, "bold"), anchor="w",
                 justify="left", wraplength=240).pack(fill="x", padx=16, pady=(20, 8))
        tk.Label(card, text=desc, bg=BG_CARD, fg=TEXT_SEC,
                 font=("Segoe UI", 9), anchor="w", justify="left",
                 wraplength=240).pack(fill="x", padx=16, pady=(0, 16))
        btn = tk.Button(card, text="Open \u25b6", command=command, bg=ACCENT, fg="white",
                         activebackground=ACCENT_HOVER, activeforeground="white",
                         relief="flat", font=("Segoe UI", 10, "bold"), cursor="hand2", pady=8)
        btn.pack(fill="x", padx=16, pady=(0, 20))
        return card

    def _open_citation_search(self):
        if self._citation_win is not None and self._citation_win.winfo_exists():
            self._citation_win.deiconify()
            self._citation_win.lift()
            self._citation_win.focus_force()
            return
        self._citation_win = open_citation_search_window(master=self)
        self._citation_win.protocol("WM_DELETE_WINDOW", self._citation_win.destroy)

    def _open_validation(self):
        if self._validation_win is not None and self._validation_win.winfo_exists():
            self._validation_win.deiconify()
            self._validation_win.lift()
            self._validation_win.focus_force()
            return
        self._validation_win = open_validation_window(master=self)
        self._validation_win.protocol("WM_DELETE_WINDOW", self._validation_win.destroy)


def launch_combined_gui():
    app = CombinedLauncher()
    app.mainloop()


# ══════════════════════════════════════════════════════
#  ENTRY POINT
# ══════════════════════════════════════════════════════
if __name__ == "__main__":
    if "--cli" in sys.argv:
        # Headless Validation run only (Citation Search has no CLI mode) —
        # identical behavior to the original standalone validation script.
        main()
    else:
        try:
            import ctypes
            ctypes.windll.user32.ShowWindow(
                ctypes.windll.kernel32.GetConsoleWindow(), 0
            )
        except Exception:
            pass

        launch_combined_gui()
