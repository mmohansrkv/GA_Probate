"""
core.py — LNGA GA Probate Courts, headless core logic.

Ported without behavioral changes from LNGA_ProbateCourts_Combined_01.py:
all Tkinter GUI code (CitationSearchApp, ValidationApp, App) has been
removed. Every scraping / parsing / Excel-writing function below is the
same code that ran inside the desktop app, just called directly from
app.py (Flask) instead of from tkinter button callbacks.

Two entry points other modules should use:
    run_live_pipeline(cfg, log, on_done)   -> Citation Search (live scrape)
    run_validation(offline_dir, output_dir, input_dir, result_base_dir)
                                            -> Validation tool
"""

import os
import re
import sys
import time
import glob
import logging
import traceback
from datetime import datetime

import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from bs4 import BeautifulSoup
BS4_AVAILABLE = True

try:
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.common.keys import Keys
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.common.exceptions import (
        TimeoutException, NoSuchElementException, StaleElementReferenceException
    )
    SELENIUM_AVAILABLE = True
except ImportError:
    SELENIUM_AVAILABLE = False


SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

log = logging.getLogger("lnga_probate")
if not log.handlers:
    _h = logging.StreamHandler()
    _h.setFormatter(logging.Formatter("%(asctime)s %(levelname)s %(message)s"))
    log.addHandler(_h)
    log.setLevel(logging.INFO)

# ─────────────────────────────────────────────────────────────────────────
# CONFIG — defaults (all overridable by the caller; folders auto-created).
# On the web app these are just scratch/staging directories under /tmp or
# a mounted disk — app.py passes real per-request paths in, these are only
# used as fallbacks.
# ─────────────────────────────────────────────────────────────────────────
INPUT_FOLDER       = os.environ.get("LNGA_INPUT_FOLDER", "/tmp/lnga/Input")
OUTPUT_FOLDER       = os.environ.get("LNGA_OUTPUT_FOLDER", "/tmp/lnga/Output")
OFFLINE_FOLDER       = os.environ.get("LNGA_OFFLINE_FOLDER", "/tmp/lnga/Offline")
TOOL_INPUT_FOLDER   = os.environ.get("LNGA_TOOL_INPUT_FOLDER", "/tmp/lnga/Tool_Input")


OFFLINE_HTML_DIR = os.environ.get("LNGA_VALIDATION_OFFLINE_DIR", "/tmp/lnga/Validation/Offline")
OUTPUT_EXCEL_DIR = os.environ.get("LNGA_VALIDATION_OUTPUT_DIR", "/tmp/lnga/Validation/Output")
INPUT_EXCEL_DIR  = os.environ.get("LNGA_VALIDATION_INPUT_DIR", "/tmp/lnga/Validation/Input")
RESULT_BASE_DIR  = os.environ.get("LNGA_VALIDATION_RESULT_DIR", "/tmp/lnga/Validation")

MAX_ROW_INDEX = 50   # HTML citation rows are indexed _0 through _50 inclusive

# Field order -- this drives the exact column order in the result sheet,
# matching your Output Excel's header pattern.
FIELD_ORDER = [
    "Case #", "County", "Violation Date", "Court Date", "Name", "Address",
    "Race", "Sex", "Citation Number", "Code", "Description", "Fine",
    "Disposition", "Disposition Date", "Sentence",
]

# --- HTML tag (control id SUFFIX) map for single (non-repeating, case-level) fields ---
SINGLE_FIELD_SUFFIXES = {
    "County":          "_lblCourtName",
    "Case #":          "_lblCaseNo",
    "Violation Date":  "lblViolationDate",
    "Court Date":      "_lblCourtDate",
    "Name":            "_lblOffender",
    "Address":         "_lblCityStateZip",
    "Race":            "_lblRace",
    "Sex":             "_lblSex",
}

# --- HTML tag (control id PREFIX, index appended) map for repeating (per-citation) fields ---
ROW_FIELD_PREFIXES = {
    "Citation Number": "_lblCitationNo_",
    "Code":            "_lblCitCode_",
    "Description":     "lblCitCodeDesc_",
    "Fine":            "lblFineAmt_",
    "Disposition":     "_lblDispDesc_",
    "Disposition Date":"_lblDispDate_",
    "Sentence":        "_lblSentencing_",
}

SINGLE_FIELDS = set(SINGLE_FIELD_SUFFIXES.keys())
ROW_FIELDS = set(ROW_FIELD_PREFIXES.keys())

# --- Column header candidates to look for in the OUTPUT excel, for each field. ---
# The script first tries "<field> (Output)" (matching your exact header pattern);
# if that's not found it falls back to these synonyms as a bare column name.
OUTPUT_COLUMN_CANDIDATES = {
    "Case #":            ["case #", "case no", "case number"],
    "County":            ["county", "court name"],
    "Violation Date":    ["violation date"],
    "Court Date":        ["court date"],
    "Name":              ["name", "offender", "offender name"],
    "Address":           ["address", "city state zip"],
    "Race":              ["race"],
    "Sex":               ["sex", "gender"],
    "Citation Number":   ["citation number", "citation no"],
    "Code":              ["code", "cit code"],
    "Description":       ["description", "cit code desc"],
    "Fine":              ["fine", "fine amt", "fine amount"],
    "Disposition":       ["disposition", "disp desc"],
    "Disposition Date":  ["disposition date", "disp date"],
    "Sentence":          ["sentence", "sentencing"],
}

# --- Column header candidates to look for in the INPUT excel ---
INPUT_COLUMN_CANDIDATES = {
    "Case #": ["case #", "case no", "case number"],
    "County": ["county", "court name"],
}

# ─────────────────────────────────────────────────────────────────────────
# CONFIG — defaults (all overridable per-request from app.py, all auto-created)
# ─────────────────────────────────────────────────────────────────────────
SEARCH_URL = "https://georgiaprobaterecords.com/Traffic/SearchCitations.aspx"

MATCH_STATUS_FILE_PREFIX = "CaseNumber_MatchStatus_"

DEBUG_HTML_BASE_HREF = "https://georgiaprobaterecords.com/Traffic/"
PAGE_LOAD_TIMEOUT = 25
ELEMENT_TIMEOUT = 20


def ensure_folder(path):
    os.makedirs(path, exist_ok=True)
    return path


# ─────────────────────────────────────────────────────────────────────────
# DUPLICATE / LIKELY-MISSING CASE NUMBER DETECTION  (from Fixed_02)
#
# A duplicate found in the DATA is a signal that some OTHER case number is
# probably missing (the site's RadGrid can occasionally render a stale
# re-render that duplicates one row and silently drops a different,
# adjacent one).
#
# Case numbers on this site follow a PREFIX + zero-padded NUMBER pattern
# (e.g. "26T0313", "26T0314", ...). When duplicates are found, this helper
# looks at the run of numbers sharing each prefix and reports any gaps in
# that numeric sequence -- the most likely candidates for the case
# number(s) that got silently dropped.
# ─────────────────────────────────────────────────────────────────────────
_CASE_NO_SPLIT_RE = re.compile(r"^([A-Za-z0-9]*?)(\d+)$")


def _split_case_no(case_no):
    """Splits e.g. '26T0313' -> ('26T', '0313', 313). Returns None if the
    case number doesn't end in a run of digits."""
    m = _CASE_NO_SPLIT_RE.match((case_no or "").strip())
    if not m:
        return None
    prefix, digits = m.group(1), m.group(2)
    return prefix, digits, int(digits)


def find_duplicate_case_numbers(case_numbers):
    """Given a list of case-number strings (in the order encountered),
    returns (duplicates, first_seen_index) where `duplicates` is a sorted
    list of every case number that appeared more than once, and
    `first_seen_index` maps each case number to the row index of its
    FIRST occurrence (so callers can report exactly where in the run the
    duplicate showed up)."""
    seen_at = {}
    dupes = set()
    for i, cn in enumerate(case_numbers):
        cn = (cn or "").strip()
        if not cn:
            continue
        if cn in seen_at:
            dupes.add(cn)
        else:
            seen_at[cn] = i
    return sorted(dupes), seen_at


def infer_likely_missing_case_numbers(case_numbers):
    """Best-effort guess at which case number(s) were silently dropped when
    a duplicate is found. Groups all case numbers by their non-numeric
    prefix, looks at the min/max of the numeric suffix seen for that
    prefix, and reports any integer in that inclusive range which never
    appeared at all. This intentionally only reports gaps INSIDE the
    observed min..max span (not a guess about what might exist before/after
    it), since that's the span the tool actually has evidence for.

    Returns a dict: {prefix: [missing_case_no, ...], ...}. Empty dict if
    every prefix's numeric run is unbroken (or nothing parses as
    prefix+number).
    """
    by_prefix = {}
    for cn in case_numbers:
        parsed = _split_case_no(cn)
        if not parsed:
            continue
        prefix, digits, num = parsed
        by_prefix.setdefault(prefix, {"width": len(digits), "nums": set()})
        by_prefix[prefix]["nums"].add(num)

    missing = {}
    for prefix, info in by_prefix.items():
        nums = info["nums"]
        if len(nums) < 2:
            continue
        width = info["width"]
        lo, hi = min(nums), max(nums)
        gap = [n for n in range(lo, hi + 1) if n not in nums]
        if gap:
            missing[prefix] = [f"{prefix}{str(n).zfill(width)}" for n in gap]
    return missing


# ─────────────────────────────────────────────────────────────────────────
# INPUT EXCEL loading  (County | Starting Date | Ending Date)
# ─────────────────────────────────────────────────────────────────────────
def find_latest_excel(folder, label=""):
    if not os.path.isdir(folder):
        raise FileNotFoundError(f"{label or 'Folder'} not found: {folder}")
    candidates = [
        os.path.join(folder, f) for f in os.listdir(folder)
        if f.lower().endswith((".xlsx", ".xls")) and not f.startswith("~$")
    ]
    if not candidates:
        raise FileNotFoundError(f"No Excel file found in {label or folder}: {folder}")
    return max(candidates, key=os.path.getmtime)


# Filenames that are OUTPUTS this same tool writes (combined results,
# match-status/validation workbooks, the seeded baseline case-number
# list). These must never be auto-picked as the County/Date INPUT file,
# even if they land in the same folder and happen to be the newest file
# there (e.g. Output/Tool Input folder pointed at the Input folder).
_NON_INPUT_NAME_PREFIXES = (
    "result_",
    MATCH_STATUS_FILE_PREFIX.lower(),
    "lnga_casenumber_baseline",
)


def _looks_like_output_file(fname):
    lower = fname.lower()
    return any(lower.startswith(p) for p in _NON_INPUT_NAME_PREFIXES)


def _file_has_input_headers(path):
    """Peeks the header row only (cheap) and reports whether it contains
    County / Start / End -style columns, without loading the whole file."""
    try:
        ext = os.path.splitext(path)[1].lower()
        if ext == ".txt":
            with open(path, "r", newline="", encoding="utf-8-sig") as f:
                header = f.readline().lower()
            return "county" in header
        else:
            df = pd.read_excel(path, dtype=object, nrows=0)
            cols = [str(c).strip().lower().replace("_", " ") for c in df.columns]
            return any("county" in c for c in cols)
    except Exception:
        return False


def find_latest_input_file(folder, label=""):
    """Same as find_latest_excel but also considers .txt input files (the
    original weekly tab-separated format), since the Input File field now
    accepts either.

    Rejects known OUTPUT-style filenames outright (result_*, the
    CaseNumber_MatchStatus_* validation workbook, the seeded baseline
    list) and then, among what's left, only picks a file whose header row
    actually contains a County-style column — newest-first. This prevents
    silently auto-selecting a leftover output/result file that happens to
    share the Input folder and have a newer timestamp."""
    if not os.path.isdir(folder):
        raise FileNotFoundError(f"{label or 'Folder'} not found: {folder}")

    all_candidates = [
        f for f in os.listdir(folder)
        if f.lower().endswith((".xlsx", ".xls", ".txt")) and not f.startswith("~$")
    ]
    if not all_candidates:
        raise FileNotFoundError(f"No .txt or .xlsx input file found in {label or folder}: {folder}")

    candidates = [f for f in all_candidates if not _looks_like_output_file(f)]
    if not candidates:
        raise FileNotFoundError(
            f"Only output-style files (e.g. result_*, "
            f"{MATCH_STATUS_FILE_PREFIX}*) were found in {label or folder} — "
            f"no County/Starting Date/Ending Date input file to auto-detect. "
            f"Use Browse to pick the correct input file.")

    candidates.sort(key=lambda f: os.path.getmtime(os.path.join(folder, f)), reverse=True)

    for f in candidates:
        full_path = os.path.join(folder, f)
        if _file_has_input_headers(full_path):
            return full_path

    raise FileNotFoundError(
        f"Found {len(candidates)} candidate file(s) in {label or folder}, but none "
        f"has a 'County' column in its header row — nothing looks like a valid "
        f"input file. Use Browse to pick the correct one.")


def _normalize_date_cell(value):
    """County/date Excel cells may come back from Excel as real datetimes
    (if the user formatted the column as a date) or as plain DD-MM-YYYY
    text. Either way, return DD-MM-YYYY text so the rest of the script
    (unchanged from the original) keeps working exactly as before."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    if isinstance(value, (pd.Timestamp, datetime)):
        return value.strftime("%d-%m-%Y")
    text = str(value).strip()
    # Excel sometimes hands back "2026-06-20 00:00:00" for a date cell that
    # pandas didn't parse as a Timestamp for some reason.
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})", text)
    if m:
        return f"{m.group(3)}-{m.group(2)}-{m.group(1)}"
    return text


def load_input_excel(path):
    """Reads County / Starting Date / Ending Date from an Excel file.
    Header matching is case-insensitive and tolerant of minor variations
    (e.g. 'Start Date' vs 'Starting Date'). Returns the same list-of-dicts
    shape ({"county", "start_date", "end_date"}) the rest of the script
    (scraping, workbook writing) already expects."""
    df = pd.read_excel(path, dtype=object)
    df.columns = [str(c).strip() for c in df.columns]

    def _find_col(*keywords):
        for c in df.columns:
            norm = c.lower().replace("_", " ")
            if all(k in norm for k in keywords):
                return c
        return None

    county_col = _find_col("county")
    start_col = _find_col("start") or _find_col("starting")
    end_col = _find_col("end") or _find_col("ending")

    missing = [n for n, c in (("County", county_col), ("Starting Date", start_col),
                               ("Ending Date", end_col)) if c is None]
    if missing:
        raise RuntimeError(
            f"Input Excel '{path}' is missing column(s): {', '.join(missing)}.\n"
            f"Header row read as: {list(df.columns)}\n\n"
            f"This does not look like an LNGA Probate input file. The Input Excel "
            f"needs exactly these columns: County, Starting Date, Ending Date "
            f"(dates as DD-MM-YYYY). If you picked a file from a different tool's "
            f"folder by mistake, use Browse to select the correct one from the "
            f"GA_Probate Input folder instead.")

    rows = []
    for _, r in df.iterrows():
        county = str(r.get(county_col, "")).strip()
        if not county or county.lower() == "nan":
            continue
        rows.append({
            "county": county,
            "start_date": _normalize_date_cell(r.get(start_col)),
            "end_date": _normalize_date_cell(r.get(end_col)),
        })
    return rows


def load_input_txt(path):
    """Reads County / Starting Date / Ending Date from the original
    tab-separated .txt format (County\tStarting Date\tEnding Date), exactly
    as the weekly input file already comes. Returns the same list-of-dicts
    shape as load_input_excel()."""
    import csv
    rows = []
    with open(path, "r", newline="", encoding="utf-8-sig") as f:
        reader = csv.reader(f, delimiter="\t")
        next(reader, None)  # header
        for line_num, row in enumerate(reader, start=2):
            if not row or not any(c.strip() for c in row):
                continue
            if len(row) < 3:
                continue
            county, start_date, end_date = row[0].strip(), row[1].strip(), row[2].strip()
            if county:
                rows.append({"county": county, "start_date": start_date, "end_date": end_date})
    return rows


def load_input_file(path):
    """Dispatches to the .txt or .xlsx loader based on the file's
    extension, so the GUI's Input File field accepts either the original
    weekly tab-separated .txt file or an Excel workbook."""
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xls"):
        return load_input_excel(path)
    if ext == ".txt":
        return load_input_txt(path)
    raise RuntimeError(
        f"Unsupported input file type '{ext}'. Use a .txt (tab-separated: "
        f"County / Starting Date / Ending Date) or .xlsx (same columns) file.")


def to_date_input_format(ddmmyyyy_str):
    """Convert 'DD-MM-YYYY' to the site's own display format, M/D/YYYY with
    NO leading zeros (e.g. '6/20/2026')."""
    dt = datetime.strptime(ddmmyyyy_str.strip(), "%d-%m-%Y")
    return f"{dt.month}/{dt.day}/{dt.year}"


# ─────────────────────────────────────────────────────────────────────────
# Debug-HTML helper (patches saved debug_<county>.html so it can be opened
# later in a browser for troubleshooting, with CSS/JS/images resolving
# back against the live site)
# ─────────────────────────────────────────────────────────────────────────
def patch_debug_html_for_offline_viewing(html, base_href=DEBUG_HTML_BASE_HREF):
    """Inserts <base href="..."> so a saved debug_<county>.html opened later
    resolves its CSS/JS/images back against the live site."""
    base_tag = f'<base href="{base_href}">'
    if re.search(r"<head[^>]*>", html, re.IGNORECASE):
        return re.sub(r"(<head[^>]*>)", r"\1" + base_tag, html, count=1, flags=re.IGNORECASE)
    return base_tag + html


# ─────────────────────────────────────────────────────────────────────────
# Selenium driver + search-form interaction  (unchanged behavior — verified
# against the live search page; see the original script's history)
# ─────────────────────────────────────────────────────────────────────────
ID_COUNTY_DROPDOWN   = "ctl00_cpMain_ddlCounty"
ID_COUNTY_POPUP      = "ctl00_cpMain_ddlCounty_DropDown"
ID_OFFENSE_START     = "ctl00_cpMain_txtViolationStartDate_dateInput"
ID_OFFENSE_END       = "ctl00_cpMain_txtViolationEndDate_dateInput"
ID_SEARCH_BUTTON     = "ctl00_cpMain_rbSearch_input"
ID_RESULTS_GRID_CONTAINER = "ctl00_cpMain_rapGrid"

RESULT_COLUMNS = ["CASE #", "OFFENDER", "OFFENSE DATE", "COURT DATE", "DISPOSED", "COUNTY"]
_HEADER_KEY_MAP = {
    "CASE #":       "case_number",
    "OFFENDER":     "offender",
    "OFFENSE DATE": "offense_date",
    "COURT DATE":   "court_date",
    "DISPOSED":     "disposed",
    "COUNTY":       "county",
}


def build_driver(headless=True):
    options = webdriver.ChromeOptions()
    if headless:
        options.add_argument("--headless=new")
    options.add_argument("--window-size=1440,1000")
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--log-level=3")

    # In the container, CHROME_BIN / CHROMEDRIVER_PATH point at the
    # Chromium + chromedriver baked into the Docker image (see Dockerfile),
    # so Selenium doesn't try to auto-download a browser at request time.
    # Locally (no env vars set) this falls back to selenium-manager's
    # normal auto-detect behavior, unchanged from the original script.
    chrome_bin = os.environ.get("CHROME_BIN")
    if chrome_bin:
        options.binary_location = chrome_bin

    driver_path = os.environ.get("CHROMEDRIVER_PATH")
    if driver_path:
        from selenium.webdriver.chrome.service import Service
        service = Service(executable_path=driver_path)
        driver = webdriver.Chrome(options=options, service=service)
    else:
        driver = webdriver.Chrome(options=options)

    driver.set_page_load_timeout(PAGE_LOAD_TIMEOUT)
    return driver


def select_county(driver, county_name, match_timeout=6):
    wait = WebDriverWait(driver, ELEMENT_TIMEOUT)
    dropdown_div = wait.until(EC.element_to_be_clickable((By.ID, ID_COUNTY_DROPDOWN)))
    dropdown_div.click()

    popup = wait.until(EC.visibility_of_element_located((By.ID, ID_COUNTY_POPUP)))
    target = county_name.strip().lower()

    match = None
    last_seen_texts = []
    deadline = time.time() + match_timeout
    while time.time() < deadline:
        items = popup.find_elements(By.XPATH, ".//li")
        last_seen_texts = [li.text.strip() for li in items]
        match = next((li for li, txt in zip(items, last_seen_texts) if txt.lower() == target), None)
        if match is not None:
            break
        time.sleep(0.2)

    if match is None:
        raise RuntimeError(
            f"County '{county_name}' not found in RadDropDownList after "
            f"{match_timeout}s of polling. Options seen: {last_seen_texts}")
    match.click()

    try:
        WebDriverWait(driver, 6).until(EC.invisibility_of_element_located((By.ID, ID_COUNTY_POPUP)))
    except TimeoutException:
        pass
    wait_for_ajax_idle(driver)


def _set_date_via_telerik_api(driver, picker_client_id, date_str):
    script = """
        var clientId = arguments[0];
        var dateStr = arguments[1];
        var sysDefined = (typeof Sys !== 'undefined');
        var picker = (window.$find) ? $find(clientId) : null;
        var pickerFound = !!(picker && picker.set_selectedDate);
        var setSucceeded = false;
        if (pickerFound) {
            var parts = dateStr.split('/');
            var d = new Date(parseInt(parts[2], 10), parseInt(parts[0], 10) - 1, parseInt(parts[1], 10));
            picker.set_selectedDate(d);
            setSucceeded = true;
        }
        return {sys_defined: sysDefined, picker_found: pickerFound, set_succeeded: setSucceeded};
    """
    try:
        return driver.execute_script(script, picker_client_id, date_str)
    except Exception as e:
        return {"sys_defined": None, "picker_found": False, "set_succeeded": False, "js_error": str(e)}


def _type_into_date_input(driver, input_el, date_str, picker_client_id=None, field_label=""):
    driver.execute_script("arguments[0].value = '';", input_el)
    input_el.click()
    input_el.send_keys(date_str)
    input_el.send_keys(Keys.TAB)

    actual = input_el.get_attribute("value")
    if actual.strip() == date_str.strip() and picker_client_id:
        try:
            has_date = driver.execute_script(
                "var p = window.$find ? $find(arguments[0]) : null;"
                "return !!(p && p.get_selectedDate && p.get_selectedDate());",
                picker_client_id)
        except Exception:
            has_date = True
        if has_date:
            return

    m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{4})$", date_str.strip())
    if m:
        padded = f"{int(m.group(1)):02d}/{int(m.group(2)):02d}/{m.group(3)}"
        if padded != date_str.strip():
            driver.execute_script("arguments[0].value = '';", input_el)
            input_el.click()
            input_el.send_keys(padded)
            input_el.send_keys(Keys.TAB)
            actual = input_el.get_attribute("value")
            if actual.strip() in (date_str.strip(), padded):
                return

    diagnostics = None
    if picker_client_id:
        diagnostics = _set_date_via_telerik_api(driver, picker_client_id, date_str)
        if diagnostics.get("set_succeeded"):
            final_val = input_el.get_attribute("value")
            if final_val.strip() == date_str.strip():
                return

    driver.execute_script(
        "arguments[0].value = arguments[1];"
        "arguments[0].dispatchEvent(new Event('change', {bubbles:true}));"
        "arguments[0].dispatchEvent(new Event('blur', {bubbles:true}));",
        input_el, date_str)
    final_val = input_el.get_attribute("value")
    if final_val.strip() == date_str.strip():
        return

    raise RuntimeError(
        f"Could not get {field_label or 'date field'} to hold '{date_str}' after typing, "
        f"the Telerik API call, and the raw-value fallback. Final input value seen: "
        f"'{final_val}'. Telerik API diagnostics: {diagnostics}.")


def wait_for_ajax_idle(driver, timeout=8):
    try:
        WebDriverWait(driver, timeout).until(
            lambda d: d.execute_script(
                "var m = window.Sys && Sys.WebForms && Sys.WebForms.PageRequestManager;"
                "return (document.readyState === 'complete') && "
                "(!m || !m.getInstance().get_isInAsyncPostBack());"
            )
        )
    except TimeoutException:
        pass


def _find_fresh(driver, element_id, timeout=ELEMENT_TIMEOUT):
    return WebDriverWait(driver, timeout).until(
        EC.presence_of_element_located((By.ID, element_id)))


def set_offense_date_range(driver, start_ddmmyyyy, end_ddmmyyyy):
    start_str = to_date_input_format(start_ddmmyyyy)
    end_str = to_date_input_format(end_ddmmyyyy)

    start_picker_id = ID_OFFENSE_START[: -len("_dateInput")]
    end_picker_id = ID_OFFENSE_END[: -len("_dateInput")]

    start_input = _find_fresh(driver, ID_OFFENSE_START)
    _type_into_date_input(driver, start_input, start_str, picker_client_id=start_picker_id,
                           field_label="Offense Date Start")
    wait_for_ajax_idle(driver)

    end_input = _find_fresh(driver, ID_OFFENSE_END)
    _type_into_date_input(driver, end_input, end_str, picker_client_id=end_picker_id,
                           field_label="Offense Date End")
    wait_for_ajax_idle(driver)


def click_search(driver):
    wait = WebDriverWait(driver, ELEMENT_TIMEOUT)
    try:
        btn = wait.until(EC.element_to_be_clickable((By.ID, ID_SEARCH_BUTTON)))
        btn.click()
        wait_for_ajax_idle(driver, timeout=ELEMENT_TIMEOUT)
        return
    except TimeoutException:
        pass
    strategies = [
        (By.XPATH, "//input[@type='submit' and contains(translate(@value,"
                    "'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'search')]"),
        (By.XPATH, "//button[contains(translate(text(),"
                    "'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'search')]"),
    ]
    for by, sel in strategies:
        try:
            btn = wait.until(EC.element_to_be_clickable((by, sel)))
            btn.click()
            wait_for_ajax_idle(driver, timeout=ELEMENT_TIMEOUT)
            return
        except TimeoutException:
            continue
    raise RuntimeError("Could not locate the Search button (tried several selectors).")


def set_page_size(driver, grid_container_id=ID_RESULTS_GRID_CONTAINER, target_size=30, timeout=8):
    """Forces the results grid's page-size combo box (Telerik RadComboBox,
    e.g. 'PageSizeComboBox') to `target_size` (default 30) right after a
    search, so each page of results holds up to 30 rows before the
    Next-page loop in scrape_result_rows() starts clicking through pages.

    This tries several selector strategies since the exact client ID isn't
    guaranteed across the site's grid instances. If no page-size control is
    found (e.g. too few rows for one to render, or the site changed), this
    logs nothing and simply returns False — scrape_result_rows() still
    pages through all results at whatever the default page size is, so no
    data is lost, just possibly slower to page through."""
    wait = WebDriverWait(driver, timeout)
    try:
        containers = driver.find_elements(By.ID, grid_container_id)
        search_root = containers[0] if containers else driver
    except Exception:
        search_root = driver

    target_text = str(target_size)

    # Strategy 1: plain <select> page-size dropdown (some RadGrid skins
    # render the pager's page-size picker as a native <select>).
    try:
        selects = search_root.find_elements(
            By.XPATH,
            ".//select[contains(translate(@id,'PAGESIZE','pagesize'),'pagesize') or "
            "contains(translate(@class,'PAGESIZE','pagesize'),'pagesize')]"
        )
        for sel_el in selects:
            options = sel_el.find_elements(By.TAG_NAME, "option")
            for opt in options:
                if opt.text.strip() == target_text or opt.get_attribute("value") == target_text:
                    opt.click()
                    wait_for_ajax_idle(driver, timeout=timeout)
                    time.sleep(0.4)
                    return True
    except Exception:
        pass

    # Strategy 2: Telerik RadComboBox rendered as a clickable div/input with
    # an id containing "PageSize" — open its dropdown popup and click the
    # "30" list item.
    try:
        combo_inputs = search_root.find_elements(
            By.XPATH,
            ".//*[contains(@id,'PageSizeComboBox') or contains(@id,'PageSize') "
            "or contains(@id,'pageSize')]"
        )
        for combo in combo_inputs:
            try:
                combo.click()
            except Exception:
                driver.execute_script("arguments[0].click();", combo)
            time.sleep(0.3)
            popups = driver.find_elements(
                By.XPATH,
                "//div[contains(@id,'PageSizeComboBox') and contains(@id,'DropDown')]//li | "
                "//ul[contains(@id,'PageSizeComboBox')]//li"
            )
            for li in popups:
                if li.text.strip() == target_text:
                    li.click()
                    wait_for_ajax_idle(driver, timeout=timeout)
                    time.sleep(0.4)
                    return True
    except Exception:
        pass

    # Strategy 3: raw Telerik client-side API — find any RadComboBox on the
    # page whose client ID mentions PageSize and set its value via JS.
    try:
        result = driver.execute_script("""
            var target = arguments[0];
            if (!window.Sys || !Sys.Application) { return false; }
            var comps = Sys.Application.getComponents ? Sys.Application.getComponents() : [];
            for (var i = 0; i < comps.length; i++) {
                var c = comps[i];
                var id = c.get_id ? c.get_id() : "";
                if (id && id.toLowerCase().indexOf("pagesize") !== -1 && c.set_text && c.findItemByText) {
                    var item = c.findItemByText(target);
                    if (item) { item.select(); return true; }
                }
            }
            return false;
        """, target_text)
        if result:
            wait_for_ajax_idle(driver, timeout=timeout)
            time.sleep(0.4)
            return True
    except Exception:
        pass

    return False


_PAGER_TEXT_RE = re.compile(r"page\s*size", re.IGNORECASE)

# Matches the RadGrid pager's own summary text, e.g. "96 items in 4 pages",
# rendered in a <div class="rgInfoPart"> inside the same pager row that
# _is_pager_row() excludes. Used purely as a self-check AFTER paging is
# complete — comparing the number of rows actually collected against the
# total the grid itself reports, so a silently-truncated scrape (e.g. the
# Next-click loop stopping early because of a slow AJAX postback) shows up
# as a logged warning instead of quietly shipping a partial result set.
_GRID_TOTAL_RE = re.compile(r"(\d+)\s+items?\s+in\s+(\d+)\s+pages?", re.IGNORECASE)


def _get_grid_expected_total(driver, grid_container_id=ID_RESULTS_GRID_CONTAINER):
    """Reads the RadGrid's 'N items in M pages' summary (rgInfoPart) and
    returns (expected_item_count, expected_page_count), or (None, None) if
    the summary isn't present (e.g. grid too small to show a pager)."""
    try:
        containers = driver.find_elements(By.ID, grid_container_id)
        search_root = containers[0] if containers else driver
        info_els = search_root.find_elements(By.XPATH, ".//div[contains(@class,'rgInfoPart')]")
        for el in info_els:
            m = _GRID_TOTAL_RE.search(el.text.strip())
            if m:
                return int(m.group(1)), int(m.group(2))
    except Exception:
        pass
    return None, None


def _is_pager_row(tr):
    """Identify the RadGrid pager row — the row that renders the page-number
    links ('1 2 3 4 ...') and the 'Page size:' dropdown — so it can be
    excluded from body_rows.

    This row lives inside the SAME <table> as the header and data rows, but
    it is neither. It typically renders as a single <td> with a colspan
    across the whole grid width, containing the pager UI as one text blob.
    Previously nothing excluded this row, so its entire text (e.g.
    '1 2 3 4 Page size: select') was being zipped into the first data
    column and — because it was non-empty — appended to the results as a
    bogus row on every single page.
    """
    try:
        cls = (tr.get_attribute("class") or "").lower()
        if "pager" in cls:
            return True
        tds = tr.find_elements(By.TAG_NAME, "td")
        if len(tds) == 1:
            td = tds[0]
            td_cls = (td.get_attribute("class") or "").lower()
            colspan = td.get_attribute("colspan")
            txt = td.text.strip()
            if "pager" in td_cls or _PAGER_TEXT_RE.search(txt):
                return True
            if colspan and colspan.isdigit() and int(colspan) > 1 and _PAGER_TEXT_RE.search(txt):
                return True
    except Exception:
        pass
    return False


def _click_next_numeric_page(search_root, driver, current_page_num):
    """Fallback pager strategy for grids whose pager renders as PURE page
    NUMBERS with no Next/Prev arrow controls at all (e.g. a pager showing
    just "1  2  3  4   Page size: [30 v]   96 items in 4 pages" with no
    separate Next/">" element). Looks for a clickable page-number element
    whose text is exactly str(current_page_num + 1) and clicks it. Returns
    True if a click happened, False if there's no such page (i.e. we're
    already on the last page)."""
    target_text = str(current_page_num + 1)
    try:
        candidates = search_root.find_elements(
            By.XPATH,
            ".//a[normalize-space(text())='%s'] | "
            ".//span[normalize-space(text())='%s' and "
            "(contains(@onclick,'age') or contains(@class,'age'))]"
            % (target_text, target_text)
        )
        for el in candidates:
            cls = (el.get_attribute("class") or "").lower()
            # Skip anything that's actually the page-SIZE dropdown/list
            # (e.g. an option literally reading "4") rather than a page
            # NUMBER link in the pager.
            if "pagesize" in cls or el.tag_name.lower() == "option":
                continue
            try:
                el.click()
            except Exception:
                try:
                    driver.execute_script("arguments[0].click();", el)
                except Exception:
                    continue
            return True
    except Exception:
        pass
    return False


def _wait_grid_row_count_stable(driver, grid_container_id=ID_RESULTS_GRID_CONTAINER,
                                 timeout=6, stable_polls=2, poll_interval=0.3):
    """FIX (last-page cropping): after a postback swaps in a new page,
    RadGrid can attach its <tr> row elements in more than one paint pass
    instead of all at once. The page-advance check in scrape_result_rows()
    only waits for the row SET to differ from the previous page at all --
    which is satisfied the instant the FIRST new row attaches, not once
    every row for that page has attached.

    This is invisible on a full (e.g. 30-row) page because the grid is
    already mid-flight when the first row lands and the rest usually
    follow within the same tick. It becomes visible on the LAST page of a
    result set, which is typically SHORTER than a full page (e.g. page 7
    of 7 holding only 20 of the normal 30 rows): the very first of those
    20 rows attaching is enough to satisfy "different from before", so the
    scraper can end up reading the grid after only a handful of that
    page's rows have actually painted -- silently cropping the last page
    short of its real row count.

    Polls the live row count on `grid_container_id` until it holds steady
    across `stable_polls` consecutive polls (grid has finished painting),
    or gives up after `timeout` seconds (grid was just genuinely slow;
    caller proceeds with whatever is there rather than hanging forever).
    """
    try:
        containers = driver.find_elements(By.ID, grid_container_id)
        root = containers[0] if containers else driver
    except Exception:
        root = driver

    deadline = time.time() + timeout
    last_count = -1
    stable_hits = 0
    while time.time() < deadline:
        try:
            count = len(root.find_elements(
                By.XPATH,
                ".//tr[contains(@class,'rgRow') or contains(@class,'rgAltRow')]"
            ))
        except Exception:
            count = last_count
        if count == last_count:
            stable_hits += 1
            if stable_hits >= stable_polls:
                return
        else:
            stable_hits = 0
        last_count = count
        time.sleep(poll_interval)


def scrape_result_rows(driver, timeout=ELEMENT_TIMEOUT, max_pages=200, log=None):
    if log is None:
        log = lambda *a, **k: None
    results = []
    try:
        WebDriverWait(driver, timeout).until(
            lambda d: d.find_elements(By.ID, ID_RESULTS_GRID_CONTAINER)
            and d.find_element(By.ID, ID_RESULTS_GRID_CONTAINER).find_elements(By.TAG_NAME, "table")
        )
    except TimeoutException:
        return results

    # Same "don't read mid-paint" protection as page-advances (see
    # _wait_grid_row_count_stable) applied to the very first page too,
    # since a short FIRST page (e.g. a search with under 30 total results,
    # so page 1 is also the last/only page) can race exactly the same way.
    _wait_grid_row_count_stable(driver, timeout=5)

    stale_retries = 0
    MAX_STALE_RETRIES = 5
    page_num = 1

    expected_total, expected_pages = _get_grid_expected_total(driver)

    for _ in range(max_pages):
        try:
            containers = driver.find_elements(By.ID, ID_RESULTS_GRID_CONTAINER)
            search_root = containers[0] if containers else driver
            tables = search_root.find_elements(By.TAG_NAME, "table")

            target_table, header_keys, header_row_el = None, None, None
            for t in tables:
                all_trs = t.find_elements(By.XPATH, ".//tr")
                if not all_trs:
                    continue
                # Identify the header row by CONTENT (looking for "CASE #"),
                # not by assuming it's always tr[1]. Telerik RadGrid
                # sometimes renders an extra <tr> (a filter row, a hidden
                # group row, etc.) ahead of the real header in document
                # order. If we blindly treat "tr[1]" as the header and skip
                # it, that extra row gets skipped instead of the real
                # header, and the true header row then gets counted as a
                # body row — which pushes every genuine data row down by
                # one and causes the actual first data row (the one on top
                # of the visible grid) to be silently dropped, no matter
                # which county is searched.
                for tr in all_trs:
                    cells = tr.find_elements(By.XPATH, ".//th | .//td")
                    texts = [c.text.strip().upper() for c in cells]
                    if "CASE #" in texts or any("CASE" in t2 for t2 in texts):
                        target_table = t
                        header_row_el = tr
                        header_keys = [_HEADER_KEY_MAP.get(t2, t2.lower().replace(" ", "_")) for t2 in texts]
                        break
                if target_table is not None:
                    break

            if target_table is None:
                break

            # Exclude ONLY the specific <tr> element we identified as the
            # header (matched by its Selenium element id), rather than by
            # position. This guarantees every real data row is kept even
            # if the header isn't literally the first <tr> in the DOM.
            all_trs = target_table.find_elements(By.XPATH, ".//tr")
            if header_row_el is not None:
                body_rows = [tr for tr in all_trs if tr.id != header_row_el.id and not _is_pager_row(tr)]
            else:
                body_rows = [tr for tr in all_trs[1:] if not _is_pager_row(tr)]

            if not body_rows and not results:
                # The grid's header rendered but no data rows are there yet —
                # this can happen right after an AJAX postback where the
                # header binds a beat before the row data does. Give it one
                # short extra look before concluding the grid is truly empty.
                time.sleep(1.0)
                all_trs = target_table.find_elements(By.XPATH, ".//tr")
                if header_row_el is not None:
                    body_rows = [tr for tr in all_trs if tr.id != header_row_el.id and not _is_pager_row(tr)]
                else:
                    body_rows = [tr for tr in all_trs[1:] if not _is_pager_row(tr)]

            # FIX (page 7 / BUTTS-county bug): RadGrid can occasionally
            # render a SAME-PAGE duplicate -- two rows on the very page
            # currently on screen sharing the identical case number/RECID
            # -- in place of what should have been two DIFFERENT rows
            # (confirmed live: page 7 rendered case 26T0314 twice back to
            # back, and the case number that belonged in that first slot,
            # 26T0313, never appeared anywhere in the result set at all).
            # This is a stale/partial re-render, not a legitimate repeat,
            # so retry re-reading THIS SAME page (no navigation) a few
            # times before accepting it, giving the grid a chance to
            # finish painting the real distinct row.
            PAGE_DUPLICATE_RETRIES = 3
            page_rows = []
            page_dupes = []
            for attempt in range(1, PAGE_DUPLICATE_RETRIES + 1):
                page_rows = []
                for r in body_rows:
                    cells = r.find_elements(By.TAG_NAME, "td")
                    if not cells:
                        continue
                    # Defensive second check: even if a pager row slipped
                    # past _is_pager_row() (e.g. an unexpected skin/markup
                    # variant), never let a cell whose text matches the
                    # pager pattern (or is just the page-number sequence)
                    # get recorded as a case number.
                    values = [c.text.strip() for c in cells]
                    if len(values) == 1 and _PAGER_TEXT_RE.search(values[0]):
                        continue
                    values = (values + [""] * len(header_keys))[:len(header_keys)]
                    row_dict = dict(zip(header_keys, values))
                    if row_dict.get("case_number") and not _PAGER_TEXT_RE.search(row_dict["case_number"]):
                        page_rows.append(row_dict)

                page_dupes, _ = find_duplicate_case_numbers(
                    [row.get("case_number", "") for row in page_rows]
                )
                if not page_dupes:
                    break

                log(f"    [WARNING] Page {page_num}: duplicate case "
                    f"number(s) rendered within this SAME page "
                    f"({', '.join(page_dupes)}) -- likely a stale/partial "
                    f"re-render masking a different row. Re-reading this "
                    f"page (attempt {attempt}/{PAGE_DUPLICATE_RETRIES})...")
                time.sleep(1.0)
                _wait_grid_row_count_stable(driver, timeout=5)
                try:
                    all_trs = target_table.find_elements(By.XPATH, ".//tr")
                    if header_row_el is not None:
                        body_rows = [tr for tr in all_trs if tr.id != header_row_el.id and not _is_pager_row(tr)]
                    else:
                        body_rows = [tr for tr in all_trs[1:] if not _is_pager_row(tr)]
                except Exception:
                    break

            if page_dupes:
                log(f"    [WARNING] Page {page_num}: duplicate case "
                    f"number(s) persisted after {PAGE_DUPLICATE_RETRIES} "
                    f"re-read attempt(s): {', '.join(page_dupes)}. A "
                    f"different case number was likely dropped in its "
                    f"place -- please spot-check this page against the "
                    f"live site.")

            results.extend(page_rows)

            # NOTE: on this site (georgiaprobaterecords.com) the Next/Last
            # page pager controls render as <input type="button"
            # class="rgPageNext"> / rgPageLast, NOT as <a> tags. The old
            # XPath only ever looked for an <a class="rgPageNext">, so it
            # never found this site's Next button, next_candidates came
            # back empty every time, and the loop below broke out after
            # page 1 (only ~30 rows), silently discarding every remaining
            # page (e.g. pages 2-7 of a 7-page / page-size-30 result set).
            # Matching the <input> variant fixes that.
            # Snapshot of the WHOLE current page's row content (not just the
            # first row) — used below to confirm the postback really swapped
            # in a new page. Comparing every row (not only row[0]) avoids a
            # false "unchanged" read on the rare page boundary where two
            # consecutive pages happen to share the same first case number
            # text, and avoids a false "changed" read from a half-rendered
            # partial repaint (which would only touch some rows).
            prev_rows_snapshot = tuple(
                r.get_attribute("outerHTML") for r in body_rows
            ) if body_rows else tuple()

            # FIX (root cause of "N items across M pages, but only X rows
            # collected" + "removed Y duplicate case numbers" appearing
            # TOGETHER on the same county): the old code clicked Next/click-
            # numeric ONCE, waited once for the grid's first row to change,
            # and if that wait timed out it just `pass`-ed through and still
            # incremented page_num. Two different failure modes then hid
            # behind that single silent `pass`:
            #   1) The click never actually registered (stale/intercepted
            #      element, slow site, or fired mid-postback) -> the grid
            #      never advances, so the NEXT loop pass re-scrapes the
            #      exact same page it just scraped -> duplicate case
            #      numbers.
            #   2) page_num was incremented even though the grid didn't
            #      move -> if the arrow Next control is gone on a later pass
            #      and the code falls back to the pure-numeric pager, it
            #      targets str(page_num + 1), which is now off-by-one from
            #      the page actually on screen -> an entire real page of
            #      rows gets skipped, never scraped at all.
            # Together those two bugs produce exactly the symptom seen in
            # the log: fewer rows collected than the grid reports AND
            # duplicate rows removed, in the same county, at the same time.
            #
            # This is worse on a slow/loaded live site, where a single
            # ELEMENT_TIMEOUT-length wait can legitimately not be enough.
            #
            # Fix, two layers:
            #  (a) After clicking, wait on ASP.NET's own PageRequestManager
            #      flag (wait_for_ajax_idle) FIRST — this is the site
            #      telling us the postback actually finished, rather than
            #      us guessing from polling the DOM.
            #  (b) THEN confirm the grid's row content actually changed
            #      (whole-page snapshot, not just row 0). Only trust and
            #      advance page_num once both signals agree. Retry the
            #      click+wait cycle several times, with a growing timeout
            #      budget each attempt, before giving up.
            NAV_RETRIES = 5
            advanced = False
            for nav_attempt in range(1, NAV_RETRIES + 1):
                attempt_timeout = timeout + (nav_attempt - 1) * 5  # 20,25,30,35,40s

                next_candidates = search_root.find_elements(
                    By.XPATH,
                    ".//input[contains(@class,'rgPageNext') and not(contains(@class,'Disabled')) and not(@disabled)] | "
                    ".//a[contains(@class,'rgPageNext') and not(contains(@class,'Disabled'))] | "
                    ".//a[normalize-space(text())='Next'] | .//a[normalize-space(text())='>']"
                )

                # IMPORTANT (confirmed from real site HTML, "37 items in 2
                # pages" case): on the LAST page, this site's Next button
                # keeps class="rgPageNext" — no "Disabled" class, no
                # disabled attribute at all. The ONLY difference is its
                # onclick handler is prefixed with "return false;" (e.g.
                # onclick="return false;__doPostBack('...ctl10','')"),
                # which silently blocks the postback client-side. The
                # enabled Next button's onclick has no such prefix
                # (onclick="javascript:__doPostBack('...ctl10','')").
                # Without filtering these out, next_candidates never comes
                # up empty on the last page.
                next_candidates = [
                    el for el in next_candidates
                    if not (el.get_attribute("onclick") or "").strip().startswith("return false;")
                ]

                if next_candidates:
                    try:
                        next_candidates[0].click()
                    except Exception:
                        try:
                            driver.execute_script("arguments[0].click();", next_candidates[0])
                        except Exception:
                            pass
                    clicked = True
                else:
                    # No usable Next/Prev arrow control — try the pure-
                    # numeric pager fallback ("1  2  3  4   Page size: ...
                    # 96 items in 4 pages"). page_num here always reflects
                    # the last CONFIRMED page (see below), so the target
                    # text stays correct even after a retry.
                    clicked = _click_next_numeric_page(search_root, driver, page_num)

                if not clicked:
                    # No usable Next/Prev control at all — genuinely the
                    # last page.
                    break

                # (a) Wait for the ASP.NET AJAX postback itself to finish —
                # the site's own ground-truth signal, not a DOM guess.
                wait_for_ajax_idle(driver, timeout=attempt_timeout)

                # (b) Confirm the grid's row content actually changed
                # (whole-page snapshot comparison).
                try:
                    WebDriverWait(driver, attempt_timeout).until(
                        lambda d: (
                            d.find_elements(By.ID, ID_RESULTS_GRID_CONTAINER)
                            and d.find_element(By.ID, ID_RESULTS_GRID_CONTAINER)
                                .find_elements(By.XPATH, ".//tr[contains(@class,'rgRow') or contains(@class,'rgAltRow')]")
                            and (
                                not prev_rows_snapshot or
                                tuple(
                                    tr.get_attribute("outerHTML")
                                    for tr in d.find_element(By.ID, ID_RESULTS_GRID_CONTAINER)
                                        .find_elements(By.XPATH, ".//tr[contains(@class,'rgRow') or contains(@class,'rgAltRow')]")
                                ) != prev_rows_snapshot
                            )
                        )
                    )
                    advanced = True
                    break
                except TimeoutException:
                    # Click may not have registered, or fired before the
                    # previous postback fully settled, or the site is just
                    # slow. Let the AJAX queue drain and retry the click
                    # (with a longer budget next attempt) rather than
                    # silently assuming we moved on.
                    log(f"    [INFO] Page {page_num} -> {page_num + 1} not "
                        f"confirmed on attempt {nav_attempt}/{NAV_RETRIES} "
                        f"(timeout {attempt_timeout}s) — retrying.")
                    wait_for_ajax_idle(driver, timeout=5)
                    time.sleep(1.0)
                    continue

            if not advanced:
                log(f"    [WARNING] Could not confirm the grid advanced past "
                    f"page {page_num} after {NAV_RETRIES} attempt(s) — "
                    f"stopping pagination here.")
                break

            page_num += 1

            # FIX (last-page cropping): don't trust the grid is fully
            # painted just because SOME row differs from the previous page
            # (see _wait_grid_row_count_stable docstring). This matters
            # most on a short final page, where the old single fixed
            # time.sleep(0.5)/(0.6) could fire before all of that page's
            # rows had actually attached, so the NEXT loop pass would read
            # (and scrape) an incomplete row set for the last page.
            _wait_grid_row_count_stable(driver, timeout=5)
            time.sleep(0.3)

        except StaleElementReferenceException:
            stale_retries += 1
            if stale_retries > MAX_STALE_RETRIES:
                break
            wait_for_ajax_idle(driver, timeout=3)
            time.sleep(0.3)
            continue
        except Exception:
            break

    if expected_total is not None and len(results) != expected_total:
        log(f"    [WARNING] Grid reported {expected_total} item(s) across "
            f"{expected_pages} page(s), but only {len(results)} row(s) were "
            f"collected — pagination may have stopped early.")

    # De-duplicate by case number. The RadGrid can occasionally render the
    # same row twice across a paging cycle (e.g. a partial AJAX re-render
    # that overlaps with the tail of the previous page, or a slow postback
    # that gets read before AND after it settles), which otherwise inflates
    # the Output workbook with repeated case numbers. Keep the first
    # occurrence of each case number and drop the rest.
    seen_case_numbers = set()
    deduped_results = []
    duplicate_count = 0
    duplicated_case_nos = []
    for row_dict in results:
        cn = row_dict.get("case_number", "").strip()
        if cn in seen_case_numbers:
            duplicate_count += 1
            duplicated_case_nos.append(cn)
            continue
        seen_case_numbers.add(cn)
        deduped_results.append(row_dict)

    if duplicate_count:
        log(f"    [INFO] Removed {duplicate_count} duplicate case number "
            f"row(s) (same case rendered more than once during paging): "
            f"{', '.join(sorted(set(duplicated_case_nos)))}")
        # FIX (page 7 / BUTTS-county bug): a duplicated row usually means a
        # DIFFERENT, adjacent case number was silently dropped in its
        # place -- simply removing the duplicate and moving on (the old
        # behavior) hides that loss completely. Surface the best-effort
        # guess at what's actually missing so it doesn't go unnoticed.
        likely_missing = infer_likely_missing_case_numbers(
            [r.get("case_number", "") for r in results]
        )
        if likely_missing:
            for prefix, nums in likely_missing.items():
                log(f"    [WARNING] Possible MISSING case number(s) (gap "
                    f"in the {prefix}#### sequence, likely dropped when a "
                    f"duplicate rendered in its place): {', '.join(nums)}")

    return deduped_results


def _html_is_pager_row(tr):
    cls = " ".join(tr.get("class", []) or []).lower()
    if "pager" in cls:
        return True
    tds = tr.find_all("td")
    if len(tds) == 1:
        td = tds[0]
        td_cls = " ".join(td.get("class", []) or []).lower()
        colspan = td.get("colspan")
        txt = td.get_text(" ", strip=True)
        if "pager" in td_cls or _PAGER_TEXT_RE.search(txt):
            return True
        if colspan and colspan.isdigit() and int(colspan) > 1 and _PAGER_TEXT_RE.search(txt):
            return True
    return False


def _parse_offline_page(path, grid_container_id=ID_RESULTS_GRID_CONTAINER):
    """Parses one saved HTML page and returns (case_numbers, expected_total,
    expected_pages). expected_total/pages come from that page's own
    'N items in M pages' rgInfoPart summary, or (None, None) if absent."""
    with open(path, "r", encoding="utf-8", errors="replace") as f:
        html = f.read()
    soup = BeautifulSoup(html, "html.parser")
    container = soup.find(id=grid_container_id)
    search_root = container if container else soup

    expected_total, expected_pages = None, None
    for el in search_root.find_all("div", class_="rgInfoPart"):
        m = _GRID_TOTAL_RE.search(el.get_text(" ", strip=True))
        if m:
            expected_total, expected_pages = int(m.group(1)), int(m.group(2))
            break

    tables = search_root.find_all("table")
    target_table, header_row_el = None, None
    for t in tables:
        all_trs = t.find_all("tr")
        if not all_trs:
            continue
        for tr in all_trs:
            cells = tr.find_all(["th", "td"])
            texts = [c.get_text(" ", strip=True).upper() for c in cells]
            if "CASE #" in texts or any("CASE" in t2 for t2 in texts):
                target_table, header_row_el = t, tr
                break
        if target_table is not None:
            break

    if target_table is None:
        return [], expected_total, expected_pages

    all_trs = target_table.find_all("tr")
    body_rows = [tr for tr in all_trs if tr is not header_row_el and not _html_is_pager_row(tr)]

    case_numbers = []
    for r in body_rows:
        cells = r.find_all("td")
        if not cells:
            continue
        values = [c.get_text(" ", strip=True) for c in cells]
        if len(values) == 1 and _PAGER_TEXT_RE.search(values[0]):
            continue
        case_no = values[0] if values else ""
        if case_no and not _PAGER_TEXT_RE.search(case_no):
            case_numbers.append(case_no)
    return case_numbers, expected_total, expected_pages


def verify_offline_html_pages(paths):
    """Runs the offline pagination check across one or more saved HTML
    pages belonging to the SAME search result set. Returns (ok, report_text).

    ok=True  -> total rows extracted across all supplied pages matches the
                grid's own reported total, and no case number repeats
                across pages (i.e. nothing was cropped or double-read).
    ok=False -> mismatch or duplicates found; report_text explains why.
    """
    if not BS4_AVAILABLE:
        return False, ("BeautifulSoup is required for this check.\n"
                        "Install it with:  pip install beautifulsoup4")
    if not paths:
        return False, "No HTML file(s) selected."

    lines = []
    all_case_numbers = []
    grid_total, grid_pages = None, None

    for path in paths:
        try:
            rows, exp_total, exp_pages = _parse_offline_page(path)
        except Exception as e:
            lines.append(f"{os.path.basename(path)}: ERROR reading/parsing file ({e})")
            continue
        summary = f"  (grid reports: {exp_total} items in {exp_pages} pages)" if exp_total else ""
        lines.append(f"{os.path.basename(path)}: {len(rows)} row(s) extracted{summary}")
        if exp_total is not None:
            grid_total, grid_pages = exp_total, exp_pages
        all_case_numbers.extend(rows)

    total_extracted = len(all_case_numbers)
    unique_case_numbers = set(all_case_numbers)
    dupe_count = total_extracted - len(unique_case_numbers)
    dupe_case_nos, _ = find_duplicate_case_numbers(all_case_numbers)

    lines.append("-" * 50)
    lines.append(f"TOTAL extracted across {len(paths)} page(s): {total_extracted}")
    if dupe_count:
        lines.append(f"Unique case numbers: {len(unique_case_numbers)}  "
                      f"({dupe_count} DUPLICATE(S) found across pages)")
        lines.append(f"  Duplicated case number(s): {', '.join(dupe_case_nos)}")
        # A duplicate render is a strong signal that some OTHER case
        # number was silently dropped in its place (see
        # infer_likely_missing_case_numbers docstring) -- surface the
        # best-effort guess right next to the duplicate that caused it.
        likely_missing = infer_likely_missing_case_numbers(all_case_numbers)
        if likely_missing:
            lines.append("  [LIKELY MISSING] A duplicate usually means a different case "
                         "number was dropped in its place. Gap(s) found in the numbered "
                         "sequence:")
            for prefix, nums in likely_missing.items():
                lines.append(f"    {prefix}...: {', '.join(nums)}")
    else:
        lines.append(f"Unique case numbers: {len(unique_case_numbers)}")

    ok = True
    if grid_total is not None:
        lines.append(f"Grid's own reported total: {grid_total} items in {grid_pages} pages")
        if total_extracted != grid_total:
            lines.append(f"[MISMATCH] Extracted {total_extracted} but grid says {grid_total} "
                          f"-- pagination may have stopped early / a page was cropped.")
            ok = False
        else:
            lines.append("[OK] Extracted count matches the grid's reported total.")
    else:
        lines.append("(No 'N items in M pages' summary found on any supplied page -- "
                      "can't cross-check total. Include the page that shows the pager row.)")

    if dupe_count:
        lines.append(f"[WARNING] {dupe_count} duplicate case number(s) across the supplied pages "
                      f"-- could mean the same page was read twice, overlapping page boundaries, "
                      f"OR (as seen on page 7 / BUTTS county) a stale grid re-render that duplicated "
                      f"one row and silently dropped a different, adjacent case number.")
        ok = False

    lines.append("-" * 50)
    lines.append("RESULT: PASS" if ok else "RESULT: FAIL")
    return ok, "\n".join(lines)


# ─────────────────────────────────────────────────────────────────────────
# Excel styling
# ─────────────────────────────────────────────────────────────────────────
CS_HEADER_FONT  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
CS_HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
CELL_FONT    = Font(name="Arial", size=10)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN         = Side(style="thin", color="D9D9D9")
CELL_BORDER  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
NORMAL_FILL  = PatternFill("solid", fgColor="FFFFFF")
ALT_FILL     = PatternFill("solid", fgColor="F2F6FA")

STATUS_FONT = {
    "Match":                   Font(name="Arial", size=10, color="1E7B34"),
    "New (scraped only)":      Font(name="Arial", size=10, bold=True, color="9C5700"),
    "Missing (existing only)": Font(name="Arial", size=10, bold=True, color="C00000"),
}
STATUS_FILL = {
    "Match":                   PatternFill("solid", fgColor="E2F0D9"),
    "New (scraped only)":      PatternFill("solid", fgColor="FFF2CC"),
    "Missing (existing only)": PatternFill("solid", fgColor="FCE4E4"),
}


def build_output_path(output_folder, tag=""):
    """Saves a flat file directly in output_folder (e.g.
    D:\\Mohan\\GA_Probate\\Input Folder\\result_07212026_live.xlsx) —
    no dated subfolder is created."""
    stamp = datetime.now().strftime("%m%d%Y")
    ensure_folder(output_folder)
    suffix = f"_{tag}" if tag else ""
    return os.path.join(output_folder, f"result_{stamp}{suffix}.xlsx")


def build_match_status_path(tool_input_folder, tag=""):
    ensure_folder(tool_input_folder)
    stamp = datetime.now().strftime("%m%d%Y")
    suffix = f"_{tag}" if tag else ""
    return os.path.join(tool_input_folder, f"{MATCH_STATUS_FILE_PREFIX}{stamp}{suffix}.xlsx")


def write_combined_workbook(county_results, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Case Results"

    headers = ["Case #", "Offender", "Offense Date", "Court Date", "Disposed", "County"]
    row_keys = ["case_number", "offender", "offense_date", "court_date", "disposed", "county"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font, c.fill, c.alignment, c.border = CS_HEADER_FONT, CS_HEADER_FILL, CENTER_ALIGN, CELL_BORDER
    ws.freeze_panes = "A2"

    row_idx = 2
    toggle = False
    for r in county_results:
        if r["status"] != "ok" or not r["rows"]:
            continue
        toggle = not toggle
        fill = ALT_FILL if toggle else NORMAL_FILL
        for row_dict in r["rows"]:
            values = [row_dict.get(k, "") for k in row_keys]
            if not values[-1]:
                values[-1] = r["county"]
            for ci, v in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=ci, value=v)
                cell.font, cell.fill, cell.border, cell.alignment = CELL_FONT, fill, CELL_BORDER, CENTER_ALIGN
            row_idx += 1

    for ci, w in enumerate([14, 26, 14, 14, 14, 14], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    if row_idx > 2:
        ws.auto_filter.ref = f"A1:F{row_idx - 1}"

    ws2 = wb.create_sheet("Run Summary")
    summary_headers = ["County", "Starting Date", "Ending Date", "Status", "Rows Found"]
    for ci, h in enumerate(summary_headers, 1):
        c = ws2.cell(row=1, column=ci, value=h)
        c.font, c.fill, c.alignment, c.border = CS_HEADER_FONT, CS_HEADER_FILL, CENTER_ALIGN, CELL_BORDER
    for ri, r in enumerate(county_results, 2):
        values = [r["county"], r["start_date"], r["end_date"], r["status"], len(r["rows"])]
        for ci, v in enumerate(values, 1):
            cell = ws2.cell(row=ri, column=ci, value=v)
            cell.font, cell.border, cell.alignment = CELL_FONT, CELL_BORDER, CENTER_ALIGN
    for ci, w in enumerate([18, 14, 14, 12, 12], 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w

    wb.save(output_path)
    return output_path


def find_compare_file(tool_input_folder, explicit_path=None, log=print):
    if explicit_path:
        if os.path.isfile(explicit_path):
            return explicit_path
        log(f"⚠ Compare file was given but not found: {explicit_path}")
        return None

    if not os.path.isdir(tool_input_folder):
        log(f"⚠ Tool Input folder does not exist, skipping validation: {tool_input_folder}")
        return None

    candidates = [
        os.path.join(tool_input_folder, f)
        for f in os.listdir(tool_input_folder)
        if f.lower().endswith(".xlsx") and not f.startswith(MATCH_STATUS_FILE_PREFIX)
    ]
    if not candidates:
        log(f"⚠ No existing case-number workbook (.xlsx) found in {tool_input_folder} — "
            f"skipping match/mismatch validation.")
        return None
    if len(candidates) > 1:
        candidates.sort(key=os.path.getmtime, reverse=True)
        log(f"⚠ Multiple candidate case-number workbooks found — using most recent: "
            f"{os.path.basename(candidates[0])}")
    return candidates[0]


def read_existing_case_numbers(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
    case_col_idx = None
    for idx, h in enumerate(header_row):
        if h is None:
            continue
        norm = re.sub(r"[^A-Z]", "", str(h).upper())
        if norm in ("CASE", "CASENO", "CASENUMBER", "CASENUM"):
            case_col_idx = idx
            break
    if case_col_idx is None:
        raise RuntimeError(
            f"Could not find a 'Case #' column in {path}. Header row read as: {header_row}")

    numbers = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if case_col_idx < len(row) and row[case_col_idx] not in (None, ""):
            numbers.add(str(row[case_col_idx]).strip())
    return numbers


def write_match_status_workbook(existing_numbers, scraped_numbers, compare_file_path, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Match Status"

    headers = ["Case #", "Status"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font, c.fill, c.alignment, c.border = CS_HEADER_FONT, CS_HEADER_FILL, CENTER_ALIGN, CELL_BORDER
    ws.freeze_panes = "A2"

    all_numbers = sorted(existing_numbers | scraped_numbers)
    row_idx = 2
    match_count = new_count = missing_count = 0
    for case_no in all_numbers:
        in_existing = case_no in existing_numbers
        in_scraped = case_no in scraped_numbers
        if in_existing and in_scraped:
            status = "Match"
            match_count += 1
        elif in_scraped and not in_existing:
            status = "New (scraped only)"
            new_count += 1
        else:
            status = "Missing (existing only)"
            missing_count += 1

        c1 = ws.cell(row=row_idx, column=1, value=case_no)
        c2 = ws.cell(row=row_idx, column=2, value=status)
        for c in (c1, c2):
            c.font, c.border, c.alignment = STATUS_FONT[status], CELL_BORDER, CENTER_ALIGN
            c.fill = STATUS_FILL[status]
        row_idx += 1

    for ci, w in enumerate([26, 24], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    if row_idx > 2:
        ws.auto_filter.ref = f"A1:B{row_idx - 1}"

    ws2 = wb.create_sheet("Validation Summary")
    ws2["A1"], ws2["B1"] = "Compared against", compare_file_path
    ws2["A2"], ws2["B2"] = "Existing case numbers", len(existing_numbers)
    ws2["A3"], ws2["B3"] = "Scraped case numbers", len(scraped_numbers)
    ws2["A4"], ws2["B4"] = "Match", match_count
    ws2["A5"], ws2["B5"] = "New (scraped only)", new_count
    ws2["A6"], ws2["B6"] = "Missing (existing only)", missing_count
    for r in range(1, 7):
        ws2.cell(row=r, column=1).font = Font(bold=True, name="Arial", size=10)
        ws2.cell(row=r, column=2).font = CELL_FONT
    ws2.column_dimensions["A"].width = 26
    ws2.column_dimensions["B"].width = 60

    wb.save(output_path)

    report_rows = [{"Case Number": n,
                     "Status": (
                         "Match" if (n in existing_numbers and n in scraped_numbers) else
                         "New (scraped only)" if n in scraped_numbers else
                         "Missing (existing only)")}
                    for n in all_numbers]
    report_df = pd.DataFrame(report_rows) if report_rows else pd.DataFrame(
        columns=["Case Number", "Status"])

    return output_path, match_count, new_count, missing_count, report_df


# ─────────────────────────────────────────────────────────────────────────
# LIVE SEARCH pipeline (Selenium, drives the real site)
# ─────────────────────────────────────────────────────────────────────────
def run_live_pipeline(cfg, log, on_done):
    """cfg keys: input_file, output_folder, offline_folder, tool_input_folder,
    delay_seconds_minimum, delay_seconds_maximum, headless, limit"""
    try:
        if not SELENIUM_AVAILABLE:
            raise RuntimeError(
                "selenium is not installed. Install it with:  pip install selenium")

        log("=" * 62)
        log("  LNGA Probate Courts — Live Search")
        log("=" * 62)
        log(f"  Input Excel     : {cfg['input_file']}")
        log(f"  Output folder   : {cfg['output_folder']}")
        log(f"  Offline folder  : {cfg['offline_folder']}")
        log(f"  Tool Input      : {cfg['tool_input_folder']}")

        county_rows = load_input_file(cfg["input_file"])
        limit = cfg.get("limit") or 0
        if limit:
            county_rows = county_rows[:limit]
        if not county_rows:
            raise RuntimeError(f"No county rows parsed from: {cfg['input_file']}")

        log(f"  Counties to process : {len(county_rows)}")
        log(f"  Browser mode        : {'headless' if cfg.get('headless', True) else 'visible'}\n")

        ensure_folder(cfg["offline_folder"])
        driver = build_driver(headless=cfg.get("headless", True))
        county_results = []
        MAX_ATTEMPTS = 3
        delay_min = float(cfg.get("delay_seconds_minimum", 0) or 0)
        delay_max = float(cfg.get("delay_seconds_maximum", 0) or 0)

        try:
            for i, row in enumerate(county_rows, 1):
                county, start_date, end_date = row["county"], row["start_date"], row["end_date"]
                log("-" * 60)
                log(f"[{i}/{len(county_rows)}] {county}  ({start_date} -> {end_date})")

                last_error = None
                result_rows, succeeded = [], False
                for attempt in range(1, MAX_ATTEMPTS + 1):
                    try:
                        driver.get(SEARCH_URL)
                        select_county(driver, county)
                        set_offense_date_range(driver, start_date, end_date)
                        click_search(driver)
                        set_page_size(driver, target_size=30)
                        wait_for_ajax_idle(driver, timeout=8)
                        result_rows = scrape_result_rows(driver, log=log)

                        if not result_rows and attempt < MAX_ATTEMPTS:
                            # A clean 0-row scrape is ambiguous: it could be a
                            # genuinely empty result set, or the grid's AJAX
                            # postback simply hadn't finished rendering data
                            # rows yet when we read the table. Give it one
                            # more real search attempt (with an extra settle
                            # pause) before accepting 0 as the final answer,
                            # instead of trusting the first read.
                            log(f"  [WARNING] Attempt {attempt}/{MAX_ATTEMPTS}: 0 rows scraped — "
                                f"retrying in case the grid was still loading...")
                            time.sleep(1.5)
                            continue

                        succeeded = True
                        break
                    except StaleElementReferenceException as e:
                        last_error = e
                        log(f"  [WARNING] Attempt {attempt}/{MAX_ATTEMPTS}: stale element — retrying...")
                        time.sleep(1.0)
                        continue
                    except Exception as e:
                        last_error = e
                        break

                if succeeded:
                    log(f"  {len(result_rows)} row(s) found")
                    if len(result_rows) == 0:
                        dbg_path = os.path.join(cfg["offline_folder"], f"debug_{county}_zero_rows.html")
                        try:
                            patched = patch_debug_html_for_offline_viewing(driver.page_source)
                            with open(dbg_path, "w", encoding="utf-8") as f:
                                f.write(patched)
                            log(f"    (0 rows — page HTML saved to {os.path.basename(dbg_path)} for troubleshooting)")
                        except Exception:
                            pass
                    county_results.append({"county": county, "start_date": start_date, "end_date": end_date,
                                            "rows": result_rows, "status": "ok"})
                else:
                    log(f"  [ERROR] {county} failed ({type(last_error).__name__}): {last_error}")
                    dbg_path = os.path.join(cfg["offline_folder"], f"debug_{county}.html")
                    try:
                        patched = patch_debug_html_for_offline_viewing(driver.page_source)
                        with open(dbg_path, "w", encoding="utf-8") as f:
                            f.write(patched)
                        log(f"    (page HTML saved to {os.path.basename(dbg_path)} for troubleshooting)")
                    except Exception:
                        pass
                    county_results.append({"county": county, "start_date": start_date, "end_date": end_date,
                                            "rows": [], "status": f"failed: {type(last_error).__name__}: {last_error}"})

                if delay_max > 0 and i < len(county_rows):
                    import random
                    time.sleep(random.uniform(delay_min, delay_max))
        finally:
            driver.quit()

        _finish_pipeline(county_results, cfg, log, on_done, tag="live")

    except Exception:
        log(f"\n[FATAL ERROR]\n{traceback.format_exc()}")
        on_done(False, None, None, {})


def _finish_pipeline(county_results, cfg, log, on_done, tag):
    """Shared tail end for both pipelines: write the combined Output
    workbook, then the Match/New/Missing Result workbook against whatever
    existing case-number file is sitting in the Tool Input / Compare folder."""
    ok = sum(1 for r in county_results if r["status"] == "ok")
    failed = len(county_results) - ok
    total_cases = sum(len(r["rows"]) for r in county_results)
    log("\n" + "=" * 62)
    log(f"SUMMARY: {ok} OK   {failed} failed   {total_cases} row(s) total")
    log("=" * 62)

    output_path = build_output_path(cfg["output_folder"], tag=tag)
    write_combined_workbook(county_results, output_path)
    log(f"\nCombined workbook saved -> {output_path}")

    scraped_numbers = {
        row_dict["case_number"].strip()
        for r in county_results if r["status"] == "ok"
        for row_dict in r["rows"]
        if row_dict.get("case_number")
    }

    compare_file = find_compare_file(cfg["tool_input_folder"],
                                      explicit_path=cfg.get("compare_file"), log=log)
    stats = {"total": total_cases, "match": 0, "new": 0, "missing": 0,
              "report_df": pd.DataFrame(columns=["Case Number", "Status"])}

    match_path = None
    if compare_file:
        try:
            existing_numbers = read_existing_case_numbers(compare_file)
            match_path = build_match_status_path(cfg["tool_input_folder"], tag=tag)
            _, matched, new, missing, report_df = write_match_status_workbook(
                existing_numbers, scraped_numbers, compare_file, match_path)
            log(f"Result (match status) workbook saved -> {match_path}")
            log(f"  Match: {matched}   New (scraped only): {new}   Missing (existing only): {missing}")
            stats.update({"match": matched, "new": new, "missing": missing, "report_df": report_df})
        except Exception as e:
            log(f"[ERROR] Validation step failed ({type(e).__name__}): {e}")
    else:
        log("[WARNING] Skipped match/mismatch validation — no existing case-number "
            "workbook (.xlsx with a 'Case #' column) found in the Tool Input / "
            "Compare folder to check today's scrape against.")
        if scraped_numbers:
            try:
                baseline_path = os.path.join(cfg["tool_input_folder"], "LNGA_CaseNumber_Baseline.xlsx")
                wb = Workbook()
                ws = wb.active
                ws.title = "Case Numbers"
                ws.cell(row=1, column=1, value="Case #").font = Font(bold=True, name="Arial", size=10)
                for ri, num in enumerate(sorted(scraped_numbers), 2):
                    ws.cell(row=ri, column=1, value=num).font = CELL_FONT
                ws.column_dimensions["A"].width = 26
                ensure_folder(cfg["tool_input_folder"])
                wb.save(baseline_path)
                log(f"  Seeded baseline case-number workbook -> {baseline_path}")
                log(f"  ({len(scraped_numbers)} case number(s) — future runs will validate "
                    f"new scrapes against this file. Replace it any time with a more "
                    f"authoritative existing list if you have one.)")
            except Exception as e:
                log(f"  [WARNING] Could not write baseline workbook: {e}")

    log("\n" + "=" * 62)
    log("  COMPLETE")
    log("=" * 62)

    on_done(True, output_path, match_path, stats)


def normalize_key(value):
    """Normalize a value for use as a join key (case #, county names etc)."""
    if value is None:
        return ""
    s = str(value).strip()
    if s.lower() in ("nan", "none", "nat"):
        return ""
    s = re.sub(r"\s+", " ", s).upper()
    return s


def normalize_header(value):
    """Normalize a column header for fuzzy matching (lowercase, strip punctuation/spaces)."""
    s = str(value).strip().lower()
    s = re.sub(r"[^a-z0-9]+", "", s)
    return s


NO_RECORDS_MARKERS = (
    "no records to display",
    "no records found",
    "no record found",
    "no results found",
    "no data to display",
    "no matching records found",
)


def is_no_records_placeholder(value):
    """True if value is one of the site's 'zero results for this county'
    placeholder strings (e.g. 'No records to display.') rather than an
    actual Case #. These show up as a single placeholder row per empty
    county in the Input Excel and must NOT be treated as a real case
    that is 'New'/'Not Found' in the Output."""
    s = normalize_value_for_compare(value).lower().rstrip(".")
    return s in NO_RECORDS_MARKERS


def normalize_value_for_compare(value):
    """Normalize a field value before comparing HTML vs Excel (trim, collapse spaces, blank-safe)."""
    if value is None:
        return ""
    s = str(value).strip()
    if s.lower() in ("nan", "none", "nat"):
        return ""
    s = re.sub(r"\s+", " ", s)
    return s


def find_output_column(df_columns, field):
    """
    Find the Output Excel column for a given field. Tries, in order:
        1. exact "<field> (Output)"  (your standard header pattern)
        2. "<field> (Output)" using each synonym from OUTPUT_COLUMN_CANDIDATES
        3. bare synonym column names (no "(Output)" suffix), as a fallback
    Returns the actual column name, or None if nothing matches.
    """
    norm_map = {normalize_header(c): c for c in df_columns}

    exact_key = normalize_header(f"{field} (Output)")
    if exact_key in norm_map:
        return norm_map[exact_key]

    for cand in OUTPUT_COLUMN_CANDIDATES.get(field, []):
        key = normalize_header(f"{cand} (Output)")
        if key in norm_map:
            return norm_map[key]

    for cand in OUTPUT_COLUMN_CANDIDATES.get(field, [field]):
        key = normalize_header(cand)
        if key in norm_map:
            return norm_map[key]

    return None


def find_column(df_columns, candidates):
    """Generic fuzzy header match (used for the Input Excel, which has no (Output) suffix)."""
    norm_map = {normalize_header(c): c for c in df_columns}
    for cand in candidates:
        norm_cand = normalize_header(cand)
        if norm_cand in norm_map:
            return norm_map[norm_cand]
    return None


# ============================================================
# STEP 1: EXTRACT DATA FROM OFFLINE HTML FILES
# ============================================================

def get_text_by_id_suffix(soup, suffix):
    """Find the first element whose id ATTRIBUTE ENDS WITH suffix, return its text."""
    pattern = re.compile(re.escape(suffix) + r"$")
    el = soup.find(id=pattern)
    if el is None:
        return ""
    return el.get_text(separator=" ", strip=True)


COUNTY_SUFFIX_PATTERN = re.compile(r"\s*County\s+Probate\s+Court\s*$", re.IGNORECASE)


def clean_county_value(value):
    """Strip a trailing 'County Probate Court' from a County value, e.g.
    'Fulton County Probate Court' -> 'Fulton'."""
    if not value:
        return value
    return COUNTY_SUFFIX_PATTERN.sub("", value).strip()


def extract_case_from_html(filepath):
    """Parse one offline HTML case file and return a dict of extracted data."""
    with open(filepath, "r", encoding="utf-8", errors="ignore") as f:
        html = f.read()
    soup = BeautifulSoup(html, "html.parser")

    case = {"_source_file": os.path.basename(filepath)}

    for field, suffix in SINGLE_FIELD_SUFFIXES.items():
        case[field] = get_text_by_id_suffix(soup, suffix)

    # "Fulton County Probate Court" -> "Fulton"
    case["County"] = clean_county_value(case.get("County", ""))

    citations = []
    for i in range(MAX_ROW_INDEX + 1):
        row = {}
        has_value = False
        for field, prefix in ROW_FIELD_PREFIXES.items():
            val = get_text_by_id_suffix(soup, f"{prefix}{i}")
            row[field] = val
            if val:
                has_value = True
        row["_index"] = i
        if has_value:
            citations.append(row)

    case["Citations"] = citations
    return case


def load_all_html_cases(html_dir):
    """Load and extract every .html/.htm file in html_dir."""
    files = sorted(
        glob.glob(os.path.join(html_dir, "*.html"))
        + glob.glob(os.path.join(html_dir, "*.htm"))
    )
    if not files:
        log.warning("No .html/.htm files found in %s", html_dir)
    cases = []
    for fp in files:
        try:
            cases.append(extract_case_from_html(fp))
        except Exception as e:
            log.error("Failed to parse %s: %s", fp, e)
    log.info("Extracted %d case(s) from %d HTML file(s) in %s", len(cases), len(files), html_dir)
    return cases


# ============================================================
# STEP 2: LOAD OUTPUT / INPUT EXCEL (all files, combined)
# ============================================================

def get_output_excel_name_tag(excel_dir):
    """Build a filename tag from the Output Excel file name(s) in excel_dir,
    e.g. 'Fulton_Output' -> 'Fulton_Output'. If multiple output excel files
    exist, their names are joined with '_'. Falls back to 'Output' if none found."""
    files = sorted(
        glob.glob(os.path.join(excel_dir, "*.xlsx"))
        + glob.glob(os.path.join(excel_dir, "*.xls"))
    )
    files = [f for f in files if not os.path.basename(f).startswith("~$")]
    if not files:
        return "Output"
    names = [os.path.splitext(os.path.basename(f))[0] for f in files]
    tag = "_".join(names)
    tag = re.sub(r"[^A-Za-z0-9_\-]+", "_", tag).strip("_")
    return tag or "Output"


def load_excel_folder(excel_dir):
    """Read every .xlsx/.xls file in excel_dir into one combined DataFrame (all sheets, all files).
    Skips Excel's hidden lock/temp files (e.g. '~$Book1.xlsx'), which are created
    while a workbook is open elsewhere and are not readable spreadsheets."""
    files = sorted(
        glob.glob(os.path.join(excel_dir, "*.xlsx"))
        + glob.glob(os.path.join(excel_dir, "*.xls"))
    )
    files = [f for f in files if not os.path.basename(f).startswith("~$")]
    if not files:
        log.warning("No .xlsx/.xls files found in %s", excel_dir)
        return pd.DataFrame()

    frames = []
    for fp in files:
        try:
            xls = pd.ExcelFile(fp)
            for sheet in xls.sheet_names:
                df = xls.parse(sheet, dtype=str)
                df["_source_file"] = os.path.basename(fp)
                df["_source_sheet"] = sheet
                frames.append(df)
        except Exception as e:
            log.error("Failed to read %s: %s", fp, e)

    if not frames:
        return pd.DataFrame()

    combined = pd.concat(frames, ignore_index=True, sort=False)
    log.info("Loaded %d row(s) from %d excel file(s) in %s", len(combined), len(files), excel_dir)
    return combined


# ============================================================
# STEP 3: FIELD-LEVEL VALIDATION -- HTML (Tool_Input) vs OUTPUT EXCEL
# ============================================================

def resolve_output_columns(output_df):
    """Resolve the actual Output Excel column name for every field, once."""
    resolved = {}
    for field in FIELD_ORDER:
        resolved[field] = find_output_column(output_df.columns, field)
    return resolved


def group_output_rows_by_case(output_df, case_col):
    """Group Output Excel rows by normalized Case # (Output), preserving row order."""
    groups = {}
    for _, row in output_df.iterrows():
        key = normalize_key(row.get(case_col, ""))
        if not key:
            continue
        groups.setdefault(key, []).append(row)
    return groups


def pair_citations(html_citations, output_rows):
    """
    Pair each HTML citation with an Output Excel row (one row = one citation).
    1) Exact match by normalized Citation Number where both sides have one.
    2) Leftover citations/rows are paired by original order (position).
    3) Any unmatched leftovers on either side are paired with None.
    Returns a list of (html_citation_or_None, output_row_or_None) tuples.
    """
    html_remaining = list(html_citations)
    output_remaining = list(output_rows)

    # exact Citation Number match first
    pairs = []
    used_output_idx = set()
    still_unmatched_html = []
    for cit in html_remaining:
        cit_no = normalize_key(cit.get("Citation Number", ""))
        matched_idx = None
        if cit_no:
            for i, out_row in enumerate(output_remaining):
                if i in used_output_idx:
                    continue
                out_cit_no = normalize_key(out_row.get("_Citation Number_resolved", ""))
                if out_cit_no and out_cit_no == cit_no:
                    matched_idx = i
                    break
        if matched_idx is not None:
            pairs.append((cit, output_remaining[matched_idx]))
            used_output_idx.add(matched_idx)
        else:
            still_unmatched_html.append(cit)

    unused_output = [r for i, r in enumerate(output_remaining) if i not in used_output_idx]

    # positional pairing for the rest
    for i in range(max(len(still_unmatched_html), len(unused_output))):
        h = still_unmatched_html[i] if i < len(still_unmatched_html) else None
        o = unused_output[i] if i < len(unused_output) else None
        pairs.append((h, o))

    if not pairs:
        # no citations on either side -- still emit one blank citation slot
        pairs = [(None, None)]

    return pairs


def field_match_status(tool_val, output_val):
    t = normalize_value_for_compare(tool_val)
    o = normalize_value_for_compare(output_val)
    if t == "" and o == "":
        return "Match"  # nothing to compare -- treated as match
    return "Match" if t.upper() == o.upper() else "Mismatch"


def build_validation_rows(html_cases, output_df):
    """
    Build the main result rows: one row per (case, citation) pair, with
    "<Field> (Tool_Input)", "<Field> (Output)", "<Field> Match" for every
    field in FIELD_ORDER, plus "Row Status".
    """
    if output_df.empty:
        log.warning("Output Excel is empty/missing - skipping field validation.")
        return [], []

    output_cols = resolve_output_columns(output_df)
    case_col = output_cols.get("Case #")
    if case_col is None:
        raise ValueError(
            "Could not find a 'Case # (Output)' (or fallback) column in the Output Excel. "
            "Check OUTPUT_COLUMN_CANDIDATES['Case #']."
        )

    # pre-resolve the citation number column so pair_citations can use it
    cit_no_col = output_cols.get("Citation Number")

    output_groups = group_output_rows_by_case(output_df, case_col)
    # attach a normalized helper key for citation-number matching
    for key, rows in output_groups.items():
        for row in rows:
            row["_Citation Number_resolved"] = row.get(cit_no_col, "") if cit_no_col else ""

    result_rows = []
    case_summary = []

    for html_case in html_cases:
        case_key = normalize_key(html_case.get("Case #", ""))
        output_rows_for_case = output_groups.get(case_key, [])
        found_in_output = bool(output_rows_for_case)

        pairs = pair_citations(html_case.get("Citations", []), output_rows_for_case)

        row_mismatch_count = 0
        for html_cit, out_row in pairs:
            row_data = {}
            mismatched_fields = []

            for field in FIELD_ORDER:
                if field in SINGLE_FIELDS:
                    tool_val = html_case.get(field, "")
                else:
                    tool_val = html_cit.get(field, "") if html_cit else ""

                output_val = out_row.get(output_cols.get(field), "") if (out_row is not None and output_cols.get(field)) else ""

                status = field_match_status(tool_val, output_val)
                if status == "Mismatch":
                    mismatched_fields.append(field)

                row_data[f"{field} (Tool_Input)"] = normalize_value_for_compare(tool_val)
                row_data[f"{field} (Output)"] = normalize_value_for_compare(output_val)
                row_data[f"{field} Match"] = status

            if mismatched_fields:
                row_data["Row Status"] = "Mismatch (" + ", ".join(mismatched_fields) + ")"
                row_mismatch_count += 1
            else:
                row_data["Row Status"] = "Match"

            row_data["_source_html_file"] = html_case.get("_source_file", "")
            result_rows.append(row_data)

        case_summary.append({
            "Case #": html_case.get("Case #", ""),
            "County": html_case.get("County", ""),
            "Rows Compared": len(pairs),
            "Rows Matched": len(pairs) - row_mismatch_count if found_in_output else 0,
            "Rows Mismatched": row_mismatch_count if found_in_output else len(pairs),
            "Overall Status": "Match" if (found_in_output and row_mismatch_count == 0) else
                              ("Case Not Found in Output Excel" if not found_in_output else "Mismatch"),
            "Source HTML File": html_case.get("_source_file", ""),
        })

    return result_rows, case_summary


# ============================================================
# STEP 4: INPUT EXCEL vs OUTPUT EXCEL -- Case # + County VALIDATION
# ============================================================

def run_input_vs_output_validation(input_df, output_df):
    """Returns (comparison_df, not_found_count).
    not_found_count = number of "No records to display." placeholder rows in
    the Input Excel (counties the source site returned zero cases for). Those
    rows are NOT real cases, so they are excluded from comparison_df / the
    New Case count, but they ARE counted here so the GUI can surface them
    under NOT FOUND instead of silently dropping them."""
    if input_df.empty:
        log.warning("Input Excel is empty/missing - skipping input-vs-output validation.")
        return pd.DataFrame(), 0
    if output_df.empty:
        log.warning("Output Excel is empty/missing - skipping input-vs-output validation.")
        return pd.DataFrame(), 0

    in_case_col = find_column(input_df.columns, INPUT_COLUMN_CANDIDATES["Case #"])
    in_county_col = find_column(input_df.columns, INPUT_COLUMN_CANDIDATES["County"])

    output_cols = resolve_output_columns(output_df)
    out_case_col = output_cols.get("Case #")
    out_county_col = output_cols.get("County")

    if in_case_col is None or out_case_col is None:
        raise ValueError("Could not locate a 'Case #' column in the Input and/or Output Excel.")

    output_lookup = {}
    for _, row in output_df.iterrows():
        key = normalize_key(row.get(out_case_col, ""))
        if key and key not in output_lookup:
            output_lookup[key] = row  # first row per case is enough for case-level County check

    results = []
    not_found_count = 0
    for _, in_row in input_df.iterrows():
        in_case = in_row.get(in_case_col, "")
        key = normalize_key(in_case)

        if not key:
            # blank Case # in the Input Excel - nothing to compare, skip this row entirely
            continue

        if is_no_records_placeholder(in_case):
            # source site's "no cases found for this county" placeholder row -
            # not a real case, so it must not be counted as New Case. Still
            # written into the result sheet (and counted) as "Not Found" so
            # it's visible in the Excel, not just the console/log.
            not_found_count += 1
            county_for_msg = in_row.get(in_county_col, "") if in_county_col else ""
            msg = f"'No records to display.' placeholder row for county={county_for_msg} -> counted as Not Found"
            log.info(msg)
            results.append({
                "Case #": in_case,
                "Input County": county_for_msg,
                "Output County": "",
                "Case # Status": "Not Found",
                "County Status": "Not Found",
                "Source Input File": in_row.get("_source_file", ""),
                "Notes": msg,
            })
            continue

        in_county = in_row.get(in_county_col, "") if in_county_col else ""

        out_row = output_lookup.get(key)
        if out_row is None:
            results.append({
                "Case #": in_case,
                "Input County": in_county,
                "Output County": "",
                "Case # Status": "New Case",
                "County Status": "New Case",
                "Source Input File": in_row.get("_source_file", ""),
                "Notes": "",
            })
            continue

        out_county = out_row.get(out_county_col, "") if out_county_col else ""
        county_status = (
            "Match"
            if normalize_value_for_compare(in_county) == normalize_value_for_compare(out_county)
            else "Mismatch"
        )
        results.append({
            "Case #": in_case,
            "Input County": in_county,
            "Output County": out_county,
            "Case # Status": "Match",
            "County Status": county_status,
            "Source Input File": in_row.get("_source_file", ""),
            "Notes": "",
        })

    columns = ["Case #", "Input County", "Output County", "Case # Status",
               "County Status", "Source Input File", "Notes"]
    result_df = pd.DataFrame(results)
    if result_df.empty:
        result_df = pd.DataFrame(columns=columns)
    else:
        result_df = result_df[columns]
    return result_df, not_found_count


# ============================================================
# STEP 5: WRITE RESULTS
# ============================================================

GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
HEADER_FILL = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _write_df_to_sheet(ws, df, match_col_predicate=None):
    """
    match_col_predicate: optional function(col_name) -> bool, marking which
    columns should be color-coded Match(green)/Mismatch(red).
    """
    if df.empty:
        ws.append(["No records to display."])
        return

    ws.append(list(df.columns))
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    if match_col_predicate:
        status_col_idx = [i for i, c in enumerate(df.columns) if match_col_predicate(c)]
    else:
        status_col_idx = []

    for _, row in df.iterrows():
        ws.append(list(row))
        r = ws.max_row
        for idx in status_col_idx:
            cell = ws.cell(row=r, column=idx + 1)
            val = str(cell.value or "")
            if val == "Match":
                cell.fill = GREEN_FILL
            elif val.startswith("Mismatch") or val == "Not Found in Output" or "Not Found" in val or val == "New Case":
                cell.fill = RED_FILL

    for i, col in enumerate(df.columns, start=1):
        max_len = max([len(str(col))] + [len(str(v)) for v in df[col].astype(str).tolist()])
        ws.column_dimensions[get_column_letter(i)].width = min(max(10, max_len + 2), 40)

    ws.freeze_panes = "A2"


def build_validation_result_df(result_rows):
    """Build the final DataFrame with columns in the exact required order.
    Rows where every single Tool_Input AND Output value is blank (nothing
    extracted from HTML and nothing in the Output Excel) are dropped -
    they carry no information."""
    columns = []
    for field in FIELD_ORDER:
        columns += [f"{field} (Tool_Input)", f"{field} (Output)", f"{field} Match"]
    columns += ["Row Status"]

    df = pd.DataFrame(result_rows)
    for col in columns:
        if col not in df.columns:
            df[col] = ""
    if df.empty:
        return pd.DataFrame(columns=columns)

    value_cols = [f"{field} (Tool_Input)" for field in FIELD_ORDER] + [f"{field} (Output)" for field in FIELD_ORDER]
    non_empty_mask = df[value_cols].apply(lambda col: col.astype(str).str.strip() != "").any(axis=1)
    df = df[non_empty_mask]

    return df[columns]


def write_results(result_rows, case_summary, input_output_df, result_dir, output_name_tag="Output"):
    os.makedirs(result_dir, exist_ok=True)
    out_path = os.path.join(result_dir, f"{output_name_tag}_Validation_Results.xlsx")

    validation_df = build_validation_result_df(result_rows)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws1 = wb.create_sheet("Validation_Result")
    _write_df_to_sheet(ws1, validation_df, match_col_predicate=lambda c: c.endswith("Match") or c == "Row Status")

    ws3 = wb.create_sheet("Input_vs_Output")
    _write_df_to_sheet(ws3, input_output_df, match_col_predicate=lambda c: c in ("Case # Status", "County Status"))

    wb.save(out_path)
    log.info("Results written to %s", out_path)
    return out_path


# ============================================================
# CORE RUN (used by both the GUI and the CLI fallback)
# ============================================================

def run_validation(offline_dir, output_dir, input_dir, result_base_dir):
    """Run the full validation pipeline against the given folders and return
    a summary dict. Raises on unrecoverable errors (caller should catch)."""
    log.info("=== GA Probate Validation Tool starting ===")

    html_cases = load_all_html_cases(offline_dir)
    output_df = load_excel_folder(output_dir)
    input_df = load_excel_folder(input_dir)

    result_rows, case_summary = build_validation_rows(html_cases, output_df)
    input_output_df, not_found_cases = run_input_vs_output_validation(input_df, output_df)

    today_str = datetime.now().strftime("%d%m%Y")
    result_dir = os.path.join(result_base_dir, f"Result_{today_str}")

    output_name_tag = get_output_excel_name_tag(output_dir)
    out_path = write_results(result_rows, case_summary, input_output_df, result_dir, output_name_tag)

    total_cases = len(case_summary)
    matched_cases = sum(1 for c in case_summary if c["Overall Status"] == "Match")
    mismatched_cases = sum(1 for c in case_summary if c["Overall Status"] == "Mismatch")

    # "New Case" = read straight from the Input_vs_Output sheet's "Case # Status"
    # column (an Input Excel case whose Case # doesn't exist anywhere in Output yet).
    if not input_output_df.empty and "Case # Status" in input_output_df.columns:
        new_case_count = int((input_output_df["Case # Status"] == "New Case").sum())
    else:
        new_case_count = 0

    # "Not Found" = "No records to display." placeholder rows in the Input
    # Excel (counties the source site returned zero cases for) - counted by
    # run_input_vs_output_validation, not part of input_output_df itself.

    log.info(
        "Done. %d case(s) processed, %d fully matched, %d mismatched, %d new case(s), %d not found (Input_vs_Output).",
        total_cases, matched_cases, mismatched_cases, new_case_count, not_found_cases,
    )
    log.info("Open the results file: %s", out_path)

    return {
        "out_path": out_path,
        "result_dir": result_dir,
        "total_cases": total_cases,
        "matched_cases": matched_cases,
        "mismatched_cases": mismatched_cases,
        "new_case_count": new_case_count,
        "not_found_cases": not_found_cases,
    }
